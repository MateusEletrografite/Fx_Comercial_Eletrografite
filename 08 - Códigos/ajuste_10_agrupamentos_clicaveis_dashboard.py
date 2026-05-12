from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
DASHBOARD = ROOT / "dashboard.html"
LOTE_DASHBOARD = LOTE / "dashboard.html"
VERSIONED = ROOT / "dashboard_AJUSTE_10_AGRUPAMENTOS_CLICAVEIS.html"
LOTE_VERSIONED = LOTE / "dashboard_AJUSTE_10_AGRUPAMENTOS_CLICAVEIS.html"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_10.json"


CSS = """
<style id="ajuste10-groups">
.groupCard{cursor:pointer;transition:transform .16s ease,box-shadow .16s ease,border-color .16s ease}
.groupCard:hover{transform:translateY(-1px);border-color:#8bc53f;box-shadow:0 12px 24px rgba(20,74,45,.12)}
.groupCard:focus{outline:3px solid #8bc53f;outline-offset:2px}
.groupCard small:after{content:" | abrir lista";font-weight:900;color:var(--leaf)}
.modal.drawer{align-items:stretch;justify-content:flex-end;padding:0;background:rgba(5,20,12,.38)}
.modal.drawer .dialog{height:100vh;max-height:100vh;max-width:980px;width:min(980px,94vw);border-radius:18px 0 0 18px;animation:slideInRight .22s ease-out}
.modal.drawer .dialogBody{padding:18px 20px 26px}
.groupSummary{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:14px}
.groupSummary .detail strong{display:block;font-size:20px;color:var(--green)}
.groupDrawerTable{overflow:auto;border:1px solid var(--line);border-radius:12px;max-height:58vh}
.groupDrawerTable table{min-width:980px}
@keyframes slideInRight{from{transform:translateX(100%)}to{transform:translateX(0)}}
@media(max-width:760px){.groupSummary{grid-template-columns:1fr}.modal.drawer .dialog{width:100vw;border-radius:0}}
</style>
"""


NEW_RENDER_GROUP = """function renderGroupPanel(rows){const el=$('groupList'); if(!el)return; const key=$('groupBy').value; el.innerHTML=''; const map={}; rows.forEach(r=>{const name=r[key]||'(sem informação)'; if(!map[name])map[name]={count:0,value:0,latest:'',clients:new Set()}; map[name].count++; map[name].value+=(+r['Valor total da proposta']||0); if((r['Data ISO']||'')>map[name].latest)map[name].latest=r['Data ISO']||''; map[name].clients.add(r['CPF/CNPJ']||r['Cliente Nome']||'')}); Object.entries(map).map(([name,v])=>({name,...v,clients:v.clients.size})).sort((a,b)=>(b.latest||'').localeCompare(a.latest||'')||b.count-a.count).slice(0,12).forEach(g=>{const d=document.createElement('button');d.type='button';d.className='groupCard';d.innerHTML=`<strong title="${g.name}">${g.name}</strong><span>${fmt.format(g.count)} propostas | ${fmt.format(g.clients)} clientes</span><small>${brl.format(g.value)} | mais recente: ${g.latest||'-'}</small>`;d.onclick=()=>openGroup(key,g.name);el.appendChild(d)}); if(!el.children.length)el.innerHTML='<p class="muted">Nenhum agrupamento para os filtros atuais.</p>'}"""


OPEN_GROUP_FUNCTION = """function openGroup(key,name){const rows=(lastRows.length?lastRows:filtered()).filter(r=>(r[key]||'(sem informação)')===name).sort((a,b)=>(b['Data ISO']||'').localeCompare(a['Data ISO']||'')); const value=rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0); const clients=uniq(rows.map(r=>r['CPF/CNPJ']||r['Cliente Nome'])); $('modalTitle').textContent='Agrupamento: '+name; $('modalBody').innerHTML=`<div class="groupSummary"><div class="detail"><span>Propostas</span><strong>${fmt.format(rows.length)}</strong></div><div class="detail"><span>Clientes</span><strong>${fmt.format(clients.length)}</strong></div><div class="detail"><span>Valor total</span><strong>${brl.format(value)}</strong></div></div><div class="groupDrawerTable"><table><thead><tr><th>Proposta</th><th>Data</th><th>Cliente</th><th>Status</th><th>Vendedor</th><th>Setor</th><th>Kit</th><th>Valor</th></tr></thead><tbody>${rows.map(r=>`<tr><td><button class="linkBtn prop" onclick="openProposalById('${String(r['Proposta ID']||'').replace(/'/g,'\\\\\\'')}')">${r['Proposta Numero']||r['Proposta ID']||''}</button></td><td>${r['Proposta Data']||''}</td><td>${r['Cliente Nome']||''}</td><td><span class="pill">${r['Status painel']||r['Status consolidado']||''}</span></td><td>${r['Nome do Vendedor']||''}</td><td>${r.Setor||''}</td><td>${r.Kit||''}</td><td class="money">${brl.format(+r['Valor total da proposta']||0)}</td></tr>`).join('')}</tbody></table></div>`; $('modal').classList.add('open','drawer')}"""


OPEN_BY_ID_FUNCTION = """function openProposalById(id){const r=[...allRows,...removed].find(x=>String(x['Proposta ID']||'')===String(id)); if(r){$('modal').classList.remove('drawer');openProposal(r)}}"""


def update_html(path: Path) -> None:
    html = path.read_text(encoding="utf-8")
    if 'id="ajuste10-groups"' not in html:
        html = html.replace("</head>", CSS + "</head>", 1)
    html, count = re.subn(r"function renderGroupPanel\(rows\)\{.*?\nfunction renderDups", NEW_RENDER_GROUP + "\nfunction renderDups", html, count=1, flags=re.S)
    if count != 1:
        raise RuntimeError("renderGroupPanel nao encontrado.")
    if "function openGroup(key,name)" not in html:
        html = html.replace("function openProposal(r){", OPEN_GROUP_FUNCTION + "\n" + OPEN_BY_ID_FUNCTION + "\nfunction openProposal(r){", 1)
    html = html.replace("$('modal').classList.add('open')}", "$('modal').classList.remove('drawer');$('modal').classList.add('open')}")
    html = html.replace("$('closeModal').onclick=()=>$('modal').classList.remove('open');", "$('closeModal').onclick=()=>$('modal').classList.remove('open','drawer');")
    html = html.replace("if(e.target.id==='modal')$('modal').classList.remove('open')", "if(e.target.id==='modal')$('modal').classList.remove('open','drawer')")
    path.write_text(html, encoding="utf-8")


def write_listagem(generated: str) -> None:
    wb = Workbook()
    sheets = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao", "dashboard_AJUSTE_10_AGRUPAMENTOS_CLICAVEIS.html"],
            ["Gerado em", generated],
            ["Melhoria", "Agrupamentos clicaveis com abertura lateral deslizavel."],
            ["Aplicacao", "Cada card em Agrupamento final abre clientes/propostas do grupo usando os filtros atuais."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Agrupamentos clicaveis", "Clique no card do agrupamento para abrir lista de propostas/clientes daquele grupo."],
            ["Abertura deslizavel", "Detalhe abre em painel lateral, sem trocar de pagina."],
            ["Filtros atuais", "A lista respeita modo, periodo e filtros ativos no dashboard."],
            ["Proposta clicavel", "Dentro da lista, o numero da proposta abre o detalhe da proposta."],
        ],
        "Linha do tempo": [
            ["Ordem", "Data", "Acao", "Resultado"],
            [1, generated, "Pedido de agrupamentos clicaveis", "Implementado no dashboard."],
            [2, generated, "Painel lateral", "Criada abertura deslizavel para lista do agrupamento."],
            [3, generated, "Versao criada", "dashboard_AJUSTE_10_AGRUPAMENTOS_CLICAVEIS.html"],
        ],
    }
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        for row in rows:
            ws.append(row)
        style(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def style(ws, table_name: str) -> None:
    thin = Side(style="thin", color="D9EAD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="107C41")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3FAF2")
    for idx, cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value or "") for cell in cells[:100]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(map(len, sample)) + 2, 72)
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)


def main() -> None:
    generated = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    update_html(DASHBOARD)
    shutil.copy2(DASHBOARD, LOTE_DASHBOARD)
    shutil.copy2(DASHBOARD, VERSIONED)
    shutil.copy2(DASHBOARD, LOTE_VERSIONED)
    write_listagem(generated)
    summary = {
        "gerado_em": generated,
        "ajuste": "Agrupamentos clicaveis com abertura lateral deslizavel",
        "dashboard": str(DASHBOARD.resolve()),
        "dashboard_lote": str(LOTE_DASHBOARD.resolve()),
        "dashboard_versionado": str(VERSIONED.resolve()),
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
