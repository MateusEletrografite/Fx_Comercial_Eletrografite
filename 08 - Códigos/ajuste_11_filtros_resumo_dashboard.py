from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
DASHBOARD = ROOT / "dashboard.html"
LOTE_DASHBOARD = LOTE / "dashboard.html"
VERSIONED = ROOT / "dashboard_AJUSTE_11_FILTROS_RESUMO.html"
LOTE_VERSIONED = LOTE / "dashboard_AJUSTE_11_FILTROS_RESUMO.html"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_11.json"


CSS = """
<style id="ajuste11-summary-filters">
.summaryFilters{background:#fff;border:1px solid var(--line);border-radius:14px;box-shadow:var(--shadow);padding:14px;margin-bottom:14px}
.summaryHead{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:12px}
.summaryHead h2{margin:0;color:var(--green);font-size:20px}
.summaryHead select{height:38px;border:1px solid var(--line);border-radius:10px;background:#fff;color:var(--ink);padding:0 10px}
.quickRows{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.quickBlock strong{display:block;color:#315743;font-size:12px;text-transform:uppercase;margin-bottom:8px}
.quickButtons{display:flex;gap:8px;flex-wrap:wrap}
.quickButtons button{border:1px solid var(--line);background:#f8fffa;color:var(--green);border-radius:999px;padding:8px 11px;font-weight:900}
.quickButtons button.active{background:var(--leaf);border-color:var(--leaf);color:#fff}
.kpis.summaryKpis{grid-template-columns:repeat(5,1fr)}
@media(max-width:1050px){.quickRows{grid-template-columns:1fr}.kpis.summaryKpis{grid-template-columns:1fr 1fr}}
@media(max-width:680px){.summaryHead{align-items:flex-start;flex-direction:column}.kpis.summaryKpis{grid-template-columns:1fr}}
</style>
"""


SUMMARY_HTML = """<section class="summaryFilters" id="summaryFilters"><div class="summaryHead"><h2>Resumo com filtros rápidos</h2><label class="field" style="margin:0;min-width:220px"><span style="font-size:12px;font-weight:900;color:#315743;text-transform:uppercase">Mês do resumo</span><select id="monthPick"></select></label></div><div class="quickRows"><div class="quickBlock"><strong>Setores</strong><div class="quickButtons" id="sectorQuick"></div></div><div class="quickBlock"><strong>Agrupamentos</strong><div class="quickButtons" id="groupQuick"></div></div></div></section><section class="kpis summaryKpis"><div class="kpi"><span>Propostas</span><strong id="kProps">0</strong></div><div class="kpi"><span>Valor total</span><strong id="kValue">R$ 0,00</strong></div><div class="kpi"><span>Valor em aberto</span><strong id="kOpenValue">R$ 0,00</strong></div><div class="kpi"><span>Valor concluído</span><strong id="kDoneValue">R$ 0,00</strong></div><div class="kpi"><span>Clientes</span><strong id="kClients">0</strong></div></section>"""


RENDER_SUMMARY = """function renderSummaryFilters(){const sq=$('sectorQuick'),gq=$('groupQuick'),mp=$('monthPick'); if(mp&&!mp.dataset.ready){const months=uniq(allRows.map(r=>r.AnoMes)).reverse(); mp.innerHTML=months.map(m=>`<option value="${m}" ${m===DATA.current_month?'selected':''}>${m}</option>`).join(''); mp.dataset.ready='1'} if(sq){const sectors=['',...uniq(allRows.map(r=>r.Setor))];sq.innerHTML='';sectors.forEach(v=>{const b=document.createElement('button');b.type='button';b.textContent=v||'Todos';b.className=($('sector').value===v?'active':'');b.onclick=()=>{$('sector').value=v;render()};sq.appendChild(b)})} if(gq){const opts=[...$('groupBy').options].map(o=>({value:o.value,label:o.textContent.replace('Agrupar por ','')}));gq.innerHTML='';opts.forEach(o=>{const b=document.createElement('button');b.type='button';b.textContent=o.label;b.className=($('groupBy').value===o.value?'active':'');b.onclick=()=>{$('groupBy').value=o.value;render()};gq.appendChild(b)})}}"""


NEW_RENDER = """function render(){renderSummaryFilters();let rows=filtered(); lastRows=rows; const value=rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0), openValue=rows.filter(r=>r['Status painel']==='Em aberto').reduce((a,r)=>a+(+r['Valor total da proposta']||0),0), doneValue=rows.filter(r=>r['Status painel']==='Concluídas').reduce((a,r)=>a+(+r['Valor total da proposta']||0),0), clients=uniq(rows.map(r=>r['CPF/CNPJ']||r['Cliente Nome'])); $('kProps').textContent=fmt.format(rows.length); $('kValue').textContent=brl.format(value); if($('kOpenValue'))$('kOpenValue').textContent=brl.format(openValue); if($('kDoneValue'))$('kDoneValue').textContent=brl.format(doneValue); $('kClients').textContent=fmt.format(clients.length); $('count').textContent=fmt.format(rows.length); $('tableTitle').textContent=mode==='open'?'Oportunidades em aberto':'Propostas comerciais'; renderGroupPanel(rows); renderSellers(rows); bars('statusBars',group(rows,'Status painel')); bars('sectorBars',group(rows,'Setor')); bars('kitBars',group(rows,'Kit')); renderTable(rows);renderDups();renderAudit()}"""


def update_html(path: Path) -> None:
    html = path.read_text(encoding="utf-8")
    if 'id="ajuste11-summary-filters"' not in html:
        html = html.replace("</head>", CSS + "</head>", 1)
    old_kpis = '<section class="kpis"><div class="kpi"><span>Propostas</span><strong id="kProps">0</strong></div><div class="kpi"><span>Valor total</span><strong id="kValue">R$ 0,00</strong></div><div class="kpi"><span>Clientes</span><strong id="kClients">0</strong></div></section>'
    if old_kpis in html:
        html = html.replace(old_kpis, SUMMARY_HTML, 1)
    html = html.replace('<button id="modeOpen" class="active">Em aberto</button><button id="modeAll">Todas</button>', '<button id="modeOpen">Em aberto</button><button id="modeAll" class="active">Todas</button>', 1)
    html = html.replace("let mode='open';", "let mode='all';", 1)
    html = html.replace("if(p==='month')return iso.slice(0,7)===DATA.current_month;", "if(p==='month'){const mp=$('monthPick');return iso.slice(0,7)===((mp&&mp.value)||DATA.current_month)}", 1)
    if "function renderSummaryFilters()" not in html:
        html = html.replace("function currentRows()", RENDER_SUMMARY + "\nfunction currentRows()", 1)
    html, count = re.subn(r"function render\(\)\{.*?\nfunction renderSellers", NEW_RENDER + "\nfunction renderSellers", html, count=1, flags=re.S)
    if count != 1:
        raise RuntimeError("Funcao render nao encontrada.")
    html = html.replace("['q','period','from','to','sector','vendor','status','uf','city','ptype'].forEach", "['q','period','monthPick','from','to','sector','vendor','status','uf','city','ptype'].forEach", 1)
    path.write_text(html, encoding="utf-8")


def write_listagem(generated: str) -> None:
    wb = Workbook()
    sheets = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao", "dashboard_AJUSTE_11_FILTROS_RESUMO.html"],
            ["Gerado em", generated],
            ["Melhoria", "Filtros rapidos no resumo para todos os setores e agrupamentos."],
            ["Preservado", "Ordenacao por cabecalho e agrupamentos clicaveis/deslizaveis foram mantidos."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Setores no resumo", "Botoes rapidos filtram Todos, Trafego, Carteira, ESM e qualquer outro setor existente."],
            ["Agrupamentos no resumo", "Botoes rapidos alternam Agrupar por vendedor, situacao, tipo de produto, setor e kit."],
            ["Mes do resumo", "Seletor de mes controla o periodo mensal do resumo."],
            ["Valores gerenciais", "Resumo mostra valor total, valor em aberto e valor concluido dos filtros atuais."],
        ],
        "Linha do tempo": [
            ["Ordem", "Data", "Acao", "Resultado"],
            [1, generated, "Pedido de filtros no resumo", "Adicionados filtros por setor e agrupamento."],
            [2, generated, "Resumo gerencial", "Incluidos valores em aberto/concluido e seletor de mes."],
            [3, generated, "Versao criada", "dashboard_AJUSTE_11_FILTROS_RESUMO.html"],
        ],
    }
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        for row in rows:
            ws.append(row)
        style(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def style(ws, table_name: str) -> None:
    thin = Side(style="thin", color="D9EAD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="107C41")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3FAF2")
    for idx, cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value or "") for cell in cells[:100]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(map(len, sample)) + 2, 72)
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)


def main() -> None:
    generated = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    update_html(DASHBOARD)
    shutil.copy2(DASHBOARD, LOTE_DASHBOARD)
    shutil.copy2(DASHBOARD, VERSIONED)
    shutil.copy2(DASHBOARD, LOTE_VERSIONED)
    write_listagem(generated)
    summary = {
        "gerado_em": generated,
        "ajuste": "Filtros no resumo para setores e agrupamentos",
        "dashboard": str(DASHBOARD.resolve()),
        "dashboard_lote": str(LOTE_DASHBOARD.resolve()),
        "dashboard_versionado": str(VERSIONED.resolve()),
        "observacao": "Mantidos os ajustes ja concluidos: ordenacao por cabecalho e agrupamentos clicaveis.",
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
