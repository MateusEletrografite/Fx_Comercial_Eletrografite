from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
DASHBOARD = ROOT / "dashboard.html"
LOTE_DASHBOARD = LOTE / "dashboard.html"
VERSIONED = ROOT / "dashboard_AJUSTE_12_AUDITORIA_SIMPLES.html"
LOTE_VERSIONED = LOTE / "dashboard_AJUSTE_12_AUDITORIA_SIMPLES.html"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_12.json"


SIMPLE_PANEL = """<section class="expandPanel open" id="auditoriaPanel">
  <div class="closeLine">
    <h2>Auditoria ERP</h2>
    <button data-close="auditoriaPanel">Fechar</button>
  </div>
  <p class="muted">Lista de conferência disponível na aba Auditoria ERP do relatório.</p>
  <div class="auditSimpleActions">
    <a class="btn" href="relatorios_e_dashboard_base_AJUSTE_7_REGRAS.xlsx">Abrir relatório</a>
  </div>
</section>
"""


SIMPLE_CSS = """
<style id="ajuste12-auditoria-simples">
.auditSimpleActions{margin-top:12px;display:flex;gap:10px;flex-wrap:wrap}
</style>
"""


def update_html(path: Path) -> None:
    html = path.read_text(encoding="utf-8")
    html = re.sub(r"\n?<style id=\"ajuste8-dashboard\">.*?</style>\n?", "\n", html, flags=re.S)
    if 'id="ajuste12-auditoria-simples"' not in html:
        html = html.replace("</head>", SIMPLE_CSS + "</head>", 1)

    start = html.find('<section class="expandPanel open" id="auditoriaPanel">')
    if start == -1:
        raise RuntimeError("auditoriaPanel nao encontrado.")
    end = html.find('<section class="tablePanel">', start)
    if end == -1:
        raise RuntimeError("fim do auditoriaPanel nao encontrado.")
    html = html[:start] + SIMPLE_PANEL + html[end:]

    html = re.sub(r"let auditView='erp';\s*function renderAudit\(\)\{.*?\nfunction render\(\)", "function renderAudit(){}\nfunction render()", html, count=1, flags=re.S)
    html = re.sub(r"\s*document\.querySelectorAll\('button\[data-audit-view\]'\).*?renderAudit\(\)\}\);", "", html, count=1)

    blocked_terms = [
        "auditTotal",
        "auditVendors",
        "auditOther",
        "auditCristina",
        "auditTech",
        "data-audit-view",
        "auditTable",
        "decisionGrid",
        "auditKpi",
        "auditIntro",
        "validar se entra na base",
        "Outro segmento</span>",
        "Regra de decisao",
    ]
    leftovers = [term for term in blocked_terms if term in html]
    if leftovers:
        raise RuntimeError("Termos indevidos ainda presentes: " + ", ".join(leftovers))
    path.write_text(html, encoding="utf-8")


def write_listagem(generated: str) -> None:
    wb = Workbook()
    sheets = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao", "dashboard_AJUSTE_12_AUDITORIA_SIMPLES.html"],
            ["Gerado em", generated],
            ["Acao", "Removidas as melhorias equivocadas da Auditoria ERP no dashboard."],
            ["Mantido", "Filtros do resumo, ordenacao por cabecalho e agrupamentos clicaveis/deslizaveis."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Auditoria ERP no dashboard", "Nao explicar, nao criar ranking, nao criar cards e nao criar botoes internos de decisao."],
            ["Auditoria ERP", "Deixar apenas como acesso simples/lista de conferencia via relatorio."],
            ["Preservar ajustes corretos", "Manter filtros de resumo, ordenacao por cabecalho e agrupamentos clicaveis."],
        ],
        "Linha do tempo": [
            ["Ordem", "Data", "Acao", "Resultado"],
            [1, generated, "Usuario pediu retirada das melhorias de Auditoria ERP", "Bloco explicativo/de decisao removido."],
            [2, generated, "Dashboard atualizado", "Auditoria ERP voltou a ser secao simples."],
            [3, generated, "Versao criada", "dashboard_AJUSTE_12_AUDITORIA_SIMPLES.html"],
        ],
    }
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        for row in rows:
            ws.append(row)
        style(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def style(ws, table_name: str) -> None:
    thin = Side(style="thin", color="D9EAD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="107C41")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3FAF2")
    for idx, cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value or "") for cell in cells[:100]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(map(len, sample)) + 2, 72)
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)


def main() -> None:
    generated = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    update_html(DASHBOARD)
    shutil.copy2(DASHBOARD, LOTE_DASHBOARD)
    shutil.copy2(DASHBOARD, VERSIONED)
    shutil.copy2(DASHBOARD, LOTE_VERSIONED)
    write_listagem(generated)
    summary = {
        "gerado_em": generated,
        "ajuste": "Remocao das melhorias equivocadas de Auditoria ERP no dashboard",
        "dashboard": str(DASHBOARD.resolve()),
        "dashboard_lote": str(LOTE_DASHBOARD.resolve()),
        "dashboard_versionado": str(VERSIONED.resolve()),
        "preservado": [
            "filtros no resumo",
            "ordenacao por cabecalho",
            "agrupamentos clicaveis com abertura deslizavel",
        ],
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
