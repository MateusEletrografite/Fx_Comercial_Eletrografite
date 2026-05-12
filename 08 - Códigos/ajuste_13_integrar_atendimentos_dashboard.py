from __future__ import annotations

import json
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = ROOT.parent
ATENDIMENTOS = DOWNLOADS / "Atendimentos"
LOTE = ROOT / "06 - Lote final organizado"
APOIO = LOTE / "apoio"
DASHBOARD = ROOT / "dashboard.html"
LOTE_DASHBOARD = LOTE / "dashboard.html"
VERSIONED = ROOT / "dashboard_AJUSTE_13_ATENDIMENTOS.html"
LOTE_VERSIONED = LOTE / "dashboard_AJUSTE_13_ATENDIMENTOS.html"
DATA_JSON = APOIO / "dashboard_data.json"
OUT_XLSX = ROOT / "relatorio_plano_ativacao_atendimentos_AJUSTE_13.xlsx"
LOTE_XLSX = LOTE / OUT_XLSX.name
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_13.json"


CSS = """
<style id="ajuste13-atendimentos">
.attendancePanel{background:#fff;border:1px solid var(--line);border-radius:14px;box-shadow:var(--shadow);padding:16px;margin:14px 0}
.attendancePanel h2{margin:0;color:var(--green);font-size:22px}
.attendanceHead{display:flex;justify-content:space-between;gap:12px;align-items:center;margin-bottom:12px}
.attendanceFilters{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0 14px}
.attendanceFilters button{border:1px solid var(--line);background:#f8fffa;color:var(--green);border-radius:999px;padding:8px 12px;font-weight:900}
.attendanceFilters button.active{background:var(--leaf);border-color:var(--leaf);color:#fff}
.attendanceGrid{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:14px}
.attendanceKpi{border:1px solid var(--line);background:#fbfffc;border-radius:12px;padding:12px}
.attendanceKpi span{display:block;font-size:11px;text-transform:uppercase;color:var(--muted);font-weight:900}
.attendanceKpi strong{display:block;margin-top:5px;font-size:22px;color:var(--green)}
.activationColumns{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.activationBox{border:1px solid var(--line);border-radius:12px;overflow:hidden;background:#fff}
.activationBox h3{margin:0;padding:12px 14px;background:#e7f5ec;color:var(--green);font-size:17px}
.activationList{max-height:430px;overflow:auto}
.activationItem{display:grid;grid-template-columns:1fr 110px;gap:8px;border-top:1px solid #edf3ef;padding:10px 12px}
.activationItem strong{display:block;color:var(--green)}
.activationItem small{display:block;color:var(--muted);margin-top:3px}
.activationItem .tag{align-self:start;text-align:center;border-radius:999px;background:#dff3e7;color:#155c38;font-weight:900;padding:5px 8px;font-size:12px}
.attendanceHistory{margin-top:16px;border:1px solid var(--line);border-radius:12px;overflow:hidden}
.attendanceHistory h3{margin:0;padding:12px 14px;background:#f8fffa;color:var(--green);font-size:17px}
.attendanceHistory table{min-width:950px}
@media(max-width:1050px){.attendanceGrid{grid-template-columns:1fr 1fr}.activationColumns{grid-template-columns:1fr}}
@media(max-width:680px){.attendanceGrid{grid-template-columns:1fr}.activationItem{grid-template-columns:1fr}}
</style>
"""


HTML_PANEL = """<section class="attendancePanel" id="attendancePanel">
  <div class="attendanceHead">
    <div>
      <h2>Atendimentos e ativação de clientes</h2>
      <p class="muted">Base: atendimentos consolidados do workspace Atendimentos.</p>
    </div>
  </div>
  <div class="attendanceFilters" id="attendanceSectorBtns"></div>
  <div class="attendanceGrid">
    <div class="attendanceKpi"><span>Clientes na base</span><strong id="attClients">0</strong></div>
    <div class="attendanceKpi"><span>Com atendimento</span><strong id="attWith">0</strong></div>
    <div class="attendanceKpi"><span>Atendidos 1 vez</span><strong id="attOnce">0</strong></div>
    <div class="attendanceKpi"><span>Sem contato 30+ dias</span><strong id="attStale">0</strong></div>
    <div class="attendanceKpi"><span>Com proposta aberta</span><strong id="attOpen">0</strong></div>
  </div>
  <div class="activationColumns">
    <div class="activationBox"><h3>Atender com proposta em aberto</h3><div class="activationList" id="activationOpen"></div></div>
    <div class="activationBox"><h3>Ativar sem proposta em aberto</h3><div class="activationList" id="activationNoOpen"></div></div>
  </div>
  <div class="attendanceHistory">
    <h3>Últimos contatos do setor</h3>
    <div class="tableWrap"><table><thead><tr><th>Cliente</th><th>Setor</th><th>Atendimentos</th><th>Último contato</th><th>Dias</th><th>Vendedor sugerido</th><th>Proposta aberta</th></tr></thead><tbody id="attendanceRows"></tbody></table></div>
  </div>
</section>
"""


JS = r"""
let attendanceSector='Todos';
function renderAttendancePanel(){
  const data=DATA.atendimentos||{}, clients=data.clientes||[], sectors=['Todos','Tráfego','Carteira','ESM'];
  const btns=$('attendanceSectorBtns'); if(!btns)return;
  btns.innerHTML=''; sectors.forEach(s=>{const b=document.createElement('button');b.type='button';b.textContent=s;b.className=attendanceSector===s?'active':'';b.onclick=()=>{attendanceSector=s;renderAttendancePanel()};btns.appendChild(b)});
  const rows=clients.filter(c=>attendanceSector==='Todos'||c.Setor===attendanceSector);
  const withAtt=rows.filter(c=>c.Qtd_atendimentos>0), once=rows.filter(c=>c.Qtd_atendimentos===1), stale=rows.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30), open=rows.filter(c=>c.Tem_proposta_aberta==='Sim');
  $('attClients').textContent=fmt.format(rows.length); $('attWith').textContent=fmt.format(withAtt.length); $('attOnce').textContent=fmt.format(once.length); $('attStale').textContent=fmt.format(stale.length); $('attOpen').textContent=fmt.format(open.length);
  function item(c){return `<div class="activationItem"><div><strong>${c.Cliente||'(sem nome)'}</strong><small>${c.Motivo_ativacao}</small><small>Último contato: ${c.Ultimo_contato||'sem registro'} | Atend.: ${fmt.format(c.Qtd_atendimentos||0)} | Vendedor: ${c.Vendedor_sugerido||''}</small></div><div class="tag">${c.Prioridade}</div></div>`}
  const openList=(data.plano||[]).filter(c=>(attendanceSector==='Todos'||c.Setor===attendanceSector)&&c.Tem_proposta_aberta==='Sim').slice(0,80);
  const noOpenList=(data.plano||[]).filter(c=>(attendanceSector==='Todos'||c.Setor===attendanceSector)&&c.Tem_proposta_aberta!=='Sim').slice(0,80);
  $('activationOpen').innerHTML=openList.length?openList.map(item).join(''):'<p class="muted" style="padding:12px">Sem clientes nesta condição.</p>';
  $('activationNoOpen').innerHTML=noOpenList.length?noOpenList.map(item).join(''):'<p class="muted" style="padding:12px">Sem clientes nesta condição.</p>';
  $('attendanceRows').innerHTML=rows.slice().sort((a,b)=>(+b.Prioridade_ordem||0)-(+a.Prioridade_ordem||0)).slice(0,220).map(c=>`<tr><td>${c.Cliente}</td><td>${c.Setor}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${c.Ultimo_contato||''}</td><td>${c.Dias_sem_contato}</td><td>${c.Vendedor_sugerido||''}</td><td>${c.Tem_proposta_aberta}</td></tr>`).join('');
}
"""


def norm_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().upper()
    return text


def digits(value: object) -> str:
    return re.sub(r"\D+", "", "" if value is None else str(value))


def cpf_cnpj_key(value: object) -> str:
    d = digits(value)
    return d if len(d) >= 11 else ""


def phone_keys(value: object) -> set[str]:
    out = set()
    for part in re.split(r"\||;|,|\s+", "" if value is None else str(value)):
        d = digits(part)
        if len(d) >= 8:
            out.add(d[-8:])
            if len(d) >= 10:
                out.add(d[-10:])
            if len(d) >= 11:
                out.add(d[-11:])
    return out


def parse_date(value: object) -> pd.Timestamp | None:
    if value is None or str(value).strip() == "":
        return None
    dt = pd.to_datetime(value, errors="coerce", dayfirst=False)
    if pd.isna(dt):
        dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    return None if pd.isna(dt) else dt


def br_date(value: pd.Timestamp | None) -> str:
    return "" if value is None or pd.isna(value) else value.strftime("%d/%m/%Y")


def load_atendimentos() -> pd.DataFrame:
    p = ATENDIMENTOS / "relatorio_atendimentos_consolidado_classificado_empresa_fornecedor_20260511-085458.xlsx"
    return pd.read_excel(p, sheet_name="Atendimentos consolidado", dtype=str, keep_default_na=False)


def build_attendance_data(dashboard_data: dict) -> tuple[dict, pd.DataFrame]:
    rows = pd.DataFrame(dashboard_data["oportunidades"])
    # Regra operacional atual informada pelo usuario.
    rows.loc[rows["Nome do Vendedor"].astype(str).str.upper().str.contains("ELENICE", na=False), "Setor"] = "Tráfego"
    dashboard_data["oportunidades"] = rows.to_dict("records")

    att = load_atendimentos()
    att["_qtd"] = pd.to_numeric(att["Quantidade de atendimentos"], errors="coerce").fillna(1).astype(int)
    att["_last"] = att.apply(
        lambda r: parse_date(r.get("Data do último contato")) or parse_date(r.get("Data do contato")) or parse_date(r.get("Data do ultimo encerramento")),
        axis=1,
    )
    att["_cnpj_key"] = att["CNPJ na plataforma de atendimento"].map(cpf_cnpj_key)
    att["_name_key"] = att["Nome na plataforma de atendimento"].map(norm_text)
    att["_phone_keys"] = att["Telefone completo em número"].map(phone_keys)

    by_cnpj: dict[str, list[int]] = defaultdict(list)
    by_phone: dict[str, list[int]] = defaultdict(list)
    by_name: dict[str, list[int]] = defaultdict(list)
    for idx, r in att.iterrows():
        if r["_cnpj_key"]:
            by_cnpj[r["_cnpj_key"]].append(idx)
        for ph in r["_phone_keys"]:
            by_phone[ph].append(idx)
        if r["_name_key"]:
            by_name[r["_name_key"]].append(idx)

    client_groups = []
    for key, grp in rows.groupby("Cliente chave", dropna=False):
        first = grp.iloc[0]
        cnpj = cpf_cnpj_key(first.get("CPF/CNPJ", ""))
        phones = set()
        for col in ["Cliente Fone", "Cliente Celular"]:
            for value in grp.get(col, []):
                phones |= phone_keys(value)
        name_key = norm_text(first.get("Cliente Nome", ""))
        matched = []
        rule = ""
        if cnpj and cnpj in by_cnpj:
            matched = by_cnpj[cnpj]
            rule = "CNPJ/CPF"
        if not matched:
            for ph in phones:
                if ph in by_phone:
                    matched.extend(by_phone[ph])
            if matched:
                rule = "Telefone"
        if not matched and name_key in by_name:
            matched = by_name[name_key]
            rule = "Nome"
        matched = sorted(set(matched))
        att_rows = att.loc[matched] if matched else pd.DataFrame()
        qtd = int(att_rows["_qtd"].sum()) if len(att_rows) else 0
        last_dates = [d for d in att_rows["_last"].tolist()] if len(att_rows) else []
        last = max([d for d in last_dates if d is not None], default=None)
        days = "" if last is None else int((pd.Timestamp("2026-05-11") - last.normalize()).days)
        open_grp = grp[grp["Status painel"].eq("Em aberto")]
        done_grp = grp[grp["Status painel"].eq("Concluídas")]
        sector_counts = grp["Setor"].value_counts()
        setor = sector_counts.index[0] if len(sector_counts) else ""
        seller_counts = grp["Nome do Vendedor"].value_counts()
        vendedor = seller_counts.index[0] if len(seller_counts) else ""
        if "ELENICE" in str(vendedor).upper():
            setor = "Tráfego"
        value_open = float(pd.to_numeric(open_grp["Valor total da proposta"], errors="coerce").fillna(0).sum()) if len(open_grp) else 0.0
        if len(open_grp) and (qtd == 0 or days == "" or int(days) >= 15):
            priority, order, reason = "Alta", 3, "Proposta em aberto sem contato recente."
        elif len(open_grp):
            priority, order, reason = "Média", 2, "Proposta em aberto com atendimento recente."
        elif qtd == 1:
            priority, order, reason = "Média", 2, "Cliente atendido somente uma vez; ativar novo contato."
        elif days == "" or int(days) >= 30:
            priority, order, reason = "Média", 2, "Cliente sem contato recente; ativar carteira."
        else:
            priority, order, reason = "Baixa", 1, "Cliente com contato recente."
        protocol = " | ".join(att_rows["Protocolo de atendimentos"].astype(str).head(5).tolist()) if len(att_rows) else ""
        client_groups.append(
            {
                "Cliente": first.get("Cliente Nome", ""),
                "Cliente_chave": key,
                "CPF_CNPJ": first.get("CPF/CNPJ", ""),
                "Setor": setor,
                "Vendedor_sugerido": vendedor,
                "Qtd_atendimentos": qtd,
                "Ultimo_contato": br_date(last),
                "Dias_sem_contato": days,
                "Atendido_uma_vez": "Sim" if qtd == 1 else "Não",
                "Tem_proposta_aberta": "Sim" if len(open_grp) else "Não",
                "Qtd_propostas_abertas": int(len(open_grp)),
                "Valor_aberto": round(value_open, 2),
                "Qtd_propostas_concluidas": int(len(done_grp)),
                "Regra_casamento": rule,
                "Protocolos": protocol,
                "Prioridade": priority,
                "Prioridade_ordem": order,
                "Motivo_ativacao": reason,
            }
        )

    clients = pd.DataFrame(client_groups)
    clients = clients.sort_values(["Prioridade_ordem", "Valor_aberto", "Dias_sem_contato"], ascending=[False, False, False], kind="stable")
    plan = clients[clients["Prioridade"].isin(["Alta", "Média"])].copy()

    by_sector = []
    for setor, grp in clients.groupby("Setor", dropna=False):
        by_sector.append(
            {
                "Setor": setor or "(sem setor)",
                "Clientes": int(len(grp)),
                "Com_atendimento": int((grp["Qtd_atendimentos"] > 0).sum()),
                "Atendidos_uma_vez": int((grp["Qtd_atendimentos"] == 1).sum()),
                "Sem_contato_30_dias": int(((grp["Dias_sem_contato"] == "") | (pd.to_numeric(grp["Dias_sem_contato"], errors="coerce").fillna(999) >= 30)).sum()),
                "Com_proposta_aberta": int((grp["Tem_proposta_aberta"] == "Sim").sum()),
                "Sem_proposta_aberta": int((grp["Tem_proposta_aberta"] != "Sim").sum()),
                "Valor_aberto": round(float(grp["Valor_aberto"].sum()), 2),
            }
        )

    attendance_data = {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "fonte": str((ATENDIMENTOS / "relatorio_atendimentos_consolidado_classificado_empresa_fornecedor_20260511-085458.xlsx").resolve()),
        "clientes": clients.head(2500).to_dict("records"),
        "plano": plan.head(1200).to_dict("records"),
        "por_setor": by_sector,
        "resumo": {
            "clientes_base": int(len(clients)),
            "clientes_com_atendimento": int((clients["Qtd_atendimentos"] > 0).sum()),
            "clientes_atendidos_uma_vez": int((clients["Qtd_atendimentos"] == 1).sum()),
            "clientes_sem_contato_30_dias": int(((clients["Dias_sem_contato"] == "") | (pd.to_numeric(clients["Dias_sem_contato"], errors="coerce").fillna(999) >= 30)).sum()),
            "clientes_com_proposta_aberta": int((clients["Tem_proposta_aberta"] == "Sim").sum()),
        },
    }
    return attendance_data, clients


def update_dashboard_html(data: dict) -> None:
    html = DASHBOARD.read_text(encoding="utf-8")
    if 'id="ajuste13-atendimentos"' not in html:
        html = html.replace("</head>", CSS + "</head>", 1)
    if 'id="attendancePanel"' not in html:
        marker = '<section class="groupPanel" id="groupPanel">'
        html = html.replace(marker, HTML_PANEL + marker, 1)
    if "function renderAttendancePanel()" not in html:
        html = html.replace("function renderSummaryFilters()", JS + "\nfunction renderSummaryFilters()", 1)
    html = html.replace("renderTable(rows);renderDups();renderAudit()", "renderTable(rows);renderDups();renderAudit();renderAttendancePanel()", 1)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    html = re.sub(r"const DATA=\{.*?\};\nconst fmt=", f"const DATA={payload};\nconst fmt=", html, count=1, flags=re.S)
    DASHBOARD.write_text(html, encoding="utf-8")
    shutil.copy2(DASHBOARD, LOTE_DASHBOARD)
    shutil.copy2(DASHBOARD, VERSIONED)
    shutil.copy2(DASHBOARD, LOTE_VERSIONED)


def write_xlsx(clients: pd.DataFrame, attendance_data: dict) -> None:
    wb = Workbook()
    sheets = {
        "Resumo atendimento": pd.DataFrame([attendance_data["resumo"]]),
        "Resumo por setor": pd.DataFrame(attendance_data["por_setor"]),
        "Plano ativacao": clients[clients["Prioridade"].isin(["Alta", "Média"])].copy(),
        "Clientes base atendimento": clients.copy(),
    }
    first = True
    for name, df in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        ws.append(list(df.columns))
        for row in df.fillna("").itertuples(index=False, name=None):
            ws.append(list(row))
        style(ws, f"Tabela{re.sub(r'[^A-Za-z0-9]', '', name)[:20]}")
    wb.save(OUT_XLSX)
    shutil.copy2(OUT_XLSX, LOTE_XLSX)


def write_listagem(generated: str) -> None:
    wb = Workbook()
    sheets = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao", "dashboard_AJUSTE_13_ATENDIMENTOS.html"],
            ["Gerado em", generated],
            ["Melhoria", "Atendimentos consolidados integrados ao dashboard."],
            ["Regra Elenice", "Elenice classificada como Trafego por enquanto."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Casamento atendimento", "Prioridade: CNPJ/CPF, depois telefone, depois nome normalizado."],
            ["Ultimo contato", "Usa Data do ultimo contato; se vazia, Data do contato ou encerramento."],
            ["Plano com proposta aberta", "Alta prioridade quando proposta aberta e sem contato recente."],
            ["Plano sem proposta aberta", "Sugere ativar clientes atendidos uma vez ou sem contato ha 30+ dias."],
            ["Divisao por setor", "Plano separado por Trafego, Carteira e ESM, respeitando vendedores do setor atual."],
        ],
        "Linha do tempo": [
            ["Ordem", "Data", "Acao", "Resultado"],
            [1, generated, "Leitura de Atendimentos consolidado", "Base mais recente do workspace Atendimentos usada."],
            [2, generated, "Casamento com clientes da base", "Clientes de propostas ganharam resumo de atendimento."],
            [3, generated, "Dashboard atualizado", "Painel de atendimentos e ativacao incluído."],
        ],
    }
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        for row in rows:
            ws.append(row)
        style(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def style(ws, table_name: str) -> None:
    thin = Side(style="thin", color="D9EAD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="107C41")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3FAF2")
    for idx, cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value or "") for cell in cells[:150]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(map(len, sample)) + 2, 68)
    if ws.max_row > 1 and ws.max_column > 1:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws.add_table(table)


def main() -> None:
    data = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    attendance_data, clients = build_attendance_data(data)
    data["atendimentos"] = attendance_data
    DATA_JSON.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    update_dashboard_html(data)
    write_xlsx(clients, attendance_data)
    generated = attendance_data["gerado_em"]
    write_listagem(generated)
    summary = {
        "gerado_em": generated,
        "dashboard": str(DASHBOARD.resolve()),
        "dashboard_versionado": str(VERSIONED.resolve()),
        "relatorio_xlsx": str(OUT_XLSX.resolve()),
        "clientes_base": attendance_data["resumo"]["clientes_base"],
        "clientes_com_atendimento": attendance_data["resumo"]["clientes_com_atendimento"],
        "clientes_atendidos_uma_vez": attendance_data["resumo"]["clientes_atendidos_uma_vez"],
        "clientes_sem_contato_30_dias": attendance_data["resumo"]["clientes_sem_contato_30_dias"],
        "clientes_com_proposta_aberta": attendance_data["resumo"]["clientes_com_proposta_aberta"],
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
