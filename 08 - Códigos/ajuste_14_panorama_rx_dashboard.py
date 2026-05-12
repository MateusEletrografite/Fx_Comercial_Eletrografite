from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
SOURCE = ROOT / "dashboard_AJUSTE_13_ATENDIMENTOS.html"
OUT = ROOT / "dashboard_AJUSTE_14_PANORAMA_RX.html"
ROOT_DASH = ROOT / "dashboard.html"
LOTE_DASH = LOTE / "dashboard.html"
LOTE_OUT = LOTE / OUT.name
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_14.json"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


CSS = r"""
<style id="ajuste14-panorama-rx">
.actions:empty{display:none}
.dashboardFooter{max-width:1420px;margin:20px auto 0;padding:18px 28px 26px}
.dashboardFooter .footerBox{background:#fff;border:1px solid var(--line);border-radius:14px;box-shadow:var(--shadow);padding:16px;display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap}
.dashboardFooter h2{margin:0;color:var(--green);font-size:20px}
.footerButtons{display:flex;gap:10px;flex-wrap:wrap}
.footerButtons button,.footerButtons a{border:1px solid var(--line);background:#fff;color:var(--green);border-radius:999px;padding:10px 14px;font-weight:900;text-decoration:none;box-shadow:none}
.footerButtons button.secondary{background:#f8fffa}
.moreFilters{border:1px solid var(--line);border-radius:12px;background:#fbfffc;margin-top:8px;padding:10px}
.moreFilters summary{cursor:pointer;font-weight:900;color:var(--green)}
.moreFilters .field,.moreFilters details{margin-top:10px}
.summaryFilters{position:relative}
.quickButtons button{min-height:40px}
.summaryKpis.openMode{grid-template-columns:repeat(3,1fr)}
.summaryKpis.openMode .hideOpenMode{display:none}
.statusClients{display:block;color:var(--muted);font-size:12px;font-weight:700;margin-top:2px}
.seller{grid-template-columns:56px 1fr 180px}
.sellerStats{display:flex;gap:8px;flex-wrap:wrap;margin-top:7px}
.sellerChip{border:1px solid var(--line);border-radius:999px;background:#f8fffa;color:#315743;padding:3px 8px;font-size:12px;font-weight:800}
.sellerDetailPanel{display:none;margin-top:14px;border:1px solid var(--line);border-radius:14px;background:#fbfffc;overflow:hidden}
.sellerDetailPanel.open{display:block;animation:slideDown .22s ease-out}
.sellerDetailHead{display:flex;justify-content:space-between;gap:12px;align-items:flex-start;padding:16px;border-bottom:1px solid var(--line);background:#fff}
.sellerDetailHead h3{margin:0;color:var(--green);font-size:24px}
.sellerDetailBody{padding:16px}
.sellerPanorama{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:12px 0}
.sellerPanorama .detail strong{display:block;font-size:20px;color:var(--green);margin-top:4px}
.actionReport{display:none;margin-top:14px;border:1px solid var(--line);border-radius:12px;overflow:hidden;background:#fff}
.actionReport.open{display:block;animation:slideDown .22s ease-out}
.actionReport h3{margin:0;padding:12px 14px;background:#e7f5ec;color:var(--green)}
.actionReport .tableWrap{max-height:420px}
.tableCompact{min-width:980px}
.tableCompact th{cursor:pointer}
@keyframes slideDown{from{opacity:0;transform:translateY(-10px)}to{opacity:1;transform:translateY(0)}}
@media(max-width:1050px){.sellerPanorama{grid-template-columns:1fr 1fr}.seller{grid-template-columns:50px 1fr}.seller .miniVal{grid-column:2;text-align:left}}
@media(max-width:680px){.dashboardFooter{padding:14px}.dashboardFooter .footerBox{align-items:flex-start}.sellerPanorama{grid-template-columns:1fr}.summaryKpis.openMode{grid-template-columns:1fr}.sellerDetailHead{flex-direction:column}}
</style>
"""


JS = r"""
<script id="ajuste14-panorama-rx">
(function(){
  const TODAY_ISO='2026-05-11';
  const TODAY_BR='11/05/2026';
  const safeText=v=>String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
  const numberFromText=t=>{
    const s=String(t??'').replace(/\s/g,'');
    if(/R\$|,\d{2}$/.test(s)) return Number(s.replace(/[^\d,-]/g,'').replace(/\./g,'').replace(',','.'))||0;
    const n=Number(s.replace(/[^\d.-]/g,''));
    return Number.isFinite(n)?n:NaN;
  };
  const dateDays=d=>{
    if(!d) return '';
    const parts=String(d).split('/');
    const iso=parts.length===3?`${parts[2]}-${parts[1]}-${parts[0]}`:String(d).slice(0,10);
    const diff=(new Date(TODAY_ISO+'T00:00:00')-new Date(iso+'T00:00:00'))/86400000;
    return Number.isFinite(diff)?Math.floor(diff):'';
  };
  const attClients=()=>((DATA.atendimentos||{}).clientes||[]);
  const attByKey=()=>{const m=new Map();attClients().forEach(c=>m.set(String(c.Cliente_chave||''),c));return m};
  const attMap=attByKey();
  const effectiveSector=()=>($('sector')&&$('sector').value)||attendanceSector||'Todos';
  const uniqueClients=rows=>uniq(rows.map(r=>r['CPF/CNPJ']||r['Cliente Nome']));
  const rowsForSeller=name=>{
    const base=(lastRows&&lastRows.length)?lastRows:filtered();
    return base.filter(r=>r['Nome do Vendedor']===name);
  };
  const sellerAttendanceStats=rows=>{
    const keys=[...new Set(rows.map(r=>String(r['Cliente chave']||'')).filter(Boolean))];
    const clients=keys.map(k=>attMap.get(k)).filter(Boolean);
    const openKeys=new Set(rows.filter(r=>r['Status painel']==='Em aberto').map(r=>String(r['Cliente chave']||'')));
    const openClients=[...openKeys].map(k=>attMap.get(k)).filter(Boolean);
    const bucket=list=>({
      d15:list.filter(c=>dateDays(c.Ultimo_contato)!==''&&dateDays(c.Ultimo_contato)<=15).length,
      d30:list.filter(c=>dateDays(c.Ultimo_contato)!==''&&dateDays(c.Ultimo_contato)<=30).length,
      d60:list.filter(c=>dateDays(c.Ultimo_contato)!==''&&dateDays(c.Ultimo_contato)<=60).length,
      over60:list.filter(c=>dateDays(c.Ultimo_contato)===''||dateDays(c.Ultimo_contato)>60).length
    });
    return {clients, openClients, ...bucket(openClients), all:bucket(clients)};
  };
  function moveButtonsToFooter(){
    if(document.getElementById('dashboardFooter')) return;
    const footer=document.createElement('footer');
    footer.id='dashboardFooter';
    footer.className='dashboardFooter';
    footer.innerHTML='<div class="footerBox"><h2>Relatórios e auditorias</h2><div class="footerButtons" id="footerButtons"></div></div>';
    const modal=document.getElementById('modal');
    (modal&&modal.parentNode?modal.parentNode:document.body).insertBefore(footer,modal||null);
    const target=document.getElementById('footerButtons');
    document.querySelectorAll('.actions a,.actions button').forEach(n=>target.appendChild(n));
    document.querySelectorAll('.sectionButtons button[data-panel]').forEach(n=>target.appendChild(n));
    document.querySelectorAll('.sectionButtons button.disabled').forEach(n=>n.remove());
    document.querySelectorAll('.sectionButtons button[data-panel]').forEach(b=>{
      b.onclick=()=>{const p=$(b.dataset.panel); if(p){p.classList.add('open');p.scrollIntoView({behavior:'smooth',block:'start'}); enhanceTables();}}
    });
  }
  function compactSidebar(){
    if(document.querySelector('.moreFilters')) return;
    const sidebar=document.querySelector('.sidebar'); if(!sidebar) return;
    const more=document.createElement('details');
    more.className='moreFilters';
    more.innerHTML='<summary>Mais filtros</summary>';
    const ids=['from','to','vendor','status','uf','city','ptype','colPreset'];
    ids.forEach(id=>{const el=$(id); const field=el&&el.closest('.field'); if(field)more.appendChild(field);});
    const cols=document.getElementById('columnsBox');
    const colDetails=cols&&cols.closest('details');
    if(colDetails) more.appendChild(colDetails);
    sidebar.appendChild(more);
  }
  function tweakLabels(){
    const monthLabel=[...document.querySelectorAll('.summaryHead span')].find(x=>/Mês do resumo|Mes do resumo/i.test(x.textContent));
    if(monthLabel) monthLabel.textContent='Período';
    const period=$('period');
    if(period){
      [...period.options].forEach(o=>{if(o.value==='month')o.textContent='Mês selecionado'});
      if(![...period.options].some(o=>o.value==='all')) period.insertAdjacentHTML('afterbegin','<option value="all">Todo período</option>');
    }
  }
  function ensureSellerPanel(){
    if($('sellerDetailPanel')) return;
    const list=$('sellerList');
    if(!list) return;
    const panel=document.createElement('div');
    panel.id='sellerDetailPanel';
    panel.className='sellerDetailPanel';
    list.insertAdjacentElement('afterend',panel);
  }
  window.renderSummaryFilters=function(){
    const sq=$('sectorQuick'),gq=$('groupQuick'),mp=$('monthPick');
    if(mp&&!mp.dataset.ready){
      const months=uniq(allRows.map(r=>r.AnoMes)).reverse();
      mp.innerHTML='<option value="">Todo período</option>'+months.map(m=>`<option value="${m}" ${m===DATA.current_month?'selected':''}>${m}</option>`).join('');
      mp.dataset.ready='1';
    }
    if(sq){
      const sectors=['',...uniq(allRows.map(r=>r.Setor))];
      sq.innerHTML='';
      sectors.forEach(v=>{const b=document.createElement('button');b.type='button';b.textContent=v||'Todos';b.className=($('sector').value===v?'active':'');b.onclick=()=>{$('sector').value=v;attendanceSector=v||'Todos';render()};sq.appendChild(b)});
    }
    if(gq){
      const opts=[...$('groupBy').options].map(o=>({value:o.value,label:o.textContent.replace('Agrupar por ','')}));
      gq.innerHTML='';
      opts.forEach(o=>{const b=document.createElement('button');b.type='button';b.textContent=o.label;b.className=($('groupBy').value===o.value?'active':'');b.onclick=()=>{$('groupBy').value=o.value;render();$('groupPanel').scrollIntoView({behavior:'smooth',block:'start'});};gq.appendChild(b)});
    }
  };
  window.inPeriod=function(row){
    const p=$('period').value, iso=row['Data ISO']||'';
    if(!iso) return true;
    if($('from').value&&iso<$('from').value) return false;
    if($('to').value&&iso>$('to').value) return false;
    const mp=$('monthPick');
    if((p==='month'||(mp&&mp.value)) && mp && mp.value) return iso.slice(0,7)===mp.value;
    if(p==='all') return true;
    if(p==='today') return iso===TODAY_ISO;
    if(p==='year') return iso.slice(0,4)==='2026';
    if(p==='quarter') return iso>='2026-04'&&iso<='2026-06';
    if(p==='week') return row.Semana==='2026-W20'||row.Semana==='2026-W19';
    return true;
  };
  function groupWithClients(rows,key){
    const m={};
    rows.forEach(r=>{const k=r[key]||'(sem informação)'; if(!m[k])m[k]={count:0,value:0,clients:new Set()}; m[k].count++; m[k].value+=(+r['Valor total da proposta']||0); m[k].clients.add(r['CPF/CNPJ']||r['Cliente Nome']||'')});
    return Object.entries(m).map(([name,v])=>({name,count:v.count,value:v.value,clients:v.clients.size})).sort((a,b)=>b.count-a.count);
  }
  function barsStatus(id,items){
    const el=$(id); if(!el) return; el.innerHTML='';
    const max=Math.max(1,...items.map(i=>i.count));
    items.slice(0,9).forEach(i=>{
      let d=document.createElement('div');d.className='barrow';
      d.innerHTML=`<div class="barlabel" title="${safeText(i.name)}">${safeText(i.name)}<span class="statusClients">${fmt.format(i.clients||0)} clientes</span></div><div class="bar"><i style="width:${i.count/max*100}%"></i></div><div class="barval">${fmt.format(i.count)}</div>`;
      el.appendChild(d);
    });
  }
  window.render=function(){
    renderSummaryFilters();
    let rows=filtered(); lastRows=rows;
    const value=rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0),
      openValue=rows.filter(r=>r['Status painel']==='Em aberto').reduce((a,r)=>a+(+r['Valor total da proposta']||0),0),
      doneValue=rows.filter(r=>r['Status painel']==='Concluídas').reduce((a,r)=>a+(+r['Valor total da proposta']||0),0),
      clients=uniqueClients(rows);
    $('kProps').textContent=fmt.format(rows.length);
    $('kValue').textContent=brl.format(value);
    if($('kOpenValue')) $('kOpenValue').textContent=brl.format(openValue);
    if($('kDoneValue')) $('kDoneValue').textContent=brl.format(doneValue);
    $('kClients').textContent=fmt.format(clients.length);
    const kpis=document.querySelector('.summaryKpis');
    if(kpis){
      kpis.classList.toggle('openMode',mode==='open');
      $('kValue').closest('.kpi').classList.toggle('hideOpenMode',mode==='open');
      $('kDoneValue').closest('.kpi').classList.toggle('hideOpenMode',mode==='open');
    }
    $('count').textContent=fmt.format(rows.length);
    $('tableTitle').textContent=mode==='open'?'Oportunidades em aberto':'Propostas comerciais';
    renderGroupPanel(rows); renderSellers(rows); barsStatus('statusBars',groupWithClients(rows,'Status painel')); bars('sectorBars',group(rows,'Setor')); bars('kitBars',group(rows,'Kit')); renderTable(rows); renderDups(); renderAudit(); renderAttendancePanel(); enhanceTables();
  };
  window.renderAttendancePanel=function(){
    const data=DATA.atendimentos||{}, clients=data.clientes||[], sectors=['Todos','Tráfego','Carteira','ESM'];
    const btns=$('attendanceSectorBtns'); if(!btns)return;
    const sec=effectiveSector();
    btns.innerHTML=''; sectors.forEach(s=>{const b=document.createElement('button');b.type='button';b.textContent=s;b.className=sec===s?'active':'';b.onclick=()=>{attendanceSector=s;if($('sector'))$('sector').value=s==='Todos'?'':s;render()};btns.appendChild(b)});
    const rows=clients.filter(c=>sec==='Todos'||c.Setor===sec);
    const withAtt=rows.filter(c=>c.Qtd_atendimentos>0), once=rows.filter(c=>c.Qtd_atendimentos===1), stale=rows.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30), open=rows.filter(c=>c.Tem_proposta_aberta==='Sim');
    $('attClients').textContent=fmt.format(rows.length); $('attWith').textContent=fmt.format(withAtt.length); $('attOnce').textContent=fmt.format(once.length); $('attStale').textContent=fmt.format(stale.length); $('attOpen').textContent=fmt.format(open.length);
    function item(c){return `<div class="activationItem"><div><strong>${safeText(c.Cliente||'(sem nome)')}</strong><small>${safeText(c.Motivo_ativacao)}</small><small>Último contato: ${safeText(c.Ultimo_contato||'sem registro')} | Atend.: ${fmt.format(c.Qtd_atendimentos||0)} | Vendedor: ${safeText(c.Vendedor_sugerido||'')}</small></div><div class="tag">${safeText(c.Prioridade)}</div></div>`}
    const openList=(data.plano||[]).filter(c=>(sec==='Todos'||c.Setor===sec)&&c.Tem_proposta_aberta==='Sim').slice(0,80);
    const noOpenList=(data.plano||[]).filter(c=>(sec==='Todos'||c.Setor===sec)&&c.Tem_proposta_aberta!=='Sim').slice(0,80);
    $('activationOpen').innerHTML=openList.length?openList.map(item).join(''):'<p class="muted" style="padding:12px">Sem clientes nesta condição.</p>';
    $('activationNoOpen').innerHTML=noOpenList.length?noOpenList.map(item).join(''):'<p class="muted" style="padding:12px">Sem clientes nesta condição.</p>';
    $('attendanceRows').innerHTML=rows.slice().sort((a,b)=>(+b.Prioridade_ordem||0)-(+a.Prioridade_ordem||0)).slice(0,220).map(c=>`<tr><td>${safeText(c.Cliente)}</td><td>${safeText(c.Setor)}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${safeText(c.Ultimo_contato||'')}</td><td>${safeText(c.Dias_sem_contato)}</td><td>${safeText(c.Vendedor_sugerido||'')}</td><td>${safeText(c.Tem_proposta_aberta)}</td></tr>`).join('');
  };
  window.renderSellers=function(rows){
    const el=$('sellerList'); if(!el)return; el.innerHTML='';
    group(rows,'Nome do Vendedor').slice(0,12).forEach(s=>{
      const sellerRows=rows.filter(r=>r['Nome do Vendedor']===s.name);
      const st=sellerAttendanceStats(sellerRows);
      const initial=(s.name||'?').trim()[0]||'?';
      let d=document.createElement('div');d.className='seller';
      d.innerHTML=`<div class="avatar">${safeText(initial)}</div><div><button>${safeText(s.name)}</button><small>${fmt.format(s.count)} propostas | ${fmt.format(uniqueClients(sellerRows).length)} clientes</small><div class="sellerStats"><span class="sellerChip">${fmt.format(st.clients.length)} com atendimento</span><span class="sellerChip">${fmt.format(st.d15)} até 15d</span><span class="sellerChip">${fmt.format(st.over60)} +60d</span></div></div><div class="miniVal">${brl.format(s.value)}</div>`;
      d.querySelector('button').onclick=()=>openSeller(s.name);
      el.appendChild(d);
    });
    ensureSellerPanel();
  };
  window.openSeller=function(name){
    ensureSellerPanel();
    const panel=$('sellerDetailPanel');
    const rows=rowsForSeller(name);
    const open=rows.filter(r=>r['Status painel']==='Em aberto'), done=rows.filter(r=>r['Status painel']==='Concluídas'), na=rows.filter(r=>r['Status painel']==='Não aprovadas');
    const valueOpen=open.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0), valueDone=done.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0);
    const rate=rows.length?Math.round(done.length/rows.length*100):0;
    const st=sellerAttendanceStats(rows);
    const initial=(name||'?').trim()[0]||'?';
    const plan=open.slice().sort((a,b)=>{
      const ad=dateDays((attMap.get(String(a['Cliente chave']||''))||{}).Ultimo_contato);
      const bd=dateDays((attMap.get(String(b['Cliente chave']||''))||{}).Ultimo_contato);
      const ap=ad===''?999:ad, bp=bd===''?999:bd;
      return bp-ap || (+b['Valor total da proposta']||0)-(+a['Valor total da proposta']||0);
    }).slice(0,120);
    panel.innerHTML=`<div class="sellerDetailHead"><div class="sellerHero" style="margin:0"><div class="photo">${safeText(initial)}</div><div><h3>${safeText(name||'Sem vendedor')}</h3><p class="muted">Panorama em ${TODAY_BR}: ${fmt.format(open.length)} propostas em aberto, ${fmt.format(done.length)} concluídas e taxa de fechamento de ${rate}%.</p></div></div><button class="close" id="closeSellerDetail">Fechar</button></div>
      <div class="sellerDetailBody">
        <div class="sellerPanorama">
          <div class="detail"><span>Em aberto</span><strong>${fmt.format(open.length)}</strong>${brl.format(valueOpen)}</div>
          <div class="detail"><span>Clientes em aberto</span><strong>${fmt.format(uniqueClients(open).length)}</strong>${fmt.format(st.openClients.length)} com atendimento</div>
          <div class="detail"><span>Concluídas</span><strong>${fmt.format(done.length)}</strong>${brl.format(valueDone)}</div>
          <div class="detail"><span>Não aprovadas</span><strong>${fmt.format(na.length)}</strong>propostas</div>
          <div class="detail"><span>Atend. em aberto até 15 dias</span><strong>${fmt.format(st.d15)}</strong>clientes</div>
          <div class="detail"><span>Atend. em aberto até 30 dias</span><strong>${fmt.format(st.d30)}</strong>clientes</div>
          <div class="detail"><span>Atend. em aberto até 60 dias</span><strong>${fmt.format(st.d60)}</strong>clientes</div>
          <div class="detail"><span>Sem contato ou +60 dias</span><strong>${fmt.format(st.over60)}</strong>clientes</div>
        </div>
        <button class="smallBtn" id="sellerActionBtn">Ação proposta</button>
        <div class="actionReport" id="sellerActionReport"><h3>Plano de ação de atendimento</h3><div class="tableWrap"><table class="tableCompact"><thead><tr><th>Prioridade</th><th>Proposta</th><th>Cliente</th><th>Valor</th><th>Último contato</th><th>Dias</th><th>Ação sugerida</th></tr></thead><tbody>${plan.map(r=>{const a=attMap.get(String(r['Cliente chave']||''))||{};const days=dateDays(a.Ultimo_contato);const label=days===''?'Sem histórico':days>60?'Mais de 60 dias':days>30?'Até 60 dias':days>15?'Até 30 dias':'Até 15 dias';const action=days===''||days>60?'Retomar contato hoje e validar interesse':days>30?'Reativar proposta e confirmar pendências':days>15?'Fazer follow-up de decisão':'Manter acompanhamento próximo';return `<tr><td>${safeText(label)}</td><td>${safeText(r['Proposta Numero']||r['Proposta ID']||'')}</td><td>${safeText(r['Cliente Nome']||'')}</td><td class="money">${brl.format(+r['Valor total da proposta']||0)}</td><td>${safeText(a.Ultimo_contato||'')}</td><td>${safeText(days)}</td><td>${safeText(action)}</td></tr>`}).join('')}</tbody></table></div></div>
        <div class="ops"><h3>Propostas por kit</h3><div class="bars" id="sellerKit"></div><h3>Propostas por período</h3><div class="bars" id="sellerPeriod"></div><h3>Status das propostas</h3><div class="bars" id="sellerStatus"></div></div>
      </div>`;
    panel.classList.add('open');
    $('closeSellerDetail').onclick=()=>panel.classList.remove('open');
    $('sellerActionBtn').onclick=()=>{$('sellerActionReport').classList.toggle('open'); enhanceTables();};
    bars('sellerKit',group(rows,'Kit'));bars('sellerPeriod',group(rows,'AnoMes'));barsStatus('sellerStatus',groupWithClients(rows,'Status painel'));
    enhanceTables();
    panel.scrollIntoView({behavior:'smooth',block:'start'});
  };
  window.openTopSeller=function(){const rows=lastRows.length?lastRows:filtered();const top=group(rows,'Nome do Vendedor')[0];if(top)openSeller(top.name)};
  const oldOpenGroup=window.openGroup;
  window.openGroup=function(key,name){
    oldOpenGroup(key,name);
    setTimeout(enhanceTables,0);
  };
  const oldOpenProposal=window.openProposal;
  if(oldOpenProposal){
    window.openProposal=function(r){
      oldOpenProposal(r);
      setTimeout(enhanceTables,0);
    };
  }
  const oldOpenClient=window.openClient;
  if(oldOpenClient){
    window.openClient=function(r){
      oldOpenClient(r);
      setTimeout(enhanceTables,0);
    };
  }
  function enhanceTables(){
    document.querySelectorAll('.groupDrawerTable table,.attendanceHistory table,.actionReport table,.ops table').forEach(table=>{
      if(table.dataset.sortReady) return;
      table.dataset.sortReady='1';
      table.querySelectorAll('thead th').forEach((th,idx)=>{
        th.classList.add('sortable');
        th.title='Clique para ordenar';
        th.onclick=()=>{
          const tbody=table.tBodies[0]; if(!tbody) return;
          const asc=th.dataset.dir!=='asc';
          table.querySelectorAll('th').forEach(x=>{x.classList.remove('activeSort','asc','desc');delete x.dataset.dir;});
          th.dataset.dir=asc?'asc':'desc'; th.classList.add('activeSort',asc?'asc':'desc');
          [...tbody.rows].sort((a,b)=>{
            const av=a.cells[idx]?.textContent.trim()||'', bv=b.cells[idx]?.textContent.trim()||'';
            const an=numberFromText(av), bn=numberFromText(bv);
            let cmp=(!Number.isNaN(an)&&!Number.isNaN(bn))?an-bn:av.localeCompare(bv,'pt-BR',{numeric:true,sensitivity:'base'});
            return asc?cmp:-cmp;
          }).forEach(r=>tbody.appendChild(r));
        };
      });
    });
  }
  moveButtonsToFooter();
  compactSidebar();
  tweakLabels();
  ensureSellerPanel();
  if($('monthPick')) $('monthPick').addEventListener('input',render);
  if($('detailSeller')) $('detailSeller').onclick=openTopSeller;
  render();
})();
</script>
"""


def ensure_dirs() -> None:
    for folder in [
        ROOT / "01 - Ultimo alteração, envie sempre para raiz ( Como se fosse o espaço para o mais atualizado para todos verem)",
        ROOT / "02 - Ajustes",
        ROOT / "03 -Ajustes",
        ROOT / "04 - Fonte",
        ROOT / "05 - Apoio",
        ROOT / "06 - Verificação manual",
        ROOT / "07 - Auditoria",
        ROOT / "08 - Códigos",
        ROOT / "09 - Regras",
        LOTE,
    ]:
        folder.mkdir(exist_ok=True)


def update_listagem(actions: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for item in actions:
        ws.append([now, "AJUSTE_14_PANORAMA_RX", item["tipo"], item["acao"], item["contexto"], item["obs"]])
    fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 26 if col != 4 else 44
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    tab = Table(displayName="TabelaLinhaTempoAjuste14", ref=f"A1:F{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)
    wb.save(LISTAGEM)


def main() -> None:
    ensure_dirs()
    html = SOURCE.read_text(encoding="utf-8")
    if "ajuste14-panorama-rx" not in html:
        html = html.replace("</head>", CSS + "\n</head>")
        html = html.replace("</body>", JS + "\n</body>")
    OUT.write_text(html, encoding="utf-8")
    shutil.copy2(OUT, ROOT_DASH)
    shutil.copy2(OUT, LOTE_OUT)
    shutil.copy2(OUT, LOTE_DASH)

    actions = [
        {"tipo": "Layout", "acao": "Mover botões de relatório, mapa mental, imprimir, duplicadas e auditoria para o rodapé.", "contexto": "Usuário pediu tirar esses botões do topo e enviar para o rodapé.", "obs": "Botões de Propostas em aberto e RX da equipe continuam no topo como navegação rápida."},
        {"tipo": "Filtros", "acao": "Compactar filtros laterais com área Mais filtros.", "contexto": "Usuário pediu retirar excesso de campos fixos para deixar a lateral mais limpa.", "obs": "Busca, período e setor ficam visíveis; demais filtros ficam recolhidos."},
        {"tipo": "Filtros globais", "acao": "Setores rápidos, período e mês passam a recalcular todo o dashboard.", "contexto": "Usuário pediu filtros rápidos aplicados ao dashboard inteiro.", "obs": "Inclui KPIs, RX, propostas, agrupamentos e atendimentos."},
        {"tipo": "Agrupamentos", "acao": "Botões de agrupamento agora abrem a área de agrupamentos e mantêm lista clicável.", "contexto": "Usuário informou que botões não abriam nada.", "obs": "Listas abertas recebem cabeçalho ordenável."},
        {"tipo": "RX da equipe", "acao": "Detalhe do vendedor passa a abrir deslizando para baixo no RX.", "contexto": "Usuário pediu abertura de vendedor sem modal central.", "obs": "Inclui propostas abertas/concluídas, valores e panorama de atendimentos."},
        {"tipo": "Atendimentos", "acao": "Integrar atendimento no RX por vendedor com 15, 30, 60 e mais de 60 dias.", "contexto": "Usuário pediu panorama de atendimento por vendedor e plano de ação.", "obs": "Botão Ação proposta abre relatório de atendimento para distribuir follow-up."},
        {"tipo": "KPIs", "acao": "No modo Em aberto, ocultar Valor total e Valor concluído.", "contexto": "Usuário pediu retirar esses valores das propostas em aberto.", "obs": "Ficam Propostas, Valor em aberto e Clientes."},
        {"tipo": "Status", "acao": "Adicionar total de clientes embaixo de cada status.", "contexto": "Usuário pediu clientes abaixo de cada status.", "obs": "Aplicado no bloco Propostas comerciais."},
        {"tipo": "Ordenação", "acao": "Adicionar ordenação em listas/tabelas secundárias.", "contexto": "Usuário pediu todas as listas com cabeçalho ordenável.", "obs": "Aplicado em agrupamentos, atendimentos, histórico do cliente e plano de ação."},
    ]
    update_listagem(actions)
    SUMMARY.write_text(json.dumps({"versao": OUT.name, "arquivos": [str(OUT), str(ROOT_DASH), str(LOTE_OUT), str(LOTE_DASH), str(LISTAGEM)], "acoes": actions}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "out": str(OUT), "root": str(ROOT_DASH), "lote": str(LOTE_DASH), "listagem": str(LISTAGEM)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
