from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
SOURCE = ROOT / "dashboard_AJUSTE_14_PANORAMA_RX.html"
OUT = ROOT / "dashboard_AJUSTE_15_LIMPO_CLICAVEL.html"
ROOT_DASH = ROOT / "dashboard.html"
LOTE_DASH = LOTE / "dashboard.html"
LOTE_OUT = LOTE / OUT.name
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_15.json"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


CSS = r"""
<style id="ajuste15-limpo-clicavel">
.main{display:grid;grid-template-columns:1fr;gap:14px}
.sidebar{position:static;display:grid;grid-template-columns:1.2fr repeat(4,minmax(150px,1fr));gap:12px;align-items:end;padding:14px 18px}
.sidebar .seg{grid-column:auto;margin:0}
.sidebar>.field{margin:0}
.sidebar .field label{font-size:11px}
.sidebar .field input,.sidebar .field select{height:44px;border-radius:8px}
.sidebar .moreFilters{grid-column:1/-1;margin-top:0}
.sidebar .moreFilters[open]{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:10px;align-items:end}
.sidebar .moreFilters summary{grid-column:1/-1}
.sidebar .moreFilters .field{margin:0}
.summaryFilters{padding:12px 18px}
.summaryHead{display:block;margin-bottom:8px}
.summaryHead h2{font-size:18px}
.summaryHead .field{display:none!important}
.quickRows{grid-template-columns:1fr 1fr;gap:18px}
.quickButtons button,.seg button,.footerButtons button,.footerButtons a{border-radius:8px}
.kpis.summaryKpis{grid-template-columns:repeat(5,1fr)}
.kpi{border-radius:10px;padding:14px 16px;box-shadow:0 8px 22px rgba(20,74,45,.07)}
.kpi.clickable,.barrow.clickable,.groupCard,.sellerRow,.attendanceRow{cursor:pointer}
.kpi.clickable:hover,.barrow.clickable:hover,.sellerRow:hover,.attendanceRow:hover{background:#f7fcf9;outline:1px solid #b9dcc8}
.rx{grid-template-columns:1.25fr .75fr}
.sellerList{gap:0;border:1px solid var(--line);border-radius:10px;overflow:hidden}
.sellerTableHead,.sellerRow{display:grid;grid-template-columns:1.6fr .8fr .75fr .8fr .9fr .75fr .75fr .75fr .75fr .9fr;gap:8px;align-items:center;padding:10px 12px}
.sellerTableHead{background:#e7f5ec;color:var(--green);font-size:11px;text-transform:uppercase;font-weight:900;position:sticky;top:0;z-index:1}
.sellerRow{background:#fff;border-top:1px solid #edf3ef;font-size:13px}
.sellerName{font-weight:900;color:var(--green);line-height:1.2}
.metricBtn{border:0;background:transparent;color:var(--green);font-weight:900;text-align:left;padding:0;text-decoration:underline}
.sellerActionMini{border:1px solid var(--line);background:#f8fffa;color:var(--leaf);border-radius:999px;padding:6px 9px;font-size:12px;font-weight:900}
.sellerDetailHead h3{font-size:20px}
.sellerHero .photo{display:none}
.sellerPanorama{grid-template-columns:repeat(4,1fr)}
.sellerPanorama .detail{border-radius:10px}
.attendancePanel{padding:16px 20px;border-radius:10px;box-shadow:0 8px 22px rgba(20,74,45,.07)}
.attendanceGrid{grid-template-columns:repeat(5,1fr)}
.attendanceKpi{border-radius:10px}
.activationColumns{display:none!important}
.attendanceUnified{margin-top:14px;border:1px solid var(--line);border-radius:10px;overflow:hidden;background:#fff}
.attendanceUnified h3{margin:0;padding:12px 14px;background:#e7f5ec;color:var(--green)}
.attendanceUnified table{min-width:1180px}
.attendanceUnified td,.attendanceUnified th{font-size:12px}
.expandPanel{display:none!important}
.expandPanel.open{display:block!important}
.dashboardFooter .footerBox{border-radius:10px}
@media(max-width:1200px){.sidebar{grid-template-columns:1fr 1fr}.rx{grid-template-columns:1fr}.sellerTableHead,.sellerRow{grid-template-columns:1.5fr .7fr .7fr .8fr .8fr}.sellerTableHead span:nth-child(n+6),.sellerRow span:nth-child(n+6),.sellerRow button:nth-child(n+6){display:none}.kpis.summaryKpis,.attendanceGrid{grid-template-columns:1fr 1fr}}
@media(max-width:700px){.sidebar{grid-template-columns:1fr}.quickRows{grid-template-columns:1fr}.kpis.summaryKpis,.attendanceGrid{grid-template-columns:1fr}.sellerTableHead{display:none}.sellerRow{grid-template-columns:1fr 1fr}.sellerName{grid-column:1/-1}}
</style>
"""


JS = r"""
<script id="ajuste15-limpo-clicavel">
(function(){
  const TODAY_ISO='2026-05-11';
  const TODAY_BR='11/05/2026';
  const safe=v=>String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
  const uniqLocal=a=>[...new Set(a.filter(Boolean))].sort();
  const clientKey=r=>r['CPF/CNPJ']||r['Cliente Nome']||r['Cliente chave']||'';
  const clientCount=rows=>uniqLocal(rows.map(clientKey)).length;
  const daysFrom=d=>{
    if(!d) return '';
    const p=String(d).split('/');
    const iso=p.length===3?`${p[2]}-${p[1]}-${p[0]}`:String(d).slice(0,10);
    const diff=(new Date(TODAY_ISO+'T00:00:00')-new Date(iso+'T00:00:00'))/86400000;
    return Number.isFinite(diff)?Math.floor(diff):'';
  };
  const attRows=()=>((DATA.atendimentos||{}).clientes||[]);
  const attMap=(()=>{const m=new Map();attRows().forEach(c=>m.set(String(c.Cliente_chave||''),c));return m})();
  const rowsByCurrent=()=>lastRows&&lastRows.length?lastRows:filtered();
  const openListModal=(title,rows)=>{
    $('modalTitle').textContent=title;
    $('modalBody').innerHTML=`<div class="groupSummary"><div class="detail"><span>Propostas</span><strong>${fmt.format(rows.length)}</strong></div><div class="detail"><span>Clientes</span><strong>${fmt.format(clientCount(rows))}</strong></div><div class="detail"><span>Valor</span><strong>${brl.format(rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0))}</strong></div></div><div class="groupDrawerTable"><table><thead><tr><th>Proposta</th><th>Data</th><th>Cliente</th><th>Status</th><th>Vendedor</th><th>Setor</th><th>Kit</th><th>Valor</th></tr></thead><tbody>${rows.slice(0,600).map(r=>`<tr><td><button class="linkBtn prop" onclick="openProposalById('${safe(String(r['Proposta ID']||''))}')">${safe(r['Proposta Numero']||r['Proposta ID']||'')}</button></td><td>${safe(r['Proposta Data']||'')}</td><td>${safe(r['Cliente Nome']||'')}</td><td><span class="pill">${safe(r['Status painel']||'')}</span></td><td>${safe(r['Nome do Vendedor']||'')}</td><td>${safe(r.Setor||'')}</td><td>${safe(r.Kit||'')}</td><td class="money">${brl.format(+r['Valor total da proposta']||0)}</td></tr>`).join('')}</tbody></table></div>`;
    $('modal').classList.add('open','drawer');
    setTimeout(()=>document.querySelectorAll('.groupDrawerTable table th').forEach((th,i)=>th.onclick=()=>sortTable(th.closest('table'),i,th)),0);
  };
  function sortTable(table,idx,th){
    const tbody=table.tBodies[0]; if(!tbody)return;
    const asc=th.dataset.dir!=='asc';
    table.querySelectorAll('th').forEach(x=>{x.classList.remove('activeSort','asc','desc');delete x.dataset.dir;});
    th.dataset.dir=asc?'asc':'desc'; th.classList.add('activeSort',asc?'asc':'desc');
    const num=t=>Number(String(t).replace(/\s/g,'').replace(/[^\d,-]/g,'').replace(/\./g,'').replace(',','.'));
    [...tbody.rows].sort((a,b)=>{
      const av=a.cells[idx]?.textContent.trim()||'', bv=b.cells[idx]?.textContent.trim()||'';
      const an=num(av), bn=num(bv);
      const cmp=Number.isFinite(an)&&Number.isFinite(bn)?an-bn:av.localeCompare(bv,'pt-BR',{numeric:true,sensitivity:'base'});
      return asc?cmp:-cmp;
    }).forEach(r=>tbody.appendChild(r));
  }
  function setupCleanFilters(){
    document.querySelectorAll('#duplicadasPanel,#auditoriaPanel').forEach(p=>p.classList.remove('open'));
    const monthField=document.querySelector('.summaryHead .field');
    if(monthField) monthField.style.display='none';
    const period=$('period');
    if(period&&!period.dataset.monthsReady){
      const months=uniqLocal(allRows.map(r=>r.AnoMes)).reverse();
      period.innerHTML='<option value="all">Todo período</option>'+months.map(m=>`<option value="month:${m}" ${m===DATA.current_month?'selected':''}>${m}</option>`).join('')+'<option value="today">Hoje</option><option value="week">Semana</option><option value="quarter">Trimestre</option><option value="year">Ano</option>';
      period.dataset.monthsReady='1';
    }
    const q=$('q'); if(q) q.closest('.field').style.gridColumn='span 1';
    let nav=document.querySelector('.sectionButtons');
    if(nav&&!$('goAttendance')){
      const b=document.createElement('button');
      b.id='goAttendance'; b.textContent='Atendimento';
      b.onclick=()=>$('attendancePanel').scrollIntoView({behavior:'smooth',block:'start'});
      const rx=$('goRx'); rx?rx.insertAdjacentElement('afterend',b):nav.appendChild(b);
    }
  }
  window.inPeriod=function(row){
    const p=$('period').value, iso=row['Data ISO']||'';
    if(!iso)return true;
    if($('from').value&&iso<$('from').value)return false;
    if($('to').value&&iso>$('to').value)return false;
    if(p&&p.startsWith('month:')) return iso.slice(0,7)===p.slice(6);
    if(p==='all')return true;
    if(p==='today')return iso===TODAY_ISO;
    if(p==='year')return iso.slice(0,4)==='2026';
    if(p==='quarter')return iso>='2026-04'&&iso<='2026-06';
    if(p==='week')return row.Semana==='2026-W20'||row.Semana==='2026-W19';
    return true;
  };
  function sellerStats(rows){
    const open=rows.filter(r=>r['Status painel']==='Em aberto'), done=rows.filter(r=>r['Status painel']==='Concluídas');
    const keys=[...new Set(rows.map(r=>String(r['Cliente chave']||'')).filter(Boolean))];
    const clients=keys.map(k=>attMap.get(k)).filter(Boolean);
    const openClients=[...new Set(open.map(r=>String(r['Cliente chave']||'')).filter(Boolean))].map(k=>attMap.get(k)).filter(Boolean);
    const b=list=>({d15:list.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=15).length,d30:list.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=30).length,d60:list.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=60).length,over60:list.filter(c=>daysFrom(c.Ultimo_contato)===''||daysFrom(c.Ultimo_contato)>60).length});
    return {open,done,clients,openClients,...b(openClients),valueOpen:open.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0),valueDone:done.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0)};
  }
  window.renderSellers=function(rows){
    const el=$('sellerList'); if(!el)return; el.innerHTML='<div class="sellerTableHead"><span>Vendedor</span><span>Propostas</span><span>Clientes</span><span>Abertas</span><span>Valor aberto</span><span>Atend.</span><span>15d</span><span>30d</span><span>+60d</span><span>Ação</span></div>';
    group(rows,'Nome do Vendedor').slice(0,18).forEach(s=>{
      const sr=rows.filter(r=>r['Nome do Vendedor']===s.name), st=sellerStats(sr);
      const line=document.createElement('div'); line.className='sellerRow';
      line.innerHTML=`<span class="sellerName">${safe(s.name||'Sem vendedor')}</span><button class="metricBtn">${fmt.format(sr.length)}</button><button class="metricBtn">${fmt.format(clientCount(sr))}</button><button class="metricBtn">${fmt.format(st.open.length)}</button><button class="metricBtn">${brl.format(st.valueOpen)}</button><button class="metricBtn">${fmt.format(st.clients.length)}</button><button class="metricBtn">${fmt.format(st.d15)}</button><button class="metricBtn">${fmt.format(st.d30)}</button><button class="metricBtn">${fmt.format(st.over60)}</button><button class="sellerActionMini">Ação</button>`;
      line.onclick=()=>openSeller(s.name);
      const btns=line.querySelectorAll('button');
      btns[0].onclick=e=>{e.stopPropagation();openListModal('Propostas de '+s.name,sr)};
      btns[1].onclick=e=>{e.stopPropagation();openListModal('Clientes/propostas de '+s.name,sr)};
      btns[2].onclick=e=>{e.stopPropagation();openListModal('Em aberto - '+s.name,st.open)};
      btns[3].onclick=e=>{e.stopPropagation();openListModal('Valor em aberto - '+s.name,st.open)};
      btns[4].onclick=e=>{e.stopPropagation();openSeller(s.name)};
      btns[5].onclick=e=>{e.stopPropagation();openSeller(s.name)};
      btns[6].onclick=e=>{e.stopPropagation();openSeller(s.name)};
      btns[7].onclick=e=>{e.stopPropagation();openSeller(s.name)};
      btns[8].onclick=e=>{e.stopPropagation();openSeller(s.name);setTimeout(()=>$('sellerActionBtn')&&$('sellerActionBtn').click(),250)};
      el.appendChild(line);
    });
    ensureSellerPanel();
  };
  window.openSeller=function(name){
    ensureSellerPanel();
    const panel=$('sellerDetailPanel');
    const rows=rowsByCurrent().filter(r=>r['Nome do Vendedor']===name);
    const st=sellerStats(rows), na=rows.filter(r=>r['Status painel']==='Não aprovadas');
    const plan=st.open.slice().sort((a,b)=>{
      const ad=daysFrom((attMap.get(String(a['Cliente chave']||''))||{}).Ultimo_contato), bd=daysFrom((attMap.get(String(b['Cliente chave']||''))||{}).Ultimo_contato);
      return (bd===''?999:bd)-(ad===''?999:ad)||(+b['Valor total da proposta']||0)-(+a['Valor total da proposta']||0);
    }).slice(0,120);
    panel.innerHTML=`<div class="sellerDetailHead"><div><h3>${safe(name||'Sem vendedor')}</h3><p class="muted">Panorama em ${TODAY_BR}. Clique nos números para abrir a lista correspondente.</p></div><button class="close" id="closeSellerDetail">Fechar</button></div>
      <div class="sellerDetailBody"><div class="sellerPanorama">
      <button class="detail clickable" id="detOpen"><span>Em aberto</span><strong>${fmt.format(st.open.length)}</strong>${brl.format(st.valueOpen)}</button>
      <button class="detail clickable" id="detClients"><span>Clientes abertos</span><strong>${fmt.format(clientCount(st.open))}</strong>${fmt.format(st.openClients.length)} com atendimento</button>
      <button class="detail clickable" id="detDone"><span>Concluídas</span><strong>${fmt.format(st.done.length)}</strong>${brl.format(st.valueDone)}</button>
      <button class="detail clickable" id="detNA"><span>Não aprovadas</span><strong>${fmt.format(na.length)}</strong>propostas</button>
      <button class="detail clickable" id="det15"><span>Atendimento até 15 dias</span><strong>${fmt.format(st.d15)}</strong>clientes</button>
      <button class="detail clickable" id="det30"><span>Atendimento até 30 dias</span><strong>${fmt.format(st.d30)}</strong>clientes</button>
      <button class="detail clickable" id="det60"><span>Atendimento até 60 dias</span><strong>${fmt.format(st.d60)}</strong>clientes</button>
      <button class="detail clickable" id="detOver"><span>Sem contato ou +60 dias</span><strong>${fmt.format(st.over60)}</strong>clientes</button>
      </div><button class="smallBtn" id="sellerActionBtn">Ação proposta</button>
      <div class="actionReport" id="sellerActionReport"><h3>Plano de ação de atendimento</h3><div class="tableWrap"><table class="tableCompact"><thead><tr><th>Prioridade</th><th>Proposta</th><th>Cliente</th><th>Valor</th><th>Último contato</th><th>Dias</th><th>Ação sugerida</th></tr></thead><tbody>${plan.map(r=>{const a=attMap.get(String(r['Cliente chave']||''))||{};const d=daysFrom(a.Ultimo_contato);const pri=d===''?'Sem histórico':d>60?'Mais de 60 dias':d>30?'31 a 60 dias':d>15?'16 a 30 dias':'Até 15 dias';const action=d===''||d>60?'Retomar contato hoje e validar interesse':d>30?'Reativar proposta e confirmar pendências':d>15?'Fazer follow-up de decisão':'Manter acompanhamento próximo';return `<tr><td>${safe(pri)}</td><td>${safe(r['Proposta Numero']||r['Proposta ID']||'')}</td><td>${safe(r['Cliente Nome']||'')}</td><td class="money">${brl.format(+r['Valor total da proposta']||0)}</td><td>${safe(a.Ultimo_contato||'')}</td><td>${safe(d)}</td><td>${safe(action)}</td></tr>`}).join('')}</tbody></table></div></div></div>`;
    panel.classList.add('open');
    $('closeSellerDetail').onclick=()=>panel.classList.remove('open');
    $('detOpen').onclick=()=>openListModal('Em aberto - '+name,st.open);
    $('detClients').onclick=()=>openListModal('Clientes com proposta aberta - '+name,st.open);
    $('detDone').onclick=()=>openListModal('Concluídas - '+name,st.done);
    $('detNA').onclick=()=>openListModal('Não aprovadas - '+name,na);
    ['det15','det30','det60','detOver'].forEach(id=>$(id).onclick=()=>{$('sellerActionReport').classList.add('open')});
    $('sellerActionBtn').onclick=()=>{$('sellerActionReport').classList.toggle('open')};
    setTimeout(()=>document.querySelectorAll('#sellerActionReport table th').forEach((th,i)=>th.onclick=()=>sortTable(th.closest('table'),i,th)),0);
    panel.scrollIntoView({behavior:'smooth',block:'start'});
  };
  window.renderAttendancePanel=function(){
    const data=DATA.atendimentos||{}, clients=data.clientes||[], sec=($('sector')&&$('sector').value)||attendanceSector||'Todos';
    const btns=$('attendanceSectorBtns'); if(!btns)return;
    const sectors=['Todos','Tráfego','Carteira','ESM'];
    btns.innerHTML=''; sectors.forEach(s=>{const b=document.createElement('button');b.type='button';b.textContent=s;b.className=sec===s?'active':'';b.onclick=()=>{attendanceSector=s;if($('sector'))$('sector').value=s==='Todos'?'':s;render()};btns.appendChild(b)});
    const rows=clients.filter(c=>sec==='Todos'||c.Setor===sec);
    const withAtt=rows.filter(c=>+c.Qtd_atendimentos>0), once=rows.filter(c=>+c.Qtd_atendimentos===1), stale=rows.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30), open=rows.filter(c=>c.Tem_proposta_aberta==='Sim');
    $('attClients').textContent=fmt.format(rows.length); $('attClients').closest('.attendanceKpi').querySelector('span').textContent='Clientes dos atendimentos';
    $('attWith').textContent=fmt.format(withAtt.length); $('attOnce').textContent=fmt.format(once.length); $('attStale').textContent=fmt.format(stale.length); $('attOpen').textContent=fmt.format(open.length);
    $('attOpen').closest('.attendanceKpi').querySelector('span').textContent='Com proposta em aberto';
    let box=$('attendanceUnified');
    if(!box){box=document.createElement('div');box.id='attendanceUnified';box.className='attendanceUnified';document.querySelector('.activationColumns').insertAdjacentElement('afterend',box);}
    const plan=(data.plano||[]).filter(c=>sec==='Todos'||c.Setor===sec).slice(0,500);
    box.innerHTML=`<h3>Fila única de atendimento</h3><div class="tableWrap"><table><thead><tr><th>Cliente</th><th>Setor</th><th>Vendedor sugerido</th><th>Proposta aberta</th><th>Prioridade</th><th>Último contato</th><th>Atendimentos</th><th>Motivo</th></tr></thead><tbody>${plan.map(c=>`<tr class="attendanceRow"><td>${safe(c.Cliente||'')}</td><td>${safe(c.Setor||'')}</td><td>${safe(c.Vendedor_sugerido||'')}</td><td>${safe(c.Tem_proposta_aberta||'')}</td><td>${safe(c.Prioridade||'')}</td><td>${safe(c.Ultimo_contato||'')}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${safe(c.Motivo_ativacao||'')}</td></tr>`).join('')}</tbody></table></div>`;
    box.querySelectorAll('th').forEach((th,i)=>th.onclick=()=>sortTable(th.closest('table'),i,th));
    $('attendanceRows').innerHTML=rows.slice(0,220).map(c=>`<tr><td>${safe(c.Cliente)}</td><td>${safe(c.Setor)}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${safe(c.Ultimo_contato||'')}</td><td>${safe(c.Dias_sem_contato)}</td><td>${safe(c.Vendedor_sugerido||'')}</td><td>${safe(c.Tem_proposta_aberta)}</td></tr>`).join('');
  };
  const oldRender=window.render;
  window.render=function(){
    oldRender();
    setupCleanFilters();
    document.querySelectorAll('.kpi').forEach(k=>k.classList.add('clickable'));
    $('kProps')?.closest('.kpi')?.addEventListener('click',()=>openListModal('Propostas filtradas',rowsByCurrent()),{once:true});
    $('kOpenValue')?.closest('.kpi')?.addEventListener('click',()=>openListModal('Valor em aberto',rowsByCurrent().filter(r=>r['Status painel']==='Em aberto')),{once:true});
    $('kDoneValue')?.closest('.kpi')?.addEventListener('click',()=>openListModal('Valor concluído',rowsByCurrent().filter(r=>r['Status painel']==='Concluídas')),{once:true});
    $('kClients')?.closest('.kpi')?.addEventListener('click',()=>openListModal('Clientes filtrados',rowsByCurrent()),{once:true});
    document.querySelectorAll('#statusBars .barrow').forEach(row=>{row.classList.add('clickable');const name=row.querySelector('.barlabel')?.childNodes[0]?.textContent?.trim();row.onclick=()=>openListModal('Status: '+name,rowsByCurrent().filter(r=>(r['Status painel']||'')===name));});
    document.querySelectorAll('#kitBars .barrow').forEach(row=>{row.classList.add('clickable');const name=row.querySelector('.barlabel')?.textContent?.trim();row.onclick=()=>openListModal('Kit: '+name,rowsByCurrent().filter(r=>(r.Kit||'(sem informação)')===name));});
    document.querySelectorAll('#sectorBars .barrow').forEach(row=>{row.classList.add('clickable');const name=row.querySelector('.barlabel')?.textContent?.trim();row.onclick=()=>{$('sector').value=name==='(sem informação)'?'':name;render();};});
    document.querySelectorAll('#duplicadasPanel,#auditoriaPanel').forEach(p=>{if(!p.dataset.userOpened)p.classList.remove('open')});
    document.querySelectorAll('.footerButtons button[data-panel]').forEach(b=>{b.onclick=()=>{const p=$(b.dataset.panel); if(p){p.dataset.userOpened='1';p.classList.add('open');p.scrollIntoView({behavior:'smooth',block:'start'});}}});
    document.querySelectorAll('button[data-close]').forEach(b=>{b.onclick=()=>{const p=$(b.dataset.close); if(p){delete p.dataset.userOpened;p.classList.remove('open');}}});
  };
  setupCleanFilters();
  render();
})();
</script>
"""


def ensure_dirs() -> None:
    for folder in [ROOT / "01 - Ultimo alteração, envie sempre para raiz ( Como se fosse o espaço para o mais atualizado para todos verem)", ROOT / "02 - Ajustes", ROOT / "03 -Ajustes", ROOT / "04 - Fonte", ROOT / "05 - Apoio", ROOT / "06 - Verificação manual", ROOT / "07 - Auditoria", ROOT / "08 - Códigos", ROOT / "09 - Regras", LOTE]:
        folder.mkdir(exist_ok=True)


def update_listagem(actions: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for item in actions:
        ws.append([now, "AJUSTE_15_LIMPO_CLICAVEL", item["tipo"], item["acao"], item["contexto"], item["obs"]])
    fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 28 if col != 4 else 48
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    tab = Table(displayName="TabelaLinhaTempoAjuste15", ref=f"A1:F{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)
    wb.save(LISTAGEM)


def main() -> None:
    ensure_dirs()
    html = SOURCE.read_text(encoding="utf-8")
    if "ajuste15-limpo-clicavel" not in html:
        html = html.replace("</head>", CSS + "\n</head>")
        html = html.replace("</body>", JS + "\n</body>")
    OUT.write_text(html, encoding="utf-8")
    shutil.copy2(OUT, ROOT_DASH)
    shutil.copy2(OUT, LOTE_OUT)
    shutil.copy2(OUT, LOTE_DASH)
    actions = [
        {"tipo": "Visual", "acao": "Reorganizar filtros no topo em formato horizontal e reduzir a lateral.", "contexto": "Usuário gostou do formato de filtro do dashboard gestor.", "obs": "Período concentra a seleção de mês; seletor duplicado do resumo fica oculto."},
        {"tipo": "Rodapé", "acao": "Manter Duplicadas e Auditoria ERP fechados por padrão.", "contexto": "Usuário pediu deixar só o botão e abrir apenas ao clicar.", "obs": "Botão Fechar recolhe novamente."},
        {"tipo": "Menu", "acao": "Ativar botão Atendimento no menu superior.", "contexto": "Usuário pediu abrir atendimento no menu.", "obs": "Botão rola para o bloco de atendimento."},
        {"tipo": "RX", "acao": "Trocar panorama de vendedor por tabela limpa e clicável.", "contexto": "Usuário achou o formato anterior poluído e pouco claro.", "obs": "Linhas e números abrem detalhe/lista."},
        {"tipo": "Atendimentos", "acao": "Unificar atendimento em uma fila única.", "contexto": "Usuário disse que separar com/sem proposta não fez sentido.", "obs": "Coluna indica se há proposta aberta; rótulo esclarece clientes dos atendimentos."},
        {"tipo": "Clicabilidade", "acao": "Adicionar clique em KPIs, barras, status, kits, vendedores e listas.", "contexto": "Usuário pediu tudo clicável.", "obs": "Abre lista filtrada ou detalhe sem deixar tudo exposto na tela inicial."},
    ]
    update_listagem(actions)
    SUMMARY.write_text(json.dumps({"versao": OUT.name, "arquivos": [str(OUT), str(ROOT_DASH), str(LOTE_OUT), str(LOTE_DASH), str(LISTAGEM)], "acoes": actions}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "out": str(OUT), "root": str(ROOT_DASH), "lote": str(LOTE_DASH), "listagem": str(LISTAGEM)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
