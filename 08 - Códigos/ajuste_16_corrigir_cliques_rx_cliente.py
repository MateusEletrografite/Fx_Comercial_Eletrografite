from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
SOURCE = ROOT / "dashboard_AJUSTE_15_LIMPO_CLICAVEL.html"
OUT = ROOT / "dashboard_AJUSTE_16_CLIQUES_RX_CLIENTE.html"
ROOT_DASH = ROOT / "dashboard.html"
LOTE_DASH = LOTE / "dashboard.html"
LOTE_OUT = LOTE / OUT.name
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_16.json"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


CSS = r"""
<style id="ajuste16-cliques-rx-cliente">
.hero{min-height:178px!important}
.hero .wrap{padding:16px 28px 20px!important}
.top{grid-template-columns:auto 1fr!important;max-width:1420px;margin:0 auto}
.top>div:first-child{display:none}
.brand{grid-column:auto!important;justify-content:flex-start!important;text-align:left!important}
.brand .logo{width:48px!important;height:48px!important;border-radius:10px!important;font-size:0!important;background:#fff url('logo.png') center/contain no-repeat}
.brand .logo::after{content:'EG';font-size:16px;color:var(--leaf);font-weight:900}
.brand h1{font-size:22px!important;white-space:normal!important;line-height:1.15!important;max-width:760px}
.brand h1 strong{color:#dfffd2}
.quickNav{margin-top:14px!important;text-align:left!important}
.quickNav h2{font-size:14px!important;margin:0 0 10px!important;color:#e9fff0!important}
.sectionButtons{justify-content:flex-start!important;gap:8px!important}
.sectionButtons button{min-height:36px!important;padding:9px 13px!important;font-size:13px!important;border-radius:8px!important}
.groupPanel{display:none!important}
#setorPanel,.tablePanel{order:20}
.rxClientPanel{background:#fff;border:1px solid var(--line);border-radius:10px;box-shadow:0 8px 22px rgba(20,74,45,.07);margin:14px 0;padding:0;display:none;overflow:hidden}
.rxClientPanel.open{display:block;animation:slideDown .2s ease-out}
.rxClientHead{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:14px 18px;border-bottom:1px solid var(--line)}
.rxClientHead h2{margin:0;color:var(--green);font-size:20px}
.rxClientBody{padding:14px 18px}
.rxClientSummary{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:12px}
.rxClientSummary .detail{border-radius:10px}
.rxClientBody table{min-width:1100px}
.rxClientBody th{cursor:pointer}
.rxClientClose{border:1px solid var(--line);background:#fff;border-radius:8px;padding:8px 10px;color:var(--green);font-weight:800}
.attendanceGroups{margin:12px 0;border:1px solid var(--line);border-radius:10px;overflow:hidden;background:#fff}
.attendanceGroups h3{margin:0;padding:10px 12px;background:#e7f5ec;color:var(--green);font-size:16px}
.attendanceGroupGrid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;padding:10px}
.attendanceGroupBtn{border:1px solid var(--line);background:#fbfffc;border-radius:8px;padding:10px;text-align:left;color:var(--green);font-weight:900}
.attendanceGroupBtn small{display:block;color:var(--muted);font-weight:700;margin-top:4px}
.attendanceKpi{cursor:pointer}
.attendanceKpi:hover,.attendanceGroupBtn:hover{outline:1px solid #b9dcc8;background:#f7fcf9}
@media(max-width:1000px){.rxClientSummary,.attendanceGroupGrid{grid-template-columns:1fr 1fr}.top{grid-template-columns:1fr!important}.brand{justify-content:center!important;text-align:center!important}.quickNav{text-align:center!important}.sectionButtons{justify-content:center!important}}
@media(max-width:650px){.rxClientSummary,.attendanceGroupGrid{grid-template-columns:1fr}.hero .wrap{padding:14px!important}.brand h1{font-size:18px!important}}
</style>
"""


JS = r"""
<script id="ajuste16-cliques-rx-cliente">
(function(){
  const TODAY_ISO='2026-05-11';
  const safe=v=>String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
  const clientKey=r=>r['CPF/CNPJ']||r['Cliente Nome']||r['Cliente chave']||'';
  const unique=a=>[...new Set(a.filter(Boolean))].sort();
  const clientCount=rows=>unique(rows.map(clientKey)).length;
  const attClients=()=>((DATA.atendimentos||{}).clientes||[]);
  const attPlan=()=>((DATA.atendimentos||{}).plano||[]);
  const currentRows=()=>lastRows&&lastRows.length?lastRows:filtered();
  const currentClientKeys=()=>new Set(currentRows().map(r=>String(r['Cliente chave']||'')).filter(Boolean));
  const daysFrom=d=>{
    if(!d)return '';
    const p=String(d).split('/');
    const iso=p.length===3?`${p[2]}-${p[1]}-${p[0]}`:String(d).slice(0,10);
    const diff=(new Date(TODAY_ISO+'T00:00:00')-new Date(iso+'T00:00:00'))/86400000;
    return Number.isFinite(diff)?Math.floor(diff):'';
  };
  function ensureRxClient(){
    let panel=$('rxClientPanel');
    if(panel)return panel;
    panel=document.createElement('section');
    panel.id='rxClientPanel';
    panel.className='rxClientPanel';
    panel.innerHTML='<div class="rxClientHead"><h2 id="rxClientTitle">RX de cliente</h2><button class="rxClientClose" id="rxClientClose">Fechar</button></div><div class="rxClientBody" id="rxClientBody"></div>';
    const rx=$('rxPanel');
    rx.insertAdjacentElement('afterend',panel);
    $('rxClientClose').onclick=()=>panel.classList.remove('open');
    return panel;
  }
  function sortTable(table,idx,th){
    const tbody=table.tBodies[0]; if(!tbody)return;
    const asc=th.dataset.dir!=='asc';
    table.querySelectorAll('th').forEach(x=>{x.classList.remove('activeSort','asc','desc');delete x.dataset.dir;});
    th.dataset.dir=asc?'asc':'desc'; th.classList.add('activeSort',asc?'asc':'desc');
    const num=t=>Number(String(t).replace(/\s/g,'').replace(/[^\d,-]/g,'').replace(/\./g,'').replace(',','.'));
    [...tbody.rows].sort((a,b)=>{
      const av=a.cells[idx]?.textContent.trim()||'', bv=b.cells[idx]?.textContent.trim()||'';
      const an=num(av), bn=num(bv);
      const cmp=Number.isFinite(an)&&Number.isFinite(bn)?an-bn:av.localeCompare(bv,'pt-BR',{numeric:true,sensitivity:'base'});
      return asc?cmp:-cmp;
    }).forEach(r=>tbody.appendChild(r));
  }
  function proposalRows(title,rows){
    const panel=ensureRxClient();
    $('rxClientTitle').textContent=title;
    $('rxClientBody').innerHTML=`<div class="rxClientSummary"><div class="detail"><span>Propostas</span><strong>${fmt.format(rows.length)}</strong></div><div class="detail"><span>Clientes</span><strong>${fmt.format(clientCount(rows))}</strong></div><div class="detail"><span>Em aberto</span><strong>${fmt.format(rows.filter(r=>r['Status painel']==='Em aberto').length)}</strong></div><div class="detail"><span>Valor</span><strong>${brl.format(rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0))}</strong></div></div><div class="tableWrap"><table><thead><tr><th>Proposta</th><th>Data</th><th>Cliente</th><th>Status</th><th>Vendedor</th><th>Setor</th><th>Kit</th><th>Valor</th></tr></thead><tbody>${rows.slice(0,800).map(r=>`<tr><td><button class="linkBtn prop" data-prop="${safe(r['Proposta ID']||'')}">${safe(r['Proposta Numero']||r['Proposta ID']||'')}</button></td><td>${safe(r['Proposta Data']||'')}</td><td><button class="linkBtn client" data-client="${safe(r['Cliente chave']||'')}">${safe(r['Cliente Nome']||'')}</button></td><td><span class="pill">${safe(r['Status painel']||'')}</span></td><td>${safe(r['Nome do Vendedor']||'')}</td><td>${safe(r.Setor||'')}</td><td>${safe(r.Kit||'')}</td><td class="money">${brl.format(+r['Valor total da proposta']||0)}</td></tr>`).join('')}</tbody></table></div>`;
    panel.classList.add('open');
    panel.scrollIntoView({behavior:'smooth',block:'start'});
    $('rxClientBody').querySelectorAll('th').forEach((th,i)=>th.onclick=()=>sortTable(th.closest('table'),i,th));
    $('rxClientBody').querySelectorAll('[data-prop]').forEach(b=>b.onclick=()=>openProposalById(b.dataset.prop));
    $('rxClientBody').querySelectorAll('[data-client]').forEach(b=>{b.onclick=()=>{const r=allRows.find(x=>String(x['Cliente chave']||'')===b.dataset.client); if(r) openClientInline(r);}});
  }
  function openClientInline(r){
    const key=r['Cliente chave'];
    proposalRows('RX do cliente: '+(r['Cliente Nome']||''),[...allRows,...removed].filter(x=>x['Cliente chave']===key));
  }
  function attendanceRows(title,rows){
    const panel=ensureRxClient();
    $('rxClientTitle').textContent=title;
    $('rxClientBody').innerHTML=`<div class="rxClientSummary"><div class="detail"><span>Clientes</span><strong>${fmt.format(rows.length)}</strong></div><div class="detail"><span>Com atendimento</span><strong>${fmt.format(rows.filter(c=>+c.Qtd_atendimentos>0).length)}</strong></div><div class="detail"><span>1 vez</span><strong>${fmt.format(rows.filter(c=>+c.Qtd_atendimentos===1).length)}</strong></div><div class="detail"><span>30+ dias</span><strong>${fmt.format(rows.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30).length)}</strong></div></div><div class="tableWrap"><table><thead><tr><th>Cliente</th><th>Setor</th><th>Vendedor sugerido</th><th>Proposta aberta</th><th>Prioridade</th><th>Último contato</th><th>Dias</th><th>Atendimentos</th><th>Motivo</th></tr></thead><tbody>${rows.slice(0,800).map(c=>`<tr><td>${safe(c.Cliente||'')}</td><td>${safe(c.Setor||'')}</td><td>${safe(c.Vendedor_sugerido||'')}</td><td>${safe(c.Tem_proposta_aberta||'')}</td><td>${safe(c.Prioridade||'')}</td><td>${safe(c.Ultimo_contato||'')}</td><td>${safe(c.Dias_sem_contato)}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${safe(c.Motivo_ativacao||'')}</td></tr>`).join('')}</tbody></table></div>`;
    panel.classList.add('open');
    panel.scrollIntoView({behavior:'smooth',block:'start'});
    $('rxClientBody').querySelectorAll('th').forEach((th,i)=>th.onclick=()=>sortTable(th.closest('table'),i,th));
  }
  function makeHero(){
    const h=document.querySelector('.brand h1');
    if(h)h.innerHTML='Olá, Thalys. <strong>Foco do dia:</strong> recuperar propostas abertas e priorizar atendimentos.';
    const q=document.querySelector('.quickNav h2');
    if(q)q.textContent='Acesso rápido';
    let nav=document.querySelector('.sectionButtons');
    if(nav&&!$('goAttendance')){
      const b=document.createElement('button'); b.id='goAttendance'; b.textContent='Atendimento'; b.onclick=()=>$('attendancePanel').scrollIntoView({behavior:'smooth',block:'start'});
      const rx=$('goRx'); rx?rx.insertAdjacentElement('afterend',b):nav.appendChild(b);
    }
  }
  function moveBlocks(){
    const content=document.querySelector('.content'), rx=ensureRxClient(), set=$('setorPanel'), table=document.querySelector('.tablePanel');
    if(content&&set&&table){
      content.appendChild(set);
      content.appendChild(table);
    }
    const gp=$('groupPanel'); if(gp)gp.remove();
  }
  function groupRowsBy(key){
    const rows=currentRows();
    const map={};
    rows.forEach(r=>{const name=r[key]||'(sem informação)'; if(!map[name])map[name]=[]; map[name].push(r);});
    const body=Object.entries(map).sort((a,b)=>b[1].length-a[1].length).flatMap(([name,rs])=>rs.map(r=>({...r,_groupName:name})));
    proposalRows('Dashboard RX de cliente por '+key,body);
  }
  function wireGroupingButtons(){
    const labels={'Nome do Vendedor':'vendedor','Status painel':'situação','Tipo de agrupamento de produto':'tipo de produto','Setor':'setor','Kit':'kit'};
    document.querySelectorAll('#groupQuick button').forEach(btn=>{
      btn.onclick=()=>{
        const opt=[...$('groupBy').options].find(o=>o.textContent.replace('Agrupar por ','')===btn.textContent);
        if(opt)$('groupBy').value=opt.value;
        groupRowsBy(($('groupBy')&&$('groupBy').value)||'Nome do Vendedor');
      };
    });
  }
  function filteredAttendance(){
    const keys=currentClientKeys();
    const sector=($('sector')&&$('sector').value)||'';
    return attClients().filter(c=>(!keys.size||keys.has(String(c.Cliente_chave||'')))&&(!sector||c.Setor===sector));
  }
  function renderAttendanceGroups(rows){
    let box=$('attendanceGroups');
    if(!box){
      box=document.createElement('div'); box.id='attendanceGroups'; box.className='attendanceGroups';
      const grid=document.createElement('div'); grid.id='attendanceGroupGrid'; grid.className='attendanceGroupGrid';
      box.innerHTML='<h3>Agrupamentos de atendimento</h3>';
      box.appendChild(grid);
      const panel=$('attendancePanel'); panel.querySelector('.attendanceGrid').insertAdjacentElement('afterend',box);
    }
    const specs=[
      ['Setor',c=>c.Setor||'(sem setor)'],
      ['Vendedor',c=>c.Vendedor_sugerido||'(sem vendedor)'],
      ['Prioridade',c=>c.Prioridade||'(sem prioridade)'],
      ['Proposta aberta',c=>c.Tem_proposta_aberta||'Não informado'],
      ['Último contato',c=>{const d=daysFrom(c.Ultimo_contato);return d===''?'Sem data':d<=15?'até 15 dias':d<=30?'16 a 30 dias':d<=60?'31 a 60 dias':'+60 dias'}],
      ['Recorrência',c=>+c.Qtd_atendimentos===1?'Atendido 1 vez':'Recorrente']
    ];
    const grid=$('attendanceGroupGrid'); grid.innerHTML='';
    specs.forEach(([label,fn])=>{
      const groups={}; rows.forEach(c=>{const k=fn(c); groups[k]=(groups[k]||0)+1;});
      Object.entries(groups).sort((a,b)=>b[1]-a[1]).slice(0,3).forEach(([name,count])=>{
        const b=document.createElement('button'); b.className='attendanceGroupBtn'; b.innerHTML=`${safe(label)}: ${safe(name)}<small>${fmt.format(count)} clientes</small>`;
        b.onclick=()=>attendanceRows(`Atendimentos - ${label}: ${name}`,rows.filter(c=>fn(c)===name));
        grid.appendChild(b);
      });
    });
  }
  const oldRender=window.render;
  window.render=function(){
    oldRender();
    makeHero();
    moveBlocks();
    wireGroupingButtons();
    document.querySelectorAll('#duplicadasPanel,#auditoriaPanel').forEach(p=>{if(!p.dataset.userOpened)p.classList.remove('open')});
    const att=filteredAttendance();
    renderAttendanceGroups(att);
    const setText=(id,text)=>{const el=$(id); if(el)el.textContent=text};
    setText('attClients',fmt.format(att.length));
    setText('attWith',fmt.format(att.filter(c=>+c.Qtd_atendimentos>0).length));
    setText('attOnce',fmt.format(att.filter(c=>+c.Qtd_atendimentos===1).length));
    setText('attStale',fmt.format(att.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30).length));
    setText('attOpen',fmt.format(att.filter(c=>c.Tem_proposta_aberta==='Sim').length));
    document.querySelectorAll('.attendanceKpi').forEach((k,i)=>{k.onclick=()=>attendanceRows(['Clientes no filtro','Com atendimento','Atendidos 1 vez','Sem contato 30+ dias','Com proposta aberta'][i]||'Atendimentos',[
      att,
      att.filter(c=>+c.Qtd_atendimentos>0),
      att.filter(c=>+c.Qtd_atendimentos===1),
      att.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30),
      att.filter(c=>c.Tem_proposta_aberta==='Sim')
    ][i]||att)});
    document.querySelectorAll('#attendanceSectorBtns button').forEach(b=>{b.onclick=()=>{if($('sector'))$('sector').value=b.textContent==='Todos'?'':b.textContent; attendanceSector=b.textContent; render();}});
    const sellerList=$('sellerList');
    if(sellerList&&!sellerList.dataset.delegated){
      sellerList.dataset.delegated='1';
      sellerList.addEventListener('click',e=>{
        const row=e.target.closest('.sellerRow'); if(!row)return;
        const name=row.querySelector('.sellerName')?.textContent.trim();
        if(!name)return;
        const rows=currentRows().filter(r=>r['Nome do Vendedor']===name);
        if(e.target.classList.contains('sellerActionMini')){openSeller(name);setTimeout(()=>$('sellerActionBtn')&&$('sellerActionBtn').click(),100);return;}
        proposalRows('Propostas de '+name,rows);
      },true);
    }
    if($('detailSeller'))$('detailSeller').onclick=()=>proposalRows('RX de clientes por vendedor',currentRows());
    if($('goAttendance'))$('goAttendance').onclick=()=>$('attendancePanel').scrollIntoView({behavior:'smooth',block:'start'});
  };
  window.openListModal=proposalRows;
  makeHero();
  render();
})();
</script>
"""


def update_listagem(actions: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for item in actions:
        ws.append([now, "AJUSTE_16_CLIQUES_RX_CLIENTE", item["tipo"], item["acao"], item["contexto"], item["obs"]])
    fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 28 if col != 4 else 48
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    tab = Table(displayName="TabelaLinhaTempoAjuste16", ref=f"A1:F{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)
    wb.save(LISTAGEM)


def main() -> None:
    html = SOURCE.read_text(encoding="utf-8")
    if "ajuste16-cliques-rx-cliente" not in html:
        html = html.replace("</head>", CSS + "\n</head>")
        html = html.replace("</body>", JS + "\n</body>")
    OUT.write_text(html, encoding="utf-8")
    shutil.copy2(OUT, ROOT_DASH)
    shutil.copy2(OUT, LOTE_OUT)
    shutil.copy2(OUT, LOTE_DASH)
    actions = [
        {"tipo": "Cliques", "acao": "Corrigir cliques do RX da equipe, ação e atendimento.", "contexto": "Usuário informou que vários cliques não funcionavam.", "obs": "Abertura agora usa painel inline para baixo."},
        {"tipo": "RX cliente", "acao": "Botões de agrupamento direcionam para Dashboard RX de cliente.", "contexto": "Usuário pediu direcionar ao RX de cliente, não ao agrupamento solto.", "obs": "Bloco Agrupamento final foi removido."},
        {"tipo": "Atendimentos", "acao": "Adicionar agrupamentos em atendimentos.", "contexto": "Usuário pediu agrupamento nos atendimentos também.", "obs": "Agrupa por setor, vendedor, prioridade, proposta aberta, último contato e recorrência."},
        {"tipo": "Layout", "acao": "Reduzir faixa superior, harmonizar frase e destacar foco.", "contexto": "Usuário pediu reduzir faixa, logotipo e frases mais harmonizadas.", "obs": "Tentativa de logo.png se existir; mantém EG como fallback."},
        {"tipo": "Ordem", "acao": "Mover propostas por setor e tabela para baixo do RX de cliente.", "contexto": "Usuário pediu levar essa parte para baixo do RX do cliente.", "obs": "Listas detalhadas ficam depois do painel de aprofundamento."},
        {"tipo": "Regra atendimento", "acao": "Atendimento passa a respeitar clientes do filtro atual de propostas.", "contexto": "Usuário questionou os 2.500 clientes.", "obs": "Cards de atendimento agora usam interseção com o filtro atual quando houver chave de cliente."},
    ]
    update_listagem(actions)
    SUMMARY.write_text(json.dumps({"versao": OUT.name, "arquivos": [str(OUT), str(ROOT_DASH), str(LOTE_OUT), str(LOTE_DASH), str(LISTAGEM)], "acoes": actions}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "out": str(OUT), "root": str(ROOT_DASH), "lote": str(LOTE_DASH), "listagem": str(LISTAGEM)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
