from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "dashboard.html"
OUT = ROOT / "dashboard_AJUSTE_17_VIDEO_APONTAMENTOS.html"
ROOT_DASH = ROOT / "dashboard.html"
APRESENTACAO = ROOT / "Apresentação"
APRESENTACAO_DASH = APRESENTACAO / "01_DASHBOARD_ATUAL.html"
LOTE = ROOT / "01 - Ultimo alteração"
LOTE.mkdir(exist_ok=True)
LOTE_DASH = LOTE / "dashboard.html"
SUPPORT = ROOT / "05 - Apoio"
SUPPORT.mkdir(exist_ok=True)
SUMMARY_JSON = SUPPORT / "resumo_dashboard_ajuste_17.json"
TIMELINE = ROOT / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


CSS = r"""
<style id="ajuste17-video-css">
.hero{min-height:auto;padding:18px 0;background:linear-gradient(112deg,#0c3b27 0%,#1d7a4d 62%,#72bf32 100%)}
.hero .wrap{padding:20px 24px 18px}
.hero .top{display:grid;grid-template-columns:auto 1fr;align-items:center;gap:18px}
.brand{justify-content:flex-start;text-align:left;gap:14px}
.logo{width:54px;height:54px;border-radius:12px;font-size:30px;box-shadow:none}
.brand h1{font-size:clamp(24px,3vw,38px);line-height:1.1;margin:0;text-shadow:0 1px 0 rgba(0,0,0,.18)}
.brand h1 strong{color:#dfffd0}
.quickNav{margin-top:14px}
.quickNav h2{font-size:18px;margin-bottom:12px}
.sectionButtons{gap:10px}
.sectionButtons button{min-height:42px;padding:10px 18px}
.rxClientPanel{background:#fff;border:1px solid var(--line);border-radius:14px;box-shadow:var(--shadow);margin:14px 0;padding:0;display:none;overflow:hidden}
.rxClientPanel.open{display:block;animation:slideDown .18s ease-out}
.rxClientHead{display:flex;align-items:center;justify-content:space-between;gap:12px;border-bottom:1px solid var(--line);padding:14px 18px;background:#fff}
.rxClientHead h2{margin:0;color:var(--green)}
.rxClientClose{border:1px solid var(--line);background:#fff;border-radius:10px;padding:9px 14px;color:var(--green);font-weight:800;cursor:pointer}
.rxClientBody{padding:16px 18px}
.rxClientSummary{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:10px;margin-bottom:14px}
.rxClientSummary .detail{border:1px solid var(--line);border-radius:10px;padding:12px;background:#fbfffc}
.rxClientSummary .detail span{display:block;font-size:12px;text-transform:uppercase;font-weight:900;color:#4e6a5d}
.rxClientSummary .detail strong{display:block;margin-top:6px;font-size:22px;color:var(--green)}
.groupMiniGrid{display:grid;grid-template-columns:repeat(3,minmax(180px,1fr));gap:10px;margin-bottom:14px}
.groupMiniBtn{border:1px solid var(--line);background:#fbfffc;border-radius:10px;padding:12px;text-align:left;cursor:pointer;color:var(--green);font-weight:900}
.groupMiniBtn small{display:block;margin-top:5px;color:#63766b;font-weight:700}
.sellerList{display:block}
.sellerTableHead,.sellerRow{display:grid;grid-template-columns:minmax(180px,1.4fr) repeat(8,minmax(70px,.55fr)) minmax(78px,.5fr);gap:8px;align-items:center}
.sellerTableHead{padding:12px 14px;background:#e7f5ec;border:1px solid var(--line);border-radius:10px 10px 0 0;font-size:12px;text-transform:uppercase;font-weight:900;color:var(--green)}
.sellerRow{border:1px solid var(--line);border-top:0;padding:10px 14px;background:#fff;cursor:pointer}
.sellerRow:hover{background:#f7fffa}
.sellerName{font-weight:900;color:var(--green)}
.metricBtn,.sellerActionMini{border:1px solid transparent;background:transparent;color:var(--green);font-weight:900;text-decoration:underline;cursor:pointer;padding:5px;border-radius:8px;text-align:left}
.metricBtn:hover,.sellerActionMini:hover{border-color:var(--line);background:#eef8f1}
.sellerDetailPanel{display:none;margin-top:14px;border:1px solid var(--line);border-radius:14px;background:#fff;overflow:hidden}
.sellerDetailPanel.open{display:block;animation:slideDown .18s ease-out}
.sellerDetailHead{display:flex;align-items:center;justify-content:space-between;gap:10px;padding:14px 18px;border-bottom:1px solid var(--line)}
.sellerDetailHead h3{margin:0;color:var(--green)}
.sellerDetailBody{padding:16px 18px}
.sellerPanorama{display:grid;grid-template-columns:repeat(4,minmax(130px,1fr));gap:10px;margin-bottom:12px}
.detail.clickable{border:1px solid var(--line);background:#fbfffc;border-radius:10px;padding:12px;text-align:left;cursor:pointer;color:var(--green)}
.detail.clickable:hover{background:#edf8f1}
.detail.clickable span{display:block;font-size:12px;text-transform:uppercase;font-weight:900;color:#526c60}
.detail.clickable strong{display:block;margin:4px 0;font-size:22px;color:var(--green)}
.actionReport{display:none;margin-top:12px;border:1px solid var(--line);border-radius:12px;overflow:hidden}
.actionReport.open{display:block;animation:slideDown .18s ease-out}
.actionReport h3{margin:0;padding:12px 14px;background:#e7f5ec;color:var(--green)}
.attendanceGroups{margin-top:14px;border:1px solid var(--line);border-radius:12px;padding:14px;background:#fff}
.attendanceGroupGrid{display:grid;grid-template-columns:repeat(3,minmax(180px,1fr));gap:10px}
.attendanceGroupBtn{border:1px solid var(--line);border-radius:10px;background:#fbfffc;color:var(--green);padding:10px 12px;text-align:left;font-weight:900;cursor:pointer}
.attendanceGroupBtn small{display:block;color:#62776d;font-weight:700;margin-top:4px}
.attendanceUnified{margin-top:14px;border:1px solid var(--line);border-radius:12px;overflow:hidden;background:#fff}
.attendanceUnified h3{margin:0;padding:12px 16px;background:#e7f5ec;color:var(--green)}
.kpi.clickable,.barrow.clickable,.attendanceKpi{cursor:pointer}
.kpi.clickable:hover,.barrow.clickable:hover,.attendanceKpi:hover{outline:2px solid rgba(31,122,77,.18)}
th{cursor:pointer;user-select:none}
th.activeSort{background:#c8e7c7!important}
th.asc::after{content:" ▲"}
th.desc::after{content:" ▼"}
#groupPanel{display:none!important}
#duplicadasPanel:not(.open),#auditoriaPanel:not(.open){display:none!important}
.summaryKpis.openMode{grid-template-columns:repeat(3,1fr)}
.summaryKpis.openMode .hideOpenMode{display:none}
@media(max-width:1100px){.sellerTableHead,.sellerRow{grid-template-columns:minmax(160px,1fr) repeat(4,80px)}.sellerTableHead span:nth-child(n+6):nth-child(-n+9),.sellerRow button:nth-of-type(n+5):nth-of-type(-n+8){display:none}.sellerPanorama,.rxClientSummary,.groupMiniGrid,.attendanceGroupGrid{grid-template-columns:1fr 1fr}}
@media(max-width:720px){.hero .top{grid-template-columns:1fr}.sellerTableHead,.sellerRow{grid-template-columns:1fr 80px 80px}.sellerTableHead span:nth-child(n+4),.sellerRow button:nth-of-type(n+3){display:none}.sellerPanorama,.rxClientSummary,.groupMiniGrid,.attendanceGroupGrid{grid-template-columns:1fr}}
</style>
"""


JS = r"""
<script id="ajuste17-video-apontamentos">
(function(){
  window.__aj17errors=[];
  window.addEventListener('error',function(e){
    window.__aj17errors.push((e.message||'erro')+' @ '+(e.lineno||'')+':'+(e.colno||''));
    try{
      let box=document.getElementById('aj17ErrorBox');
      if(!box){box=document.createElement('pre');box.id='aj17ErrorBox';box.style.cssText='position:fixed;left:8px;bottom:8px;z-index:9999;background:#fff1f1;color:#7b1111;border:1px solid #e1aaaa;padding:8px;max-width:520px;white-space:pre-wrap;font:12px monospace';document.body.appendChild(box);}
      box.textContent=window.__aj17errors.join('\n');
    }catch(_){}
  });
  const TODAY_ISO='2026-05-11';
  const TODAY_BR='11/05/2026';
  const $=id=>document.getElementById(id);
  const safe=v=>String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
  const uniq=a=>[...new Set((a||[]).filter(v=>v!==undefined&&v!==null&&String(v).trim()!==''))];
  const txt=v=>String(v??'').trim();
  const rowsAll=()=>DATA.oportunidades||[];
  const removedAll=()=>DATA.retiradas||[];
  const attClients=()=>((DATA.atendimentos||{}).clientes||[]);
  const attPlan=()=>((DATA.atendimentos||{}).plano||[]);
  const rowKey=r=>txt(r['Cliente chave']||r.Cliente_chave||r['CPF/CNPJ']||r.CPF_CNPJ||r['Cliente Nome']||r.Cliente);
  const clientCount=rows=>uniq((rows||[]).map(rowKey)).length;
  const groupRows=(rows,key)=>{const m={};(rows||[]).forEach(r=>{const k=txt(r[key])||'(sem informacao)';(m[k]||(m[k]=[])).push(r)});return m};
  const money=n=>brl.format(+n||0);
  const daysFrom=d=>{
    if(!d)return '';
    const s=String(d).slice(0,10);
    let iso=s;
    const p=s.split('/');
    if(p.length===3) iso=`${p[2]}-${p[1]}-${p[0]}`;
    const diff=(new Date(TODAY_ISO+'T00:00:00')-new Date(iso+'T00:00:00'))/86400000;
    return Number.isFinite(diff)?Math.max(0,Math.floor(diff)):'';
  };
  const attMap=new Map();
  attClients().forEach(c=>attMap.set(rowKey(c),c));

  function inPeriod17(row){
    const iso=txt(row['Data ISO']);
    if(!iso) return true;
    const periodValue=$('period')?$('period').value:'all';
    const mp=$('monthPick');
    if(mp&&mp.value&&periodValue==='all'&&iso.slice(0,7)!==mp.value) return false;
    if($('from')&&$('from').value&&iso<$('from').value) return false;
    if($('to')&&$('to').value&&iso>$('to').value) return false;
    const p=periodValue;
    if(/^\d{4}-\d{2}$/.test(p)) return iso.slice(0,7)===p;
    if(p==='all'||p==='month') return true;
    if(p==='today') return iso===TODAY_ISO;
    if(p==='year') return iso.slice(0,4)==='2026';
    if(p==='quarter') return iso>='2026-04'&&iso<='2026-06';
    if(p==='week') return row.Semana==='2026-W20'||row.Semana==='2026-W19';
    return true;
  }

  window.filtered=function(){
    const q=($('q')?$('q').value:'').toLowerCase().trim();
    const sector=$('sector')?$('sector').value:'';
    const vendor=$('vendor')?$('vendor').value:'';
    const status=$('status')?$('status').value:'';
    const uf=$('uf')?$('uf').value:'';
    const city=$('city')?$('city').value:'';
    const ptype=$('ptype')?$('ptype').value:'';
    return rowsAll().filter(r=>{
      if(mode==='open'&&r['Status painel']!=='Em aberto')return false;
      if(!inPeriod17(r))return false;
      if(sector&&r.Setor!==sector)return false;
      if(vendor&&r['Nome do Vendedor']!==vendor)return false;
      if(status&&r['Status painel']!==status)return false;
      if(uf&&r['Cliente UF']!==uf)return false;
      if(city&&r['Cliente Municipio']!==city)return false;
      if(ptype&&r['Tipo de agrupamento de produto']!==ptype)return false;
      if(q&&!JSON.stringify(r).toLowerCase().includes(q))return false;
      return true;
    });
  };

  function ensurePanel(){
    let p=$('rxClientPanel');
    if(p)return p;
    p=document.createElement('section');
    p.id='rxClientPanel';
    p.className='rxClientPanel';
    p.innerHTML='<div class="rxClientHead"><h2 id="rxClientTitle">Dashboard RX de cliente</h2><button class="rxClientClose" id="rxClientClose">Fechar</button></div><div class="rxClientBody" id="rxClientBody"></div>';
    const rx=$('rxPanel')||document.querySelector('.content');
    rx.insertAdjacentElement('afterend',p);
    $('rxClientClose').onclick=()=>p.classList.remove('open');
    return p;
  }

  function sortTable(table,idx,th){
    const body=table&&table.tBodies&&table.tBodies[0]; if(!body)return;
    const asc=th.dataset.dir!=='asc';
    table.querySelectorAll('th').forEach(h=>{h.classList.remove('activeSort','asc','desc');delete h.dataset.dir});
    th.dataset.dir=asc?'asc':'desc'; th.classList.add('activeSort',asc?'asc':'desc');
    const toNum=t=>Number(String(t).replace(/\s/g,'').replace(/[^\d,-]/g,'').replace(/\./g,'').replace(',','.'));
    [...body.rows].sort((a,b)=>{
      const av=a.cells[idx]?.textContent.trim()||'', bv=b.cells[idx]?.textContent.trim()||'';
      const an=toNum(av), bn=toNum(bv);
      const cmp=Number.isFinite(an)&&Number.isFinite(bn)?an-bn:av.localeCompare(bv,'pt-BR',{numeric:true,sensitivity:'base'});
      return asc?cmp:-cmp;
    }).forEach(r=>body.appendChild(r));
  }
  function enhanceTables(scope=document){
    scope.querySelectorAll('table').forEach(t=>t.querySelectorAll('th').forEach((th,i)=>th.onclick=()=>sortTable(t,i,th)));
  }

  function proposalTable(title,rows){
    rows=rows||[];
    const panel=ensurePanel();
    $('rxClientTitle').textContent=title;
    const value=rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0);
    panel.querySelector('.rxClientBody').innerHTML=`<div class="rxClientSummary">
      <div class="detail"><span>Propostas</span><strong>${fmt.format(rows.length)}</strong></div>
      <div class="detail"><span>Clientes</span><strong>${fmt.format(clientCount(rows))}</strong></div>
      <div class="detail"><span>Em aberto</span><strong>${fmt.format(rows.filter(r=>r['Status painel']==='Em aberto').length)}</strong></div>
      <div class="detail"><span>Valor</span><strong>${money(value)}</strong></div>
      </div><div class="tableWrap"><table><thead><tr><th>Proposta</th><th>Data</th><th>Cliente</th><th>Status</th><th>Vendedor</th><th>Setor</th><th>Kit</th><th>Valor</th></tr></thead><tbody>
      ${rows.slice(0,1000).map(r=>`<tr><td><button class="linkBtn" data-prop="${safe(r['Proposta ID']||'')}">${safe(r['Proposta Numero']||r['Proposta ID']||'')}</button></td><td>${safe(r['Proposta Data']||'')}</td><td><button class="linkBtn" data-client="${safe(rowKey(r))}">${safe(r['Cliente Nome']||'')}</button></td><td><span class="pill">${safe(r['Status painel']||'')}</span></td><td>${safe(r['Nome do Vendedor']||'')}</td><td>${safe(r.Setor||'')}</td><td>${safe(r.Kit||'')}</td><td class="money">${money(r['Valor total da proposta'])}</td></tr>`).join('')}
      </tbody></table></div>`;
    panel.classList.add('open');
    enhanceTables(panel);
    panel.querySelectorAll('[data-prop]').forEach(b=>b.onclick=()=>openProposalById(b.dataset.prop));
    panel.querySelectorAll('[data-client]').forEach(b=>b.onclick=()=>proposalTable('RX do cliente', [...rowsAll(),...removedAll()].filter(r=>rowKey(r)===b.dataset.client)));
    panel.scrollIntoView({behavior:'smooth',block:'start'});
  }
  window.openListModal=proposalTable;

  function groupPanel(key){
    const rows=filtered();
    const groups=Object.entries(groupRows(rows,key)).map(([name,rs])=>({name,rows:rs,value:rs.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0),clients:clientCount(rs)})).sort((a,b)=>b.rows.length-a.rows.length);
    const allGrouped=groups.flatMap(g=>g.rows.map(r=>Object.assign({_grupo:g.name},r)));
    const panel=ensurePanel();
    $('rxClientTitle').textContent='Dashboard RX de cliente por '+key;
    panel.querySelector('.rxClientBody').innerHTML=`<div class="groupMiniGrid">${groups.slice(0,12).map(g=>`<button class="groupMiniBtn" data-group="${safe(g.name)}">${safe(g.name)}<small>${fmt.format(g.rows.length)} propostas | ${fmt.format(g.clients)} clientes | ${money(g.value)}</small></button>`).join('')}</div>`;
    panel.classList.add('open');
    panel.querySelectorAll('[data-group]').forEach(b=>b.onclick=()=>proposalTable(`${key}: ${b.dataset.group}`, groups.find(g=>g.name===b.dataset.group)?.rows||[]));
    panel.querySelector('.rxClientBody').insertAdjacentHTML('beforeend','<div id="groupTableHost"></div>');
    proposalTable('Dashboard RX de cliente por '+key, allGrouped);
  }

  function currentKeys(){
    return new Set(filtered().map(rowKey).filter(Boolean));
  }
  function filteredAttendance(){
    const keys=currentKeys();
    const sec=$('sector')?$('sector').value:'';
    const q=($('q')?$('q').value:'').toLowerCase().trim();
    return attClients().filter(c=>{
      if(sec&&c.Setor!==sec)return false;
      if(keys.size&&keys.has(rowKey(c))===false)return false;
      if(q&&!JSON.stringify(c).toLowerCase().includes(q))return false;
      return true;
    });
  }
  function attendanceTable(title,rows){
    rows=rows||[];
    const panel=ensurePanel();
    $('rxClientTitle').textContent=title;
    panel.querySelector('.rxClientBody').innerHTML=`<div class="rxClientSummary">
      <div class="detail"><span>Clientes</span><strong>${fmt.format(rows.length)}</strong></div>
      <div class="detail"><span>Com atendimento</span><strong>${fmt.format(rows.filter(c=>+c.Qtd_atendimentos>0).length)}</strong></div>
      <div class="detail"><span>Atendidos 1 vez</span><strong>${fmt.format(rows.filter(c=>+c.Qtd_atendimentos===1).length)}</strong></div>
      <div class="detail"><span>Sem contato 30+ dias</span><strong>${fmt.format(rows.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30).length)}</strong></div>
      </div><div class="tableWrap"><table><thead><tr><th>Cliente</th><th>Setor</th><th>Vendedor sugerido</th><th>Proposta aberta</th><th>Prioridade</th><th>Ultimo contato</th><th>Dias</th><th>Atendimentos</th><th>Motivo</th></tr></thead><tbody>
      ${rows.slice(0,1000).map(c=>`<tr><td>${safe(c.Cliente||'')}</td><td>${safe(c.Setor||'')}</td><td>${safe(c.Vendedor_sugerido||'')}</td><td>${safe(c.Tem_proposta_aberta||'')}</td><td>${safe(c.Prioridade||'')}</td><td>${safe(c.Ultimo_contato||'')}</td><td>${safe(c.Dias_sem_contato)}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${safe(c.Motivo_ativacao||'')}</td></tr>`).join('')}
      </tbody></table></div>`;
    panel.classList.add('open');
    enhanceTables(panel);
    panel.scrollIntoView({behavior:'smooth',block:'start'});
  }

  function sellerStats(rows){
    const open=rows.filter(r=>r['Status painel']==='Em aberto');
    const done=rows.filter(r=>r['Status painel']==='Concluídas');
    const notok=rows.filter(r=>r['Status painel']==='Não aprovadas');
    const openClients=uniq(open.map(rowKey)).map(k=>attMap.get(k)).filter(Boolean);
    const bucket=list=>({
      d15:list.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=15).length,
      d30:list.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=30).length,
      d60:list.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=60).length,
      over60:list.filter(c=>daysFrom(c.Ultimo_contato)===''||daysFrom(c.Ultimo_contato)>60).length
    });
    return {open,done,notok,openClients,...bucket(openClients),valueOpen:open.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0),valueDone:done.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0)};
  }
  function ensureSellerDetail(){
    let p=$('sellerDetailPanel');
    if(p)return p;
    p=document.createElement('div');
    p.id='sellerDetailPanel';
    p.className='sellerDetailPanel';
    $('sellerList').insertAdjacentElement('afterend',p);
    return p;
  }
  function openSellerDetail(name){
    const rows=filtered().filter(r=>r['Nome do Vendedor']===name);
    const st=sellerStats(rows);
    const plan=st.open.slice().sort((a,b)=>{
      const ad=daysFrom((attMap.get(rowKey(a))||{}).Ultimo_contato), bd=daysFrom((attMap.get(rowKey(b))||{}).Ultimo_contato);
      return (bd===''?999:bd)-(ad===''?999:ad)||(+b['Valor total da proposta']||0)-(+a['Valor total da proposta']||0);
    });
    const p=ensureSellerDetail();
    p.innerHTML=`<div class="sellerDetailHead"><div><h3>${safe(name||'Sem vendedor')}</h3><p class="muted">Panorama em ${TODAY_BR}. Clique em qualquer numero para abrir a lista.</p></div><button class="rxClientClose" id="closeSellerDetail">Fechar</button></div>
      <div class="sellerDetailBody"><div class="sellerPanorama">
        <button class="detail clickable" id="detOpen"><span>Em aberto</span><strong>${fmt.format(st.open.length)}</strong>${money(st.valueOpen)}</button>
        <button class="detail clickable" id="detClients"><span>Clientes abertos</span><strong>${fmt.format(clientCount(st.open))}</strong>${fmt.format(st.openClients.length)} com atendimento</button>
        <button class="detail clickable" id="detDone"><span>Concluidas</span><strong>${fmt.format(st.done.length)}</strong>${money(st.valueDone)}</button>
        <button class="detail clickable" id="detNA"><span>Nao aprovadas</span><strong>${fmt.format(st.notok.length)}</strong>propostas</button>
        <button class="detail clickable" id="det15"><span>Atendimento ate 15 dias</span><strong>${fmt.format(st.d15)}</strong>clientes</button>
        <button class="detail clickable" id="det30"><span>Atendimento ate 30 dias</span><strong>${fmt.format(st.d30)}</strong>clientes</button>
        <button class="detail clickable" id="det60"><span>Atendimento ate 60 dias</span><strong>${fmt.format(st.d60)}</strong>clientes</button>
        <button class="detail clickable" id="detOver"><span>Sem contato ou +60 dias</span><strong>${fmt.format(st.over60)}</strong>clientes</button>
      </div><button class="smallBtn" id="sellerActionBtn">Acao proposta</button>
      <div class="actionReport" id="sellerActionReport"><h3>Plano de acao de atendimento</h3><div class="tableWrap"><table><thead><tr><th>Prioridade</th><th>Proposta</th><th>Cliente</th><th>Valor</th><th>Ultimo contato</th><th>Dias</th><th>Acao sugerida</th></tr></thead><tbody>${plan.slice(0,200).map(r=>{const a=attMap.get(rowKey(r))||{};const d=daysFrom(a.Ultimo_contato);const pri=d===''?'Sem historico':d>60?'Mais de 60 dias':d>30?'31 a 60 dias':d>15?'16 a 30 dias':'Ate 15 dias';const ac=d===''||d>60?'Retomar contato hoje e validar interesse':d>30?'Reativar proposta e confirmar pendencias':d>15?'Fazer follow-up de decisao':'Manter acompanhamento proximo';return `<tr><td>${safe(pri)}</td><td>${safe(r['Proposta Numero']||r['Proposta ID']||'')}</td><td>${safe(r['Cliente Nome']||'')}</td><td class="money">${money(r['Valor total da proposta'])}</td><td>${safe(a.Ultimo_contato||'')}</td><td>${safe(d)}</td><td>${safe(ac)}</td></tr>`}).join('')}</tbody></table></div></div></div>`;
    p.classList.add('open');
    $('closeSellerDetail').onclick=()=>p.classList.remove('open');
    $('detOpen').onclick=()=>proposalTable('Em aberto - '+name,st.open);
    $('detClients').onclick=()=>proposalTable('Clientes abertos - '+name,st.open);
    $('detDone').onclick=()=>proposalTable('Concluidas - '+name,st.done);
    $('detNA').onclick=()=>proposalTable('Nao aprovadas - '+name,st.notok);
    $('det15').onclick=()=>attendanceTable('Atendimento ate 15 dias - '+name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=15));
    $('det30').onclick=()=>attendanceTable('Atendimento ate 30 dias - '+name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=30));
    $('det60').onclick=()=>attendanceTable('Atendimento ate 60 dias - '+name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=60));
    $('detOver').onclick=()=>attendanceTable('Sem contato ou +60 dias - '+name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)===''||daysFrom(c.Ultimo_contato)>60));
    $('sellerActionBtn').onclick=()=>{$('sellerActionReport').classList.toggle('open');enhanceTables($('sellerActionReport'))};
    enhanceTables(p);
    p.scrollIntoView({behavior:'smooth',block:'start'});
  }
  window.openSeller=openSellerDetail;

  function renderSellerTable(rows){
    const el=$('sellerList'); if(!el)return;
    const groups=Object.entries(groupRows(rows,'Nome do Vendedor')).map(([name,rs])=>({name,rows:rs,stats:sellerStats(rs)})).sort((a,b)=>b.stats.open.length-a.stats.open.length||b.rows.length-a.rows.length);
    el.innerHTML='<div class="sellerTableHead"><span>Vendedor</span><span>Propostas</span><span>Clientes</span><span>Abertas</span><span>Valor aberto</span><span>Atend.</span><span>15d</span><span>30d</span><span>+60d</span><span>Acao</span></div>';
    groups.slice(0,24).forEach(g=>{
      const st=g.stats;
      const line=document.createElement('div');
      line.className='sellerRow';
      line.innerHTML=`<span class="sellerName">${safe(g.name||'Sem vendedor')}</span><button class="metricBtn">${fmt.format(g.rows.length)}</button><button class="metricBtn">${fmt.format(clientCount(g.rows))}</button><button class="metricBtn">${fmt.format(st.open.length)}</button><button class="metricBtn">${money(st.valueOpen)}</button><button class="metricBtn">${fmt.format(st.openClients.length)}</button><button class="metricBtn">${fmt.format(st.d15)}</button><button class="metricBtn">${fmt.format(st.d30)}</button><button class="metricBtn">${fmt.format(st.over60)}</button><button class="sellerActionMini">Acao</button>`;
      line.onclick=()=>openSellerDetail(g.name);
      const b=line.querySelectorAll('button');
      b[0].onclick=e=>{e.stopPropagation();proposalTable('Propostas - '+g.name,g.rows)};
      b[1].onclick=e=>{e.stopPropagation();proposalTable('Clientes/propostas - '+g.name,g.rows)};
      b[2].onclick=e=>{e.stopPropagation();proposalTable('Em aberto - '+g.name,st.open)};
      b[3].onclick=e=>{e.stopPropagation();proposalTable('Valor em aberto - '+g.name,st.open)};
      b[4].onclick=e=>{e.stopPropagation();attendanceTable('Atendimentos - '+g.name,st.openClients)};
      b[5].onclick=e=>{e.stopPropagation();attendanceTable('Atendimento ate 15 dias - '+g.name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=15))};
      b[6].onclick=e=>{e.stopPropagation();attendanceTable('Atendimento ate 30 dias - '+g.name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)!==''&&daysFrom(c.Ultimo_contato)<=30))};
      b[7].onclick=e=>{e.stopPropagation();attendanceTable('Sem contato ou +60 dias - '+g.name,st.openClients.filter(c=>daysFrom(c.Ultimo_contato)===''||daysFrom(c.Ultimo_contato)>60))};
      b[8].onclick=e=>{e.stopPropagation();openSellerDetail(g.name);setTimeout(()=>{$('sellerActionBtn')&&$('sellerActionBtn').click()},100)};
      el.appendChild(line);
    });
    ensureSellerDetail();
  }

  function setupQuickFilters(){
    const period=$('period');
    if(period&&!period.dataset.aj17Months){
      const months=uniq(rowsAll().map(r=>r.AnoMes)).sort().reverse();
      period.innerHTML='<option value="all">Todo período</option>'+months.map(m=>`<option value="${safe(m)}" ${m===DATA.current_month?'selected':''}>${safe(m)}</option>`).join('');
      period.dataset.aj17Months='1';
      period.oninput=()=>render();
    }
    const mp=$('monthPick');
    if(mp&&!mp.dataset.aj17){
      const months=uniq(rowsAll().map(r=>r.AnoMes)).sort().reverse();
      mp.innerHTML='<option value="">Todo periodo</option>'+months.map(m=>`<option value="${safe(m)}" ${m===DATA.current_month?'selected':''}>${safe(m)}</option>`).join('');
      mp.dataset.aj17='1';
      mp.oninput=()=>render();
    }
    const sq=$('sectorQuick');
    if(sq){
      const sectors=['',...uniq(rowsAll().map(r=>r.Setor)).sort()];
      sq.innerHTML='';
      sectors.forEach(v=>{
        const b=document.createElement('button'); b.type='button'; b.textContent=v||'Todos'; b.className=(($('sector')&&$('sector').value===v)?'active':'');
        b.onclick=()=>{if($('sector'))$('sector').value=v; render();};
        sq.appendChild(b);
      });
    }
    const gq=$('groupQuick');
    if(gq){
      const opts=[['Nome do Vendedor','vendedor'],['Status painel','situacao'],['Tipo de agrupamento de produto','tipo de produto'],['Setor','setor'],['Kit','kit']];
      gq.innerHTML='';
      opts.forEach(([key,label])=>{
        const b=document.createElement('button'); b.type='button'; b.textContent=label; b.className=(($('groupBy')&&$('groupBy').value===key)?'active':'');
        b.onclick=()=>{if($('groupBy'))$('groupBy').value=key; groupPanel(key); setupQuickFilters();};
        gq.appendChild(b);
      });
    }
  }

  function setupAttendance(rows){
    const att=filteredAttendance();
    const setText=(id,value)=>{const el=$(id); if(el)el.textContent=fmt.format(value)};
    setText('attClients',att.length);
    setText('attWith',att.filter(c=>+c.Qtd_atendimentos>0).length);
    setText('attOnce',att.filter(c=>+c.Qtd_atendimentos===1).length);
    setText('attStale',att.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30).length);
    setText('attOpen',att.filter(c=>c.Tem_proposta_aberta==='Sim').length);
    const labels=['Clientes no filtro','Com atendimento','Atendidos 1 vez','Sem contato 30+ dias','Com proposta aberta'];
    const lists=[att,att.filter(c=>+c.Qtd_atendimentos>0),att.filter(c=>+c.Qtd_atendimentos===1),att.filter(c=>c.Dias_sem_contato===''||+c.Dias_sem_contato>=30),att.filter(c=>c.Tem_proposta_aberta==='Sim')];
    document.querySelectorAll('.attendanceKpi').forEach((k,i)=>k.onclick=()=>attendanceTable(labels[i]||'Atendimentos',lists[i]||att));
    const btns=$('attendanceSectorBtns');
    if(btns){
      ['Todos','Tráfego','Carteira','ESM'].forEach(name=>{
        const b=[...btns.querySelectorAll('button')].find(x=>x.textContent.trim()===name);
        if(b){b.className=((($('sector')&&$('sector').value)||'Todos')===name||(!$('sector').value&&name==='Todos'))?'active':'';b.onclick=()=>{if($('sector'))$('sector').value=name==='Todos'?'':name;render();};}
      });
    }
    let groups=$('attendanceGroups');
    if(!groups){
      groups=document.createElement('div'); groups.id='attendanceGroups'; groups.className='attendanceGroups';
      groups.innerHTML='<h3>Agrupamentos de atendimento</h3><div class="attendanceGroupGrid" id="attendanceGroupGrid"></div>';
      const grid=document.querySelector('.attendanceGrid');
      grid&&grid.insertAdjacentElement('afterend',groups);
    }
    const specs=[
      ['Setor',c=>c.Setor||'(sem setor)'],
      ['Vendedor',c=>c.Vendedor_sugerido||'(sem vendedor)'],
      ['Prioridade',c=>c.Prioridade||'(sem prioridade)'],
      ['Proposta aberta',c=>c.Tem_proposta_aberta||'Nao informado'],
      ['Ultimo contato',c=>{const d=daysFrom(c.Ultimo_contato);return d===''?'Sem data':d<=15?'ate 15 dias':d<=30?'16 a 30 dias':d<=60?'31 a 60 dias':'+60 dias'}],
      ['Recorrencia',c=>+c.Qtd_atendimentos===1?'Atendido 1 vez':'Recorrente']
    ];
    const grid=$('attendanceGroupGrid');
    if(grid){
      grid.innerHTML='';
      specs.forEach(([label,fn])=>{
        const m={};att.forEach(c=>{const k=fn(c);(m[k]||(m[k]=[])).push(c)});
        Object.entries(m).sort((a,b)=>b[1].length-a[1].length).slice(0,3).forEach(([name,list])=>{
          const b=document.createElement('button'); b.className='attendanceGroupBtn'; b.innerHTML=`${safe(label)}: ${safe(name)}<small>${fmt.format(list.length)} clientes</small>`; b.onclick=()=>attendanceTable(`Atendimentos - ${label}: ${name}`,list); grid.appendChild(b);
        });
      });
    }
    let unified=$('attendanceUnified');
    if(!unified){
      unified=document.createElement('div'); unified.id='attendanceUnified'; unified.className='attendanceUnified';
      const cols=document.querySelector('.activationColumns')||$('attendanceGroups')||$('attendancePanel');
      cols&&cols.insertAdjacentElement('afterend',unified);
    }
    const plan=attPlan().filter(c=>!($('sector')&&$('sector').value)||c.Setor===$('sector').value).filter(c=>!currentKeys().size||currentKeys().has(rowKey(c))).slice(0,500);
    if(unified){
      unified.innerHTML=`<h3>Fila unica de atendimento</h3><div class="tableWrap"><table><thead><tr><th>Cliente</th><th>Setor</th><th>Vendedor sugerido</th><th>Proposta aberta</th><th>Prioridade</th><th>Ultimo contato</th><th>Atendimentos</th><th>Motivo</th></tr></thead><tbody>${plan.map(c=>`<tr><td>${safe(c.Cliente||'')}</td><td>${safe(c.Setor||'')}</td><td>${safe(c.Vendedor_sugerido||'')}</td><td>${safe(c.Tem_proposta_aberta||'')}</td><td>${safe(c.Prioridade||'')}</td><td>${safe(c.Ultimo_contato||'')}</td><td>${fmt.format(c.Qtd_atendimentos||0)}</td><td>${safe(c.Motivo_ativacao||'')}</td></tr>`).join('')}</tbody></table></div>`;
      enhanceTables(unified);
    }
  }

  function moveBlocks(){
    const content=document.querySelector('.content');
    const set=$('setorPanel'), table=document.querySelector('.tablePanel');
    if(content&&set)content.appendChild(set);
    if(content&&table)content.appendChild(table);
  }
  function setupHero(){
    const logo=document.querySelector('.logo'); if(logo)logo.textContent='E';
    const h=document.querySelector('.brand h1'); if(h)h.innerHTML='Olá, Thalys. <strong>Foco do dia:</strong> propostas abertas e atendimentos.';
    const sub=document.querySelector('.quickNav h2'); if(sub)sub.textContent='Acesso rapido';
    const nav=document.querySelector('.sectionButtons');
    if(nav&&!$('goAttendance')){
      let b=[...nav.querySelectorAll('button')].find(x=>x.textContent.trim()==='Atendimento');
      if(!b){b=document.createElement('button'); nav.appendChild(b);}
      b.id='goAttendance'; b.textContent='Atendimento'; b.classList.remove('disabled');
      b.onclick=()=>$('attendancePanel').scrollIntoView({behavior:'smooth',block:'start'});
    }
    const actions=document.querySelector('.actions');
    if(actions){
      let footer=document.querySelector('.dashboardFooter');
      if(!footer){
        footer=document.createElement('footer');
        footer.className='dashboardFooter';
        footer.innerHTML='<div class="footerBox"><strong>Relatórios e saída</strong><div class="footerButtons"></div></div>';
        document.body.insertBefore(footer, document.querySelector('#modal')||null);
      }
      const target=footer.querySelector('.footerButtons')||footer;
      [...actions.children].forEach(n=>target.appendChild(n));
      actions.remove();
    }
    if($('goOpen'))$('goOpen').onclick=()=>{$('modeOpen').click();$('summaryFilters').scrollIntoView({behavior:'smooth',block:'start'})};
    if($('goRx'))$('goRx').onclick=()=>$('rxPanel').scrollIntoView({behavior:'smooth',block:'start'});
    if($('goAttendance'))$('goAttendance').onclick=()=>$('attendancePanel').scrollIntoView({behavior:'smooth',block:'start'});
  }

  function setupClicks(rows){
    document.querySelectorAll('.kpi').forEach(k=>k.classList.add('clickable'));
    const bind=(id,title,list)=>{const el=$(id); if(el&&el.closest('.kpi'))el.closest('.kpi').onclick=()=>proposalTable(title,list())};
    bind('kProps','Propostas filtradas',()=>filtered());
    bind('kOpenValue','Propostas em aberto',()=>filtered().filter(r=>r['Status painel']==='Em aberto'));
    bind('kDoneValue','Propostas concluidas',()=>filtered().filter(r=>r['Status painel']==='Concluídas'));
    bind('kClients','Clientes filtrados',()=>filtered());
    document.querySelectorAll('#statusBars .barrow').forEach(row=>{const name=row.querySelector('.barlabel')?.childNodes[0]?.textContent?.trim();row.classList.add('clickable');row.onclick=()=>proposalTable('Status: '+name,filtered().filter(r=>(r['Status painel']||'')===name))});
    document.querySelectorAll('#kitBars .barrow').forEach(row=>{const name=row.querySelector('.barlabel')?.childNodes[0]?.textContent?.trim()||row.querySelector('.barlabel')?.textContent?.trim();row.classList.add('clickable');row.onclick=()=>proposalTable('Kit: '+name,filtered().filter(r=>(r.Kit||'(sem informacao)')===name))});
    document.querySelectorAll('#sectorBars .barrow').forEach(row=>{const name=row.querySelector('.barlabel')?.childNodes[0]?.textContent?.trim()||row.querySelector('.barlabel')?.textContent?.trim();row.classList.add('clickable');row.onclick=()=>{if($('sector'))$('sector').value=name==='(sem informacao)'?'':name;render()}});
    document.querySelectorAll('button[data-close]').forEach(b=>b.onclick=()=>{const p=$(b.dataset.close); if(p){delete p.dataset.userOpened;p.classList.remove('open')}});
    document.querySelectorAll('.footerButtons button[data-panel],.sectionButtons button[data-panel]').forEach(b=>b.onclick=()=>{const p=$(b.dataset.panel); if(p){p.dataset.userOpened='1';p.classList.add('open');p.scrollIntoView({behavior:'smooth',block:'start'})}});
  }

  const baseRender=window.render;
  window.render=function(){
    if(baseRender)baseRender();
    moveBlocks();
    setupHero();
    setupQuickFilters();
    const rows=filtered();
    const value=rows.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0);
    const open=rows.filter(r=>r['Status painel']==='Em aberto');
    const done=rows.filter(r=>r['Status painel']==='Concluídas');
    if($('kProps'))$('kProps').textContent=fmt.format(rows.length);
    if($('kValue'))$('kValue').textContent=money(value);
    if($('kOpenValue'))$('kOpenValue').textContent=money(open.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0));
    if($('kDoneValue'))$('kDoneValue').textContent=money(done.reduce((a,r)=>a+(+r['Valor total da proposta']||0),0));
    if($('kClients'))$('kClients').textContent=fmt.format(clientCount(rows));
    const kpis=document.querySelector('.summaryKpis');
    if(kpis){
      kpis.classList.toggle('openMode',mode==='open');
      $('kValue')?.closest('.kpi')?.classList.toggle('hideOpenMode',mode==='open');
      $('kDoneValue')?.closest('.kpi')?.classList.toggle('hideOpenMode',mode==='open');
    }
    renderSellerTable(rows);
    setupAttendance(rows);
    setupClicks(rows);
    enhanceTables();
    document.querySelectorAll('#duplicadasPanel,#auditoriaPanel').forEach(p=>{if(!p.dataset.userOpened)p.classList.remove('open')});
  };
  render();
})();
</script>
"""


def strip_old_adjustments(html: str) -> str:
    ids = [
        "ajuste14-panorama-rx",
        "ajuste15-limpo-clicavel",
        "ajuste16-cliques-rx-cliente",
        "ajuste17-video-apontamentos",
    ]
    for script_id in ids:
        pattern = rf"\s*<script\s+id=[\"']{re.escape(script_id)}[\"'][^>]*>.*?</script>\s*"
        html = re.sub(pattern, "\n", html, flags=re.S | re.I)
        style_pattern = rf"\s*<style\s+id=[\"']{re.escape(script_id)}[\"'][^>]*>.*?</style>\s*"
        html = re.sub(style_pattern, "\n", html, flags=re.S | re.I)
    html = re.sub(r"\s*<style\s+id=[\"']ajuste17-video-css[\"'][^>]*>.*?</style>\s*", "\n", html, flags=re.S | re.I)
    return html


def write_timeline() -> None:
    rows = [
        ["Data/hora", "Ação", "Regra", "Exceção/observação"],
        [datetime.now().strftime("%d/%m/%Y %H:%M"), "AJUSTE 17 - correção pelos apontamentos em vídeo", "Removidas camadas antigas de script e criada uma camada única de interação.", "Motivo técnico: havia funções duplicadas/reescritas em ajustes 14, 15 e 16."],
        [datetime.now().strftime("%d/%m/%Y %H:%M"), "Filtros rápidos", "Setor e mês passam a alimentar o mesmo filtro usado por KPIs, RX, propostas e atendimentos.", "Agrupamentos abrem o painel Dashboard RX de cliente, não uma seção lateral antiga."],
        [datetime.now().strftime("%d/%m/%Y %H:%M"), "RX da equipe", "Todos os números clicáveis abrem lista ordenável para baixo.", "Ação proposta abre o plano dentro do próprio vendedor."],
        [datetime.now().strftime("%d/%m/%Y %H:%M"), "Atendimentos", "Cards e agrupamentos de atendimento usam os clientes do filtro atual.", "Evita mostrar 2.500 quando o dashboard está filtrado em um conjunto menor de propostas/clientes."],
        [datetime.now().strftime("%d/%m/%Y %H:%M"), "Visual", "Faixa superior reduzida, frase harmonizada, logo mantido e termos importantes destacados.", "Sem alterar os dados-base."],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in "ABCD":
        ws.column_dimensions[col].width = 34 if col != "D" else 58
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    end = ws.max_row
    tab = Table(displayName="TabelaLinhaTempo", ref=f"A1:D{end}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)
    wb.save(TIMELINE)


def main() -> None:
    html = SOURCE.read_text(encoding="utf-8", errors="replace")
    html = strip_old_adjustments(html)
    html = html.replace("</head>", CSS + "\n</head>")
    html = html.replace("</body>", JS + "\n</body>")
    OUT.write_text(html, encoding="utf-8")
    shutil.copy2(OUT, ROOT_DASH)
    shutil.copy2(OUT, APRESENTACAO_DASH)
    shutil.copy2(OUT, LOTE_DASH)
    write_timeline()
    SUMMARY_JSON.write_text(
        json.dumps(
            {
                "ok": True,
                "versao": OUT.name,
                "arquivos": [str(OUT), str(ROOT_DASH), str(APRESENTACAO_DASH), str(LOTE_DASH)],
                "diagnostico": "Scripts de ajustes 14, 15 e 16 estavam acumulados e sobrescrevendo funções/handlers entre si.",
                "correcoes": [
                    "removidas camadas antigas de script",
                    "filtros rapidos aplicados ao conjunto atual",
                    "agrupamentos direcionam para Dashboard RX de cliente",
                    "RX da equipe com cliques funcionais e abertura para baixo",
                    "atendimentos filtrados pelo contexto atual",
                    "tabelas com cabeçalhos ordenáveis",
                    "faixa superior reduzida e frase harmonizada",
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"ok": True, "out": str(OUT), "dashboard": str(ROOT_DASH), "apresentacao": str(APRESENTACAO_DASH)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
