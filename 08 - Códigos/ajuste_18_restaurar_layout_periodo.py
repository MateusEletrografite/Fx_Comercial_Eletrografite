from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "04 - Fonte" / "dashboard_AJUSTE_16_CLIQUES_RX_CLIENTE.html"
OUT = ROOT / "dashboard_AJUSTE_18_RESTAURA_LAYOUT_PERIODO.html"
ROOT_DASH = ROOT / "dashboard.html"
APRESENTACAO_DASH = ROOT / "Apresentação" / "01_DASHBOARD_ATUAL.html"
LATEST_DIR = ROOT / "01 - Ultimo alteração"
LATEST_DIR.mkdir(exist_ok=True)
LATEST_DASH = LATEST_DIR / "dashboard.html"
SUPPORT = ROOT / "05 - Apoio"
SUPPORT.mkdir(exist_ok=True)
SUMMARY = SUPPORT / "resumo_dashboard_ajuste_18.json"
TIMELINE = ROOT / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


PATCH = r"""
<style id="ajuste18-periodo-lateral-css">
#summaryFilters .summaryHead .field:has(#monthPick){display:none!important}
#summaryFilters .summaryHead{display:block}
</style>
<script id="ajuste18-restaura-layout-periodo">
(function(){
  function byId(id){return document.getElementById(id)}
  function uniq(arr){return [...new Set((arr||[]).filter(Boolean))]}
  function syncPeriodOptions(){
    const period=byId('period');
    if(!period || period.dataset.ajuste18==='1') return;
    const data=(typeof DATA!=='undefined')?DATA:{};
    const rows=data.oportunidades||[];
    const months=uniq(rows.map(r=>r.AnoMes)).sort().reverse();
    const current=data.current_month||months[0]||'';
    period.innerHTML='<option value="all">Todo período</option>'+months.map(m=>'<option value="month:'+m+'" '+(m===current?'selected':'')+'>'+m+'</option>').join('');
    period.value='month:'+current;
    period.dataset.ajuste18='1';
    period.dataset.monthsReady='1';
    period.addEventListener('input',function(){
      const mp=byId('monthPick');
      if(mp) mp.value = period.value==='all' ? '' : period.value.replace(/^month:/,'');
      if(typeof window.render==='function') window.render();
    });
    const mp=byId('monthPick');
    if(mp){
      if(!mp.querySelector('option[value=""]')) mp.insertAdjacentHTML('afterbegin','<option value="">Todo período</option>');
      mp.value=current;
    }
  }
  const oldRender=window.render;
  window.render=function(){
    syncPeriodOptions();
    if(oldRender) oldRender();
    syncPeriodOptions();
    const mp=byId('monthPick'), period=byId('period');
    if(mp && period && period.value!=='all' && mp.value!==period.value.replace(/^month:/,'')) mp.value=period.value.replace(/^month:/,'');
  };
  syncPeriodOptions();
  if(typeof window.render==='function') window.render();
})();
</script>
"""


def update_timeline() -> None:
    rows = [
        ["Data/hora", "Ação", "Regra", "Exceção/observação"],
        [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            "AJUSTE 18 - restauração de layout",
            "A versão 17 foi descartada como base visual. A versão 18 voltou para o HTML do ajuste 16.",
            "Motivo: a versão 17 mexeu demais na estrutura visual e desfez apontamentos já aprovados.",
        ],
        [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            "Período lateral",
            "O filtro Período agora mostra Todo período e os meses disponíveis, começando por 2026-05.",
            "O mês superior foi ocultado visualmente para evitar filtro duplicado.",
        ],
        [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            "Escopo",
            "Correção intencionalmente mínima: não reescreve RX, atendimento, cabeçalho, cards ou agrupamentos.",
            "Ajustes de clique devem ser tratados depois em cima desta base restaurada, não em cima da versão 17.",
        ],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    for row in rows:
        ws.append(row)
    fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in "ABCD":
        ws.column_dimensions[col].width = 38 if col != "D" else 64
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    table = Table(displayName="TabelaLinhaTempo", ref=f"A1:D{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)
    wb.save(TIMELINE)


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    html = SOURCE.read_text(encoding="utf-8", errors="replace")
    html = html.replace("</body>", PATCH + "\n</body>")
    OUT.write_text(html, encoding="utf-8")
    for dest in [ROOT_DASH, APRESENTACAO_DASH, LATEST_DASH]:
        shutil.copy2(OUT, dest)
    update_timeline()
    SUMMARY.write_text(
        json.dumps(
            {
                "ok": True,
                "versao": OUT.name,
                "base_restaurada": str(SOURCE),
                "arquivos": [str(OUT), str(ROOT_DASH), str(APRESENTACAO_DASH), str(LATEST_DASH)],
                "ajuste": "Restaurado layout da versão 16 e aplicado somente o filtro Período por mês na lateral.",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"ok": True, "out": str(OUT), "dashboard": str(ROOT_DASH)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
