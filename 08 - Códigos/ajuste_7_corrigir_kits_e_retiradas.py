from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
APOIO = LOTE / "apoio"
SRC_XLSX = LOTE / "relatorios_e_dashboard_base_AJUSTE_6_REGRAS.xlsx"
OUT_XLSX = LOTE / "relatorios_e_dashboard_base_AJUSTE_7_REGRAS.xlsx"
ROOT_XLSX = ROOT / OUT_XLSX.name
AJUSTES_XLSX = ROOT / "Ajustes 7.xlsx"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


GREEN = "107C41"
LIGHT_GREEN = "E2F0D9"
MID_GREEN = "C6E0B4"
WHITE = "FFFFFF"
GRAY = "F2F2F2"


def norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower().strip()


def unique(values: list[str]) -> list[str]:
    seen = set()
    out = []
    for value in values:
        clean = value.strip()
        if clean and clean not in seen:
            seen.add(clean)
            out.append(clean)
    return out


def explicit_kits_from_markers(markers: object) -> list[str]:
    """Use only explicit Kit tags; ignore Expositor/Mostruario as kit evidence."""
    text = "" if markers is None else str(markers)
    parts = re.split(r";|\n|\r", text)
    kits = []
    for part in parts:
        clean = norm(part)
        if "kit" not in clean:
            continue
        found = re.search(r"\bkit\s*(pp|p|m|g)\b", clean)
        if found:
            kits.append(found.group(1).upper())
    return unique(kits)


def normalize_existing_kit(value: object) -> list[str]:
    text = "" if value is None else str(value).strip()
    if not text or norm(text) in {"nan", "none"}:
        return []
    if norm(text) == "sem tag de kit":
        return ["Sem tag de kit"]
    parts = re.split(r";|\n|\r", text)
    cleaned = []
    for part in parts:
        item = part.strip().replace("Kit ", "").replace("/sem tamanho claro", "").strip()
        if item:
            cleaned.append(item)
    return unique(cleaned)


def choose_one_kit(kits: list[str]) -> str:
    priority = ["PP", "G", "M", "P"]
    for item in priority:
        if item in kits:
            return item
    return kits[0] if kits else ""


def corrected_kit(markers: object, current_kit: object) -> str:
    explicit = explicit_kits_from_markers(markers)
    if explicit:
        return choose_one_kit(explicit)
    existing = normalize_existing_kit(current_kit)
    if len(existing) > 1:
        return "Sem tag de kit"
    return "; ".join(existing) if existing else "Sem tag de kit"


def load_csv(name: str) -> pd.DataFrame:
    return pd.read_csv(APOIO / name, dtype=str, keep_default_na=False, encoding="utf-8-sig")


def save_csv(df: pd.DataFrame, name: str) -> None:
    df.to_csv(APOIO / name, index=False, encoding="utf-8-sig")


def apply_kit_fix(df: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    if "Kit" not in df.columns or "Marcadores agrupados" not in df.columns:
        return df, 0, 0
    before_multi = int(df["Kit"].astype(str).str.contains(";", regex=False, na=False).sum())
    old = df["Kit"].copy()
    df["Kit"] = [
        corrected_kit(markers, kit)
        for markers, kit in zip(df["Marcadores agrupados"], df["Kit"], strict=False)
    ]
    changed = int((old != df["Kit"]).sum())
    after_multi = int(df["Kit"].astype(str).str.contains(";", regex=False, na=False).sum())
    return df, before_multi - after_multi, changed


def enrich_with_kept_proposal(ret: pd.DataFrame, props: pd.DataFrame, all_decisions: pd.DataFrame) -> pd.DataFrame:
    lookup_cols = [
        "Proposta ID",
        "Proposta Numero",
        "Status consolidado",
        "Proposta Data",
        "Marcadores agrupados",
    ]
    lookup = props[lookup_cols].drop_duplicates("Proposta ID").set_index("Proposta ID")
    fallback = all_decisions[lookup_cols].drop_duplicates("Proposta ID").set_index("Proposta ID")
    new_cols = {
        "Proposta mantida ID": [],
        "Proposta mantida numero": [],
        "Proposta mantida status": [],
        "Proposta mantida data de abertura": [],
        "Proposta mantida marcador": [],
    }
    for kept_id in ret.get("_id_mantido", pd.Series([""] * len(ret))).astype(str):
        source = lookup if kept_id in lookup.index else fallback
        if kept_id and kept_id in source.index:
            row = source.loc[kept_id]
            new_cols["Proposta mantida ID"].append(kept_id)
            new_cols["Proposta mantida numero"].append(row["Proposta Numero"])
            new_cols["Proposta mantida status"].append(row["Status consolidado"])
            new_cols["Proposta mantida data de abertura"].append(row["Proposta Data"])
            new_cols["Proposta mantida marcador"].append(row["Marcadores agrupados"])
        else:
            for values in new_cols.values():
                values.append("")
    for col, values in new_cols.items():
        ret[col] = values
    return ret


def style_sheet(ws, table_name: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor=GREEN)
    header_font = Font(color=WHITE, bold=True)
    light_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    thin = Side(style="thin", color="D9EAD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFFFA")
    for col_idx, column_cells in enumerate(ws.columns, 1):
        values = [str(cell.value or "") for cell in column_cells[:200]]
        width = min(max([len(value) for value in values] + [10]) + 2, 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if table_name and ws.max_row > 1 and ws.max_column > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws["A1"].fill = header_fill
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row and row[0].row % 2 == 0:
            row[0].fill = light_fill


def replace_sheet_from_df(wb, sheet_name: str, df: pd.DataFrame, table_name: str) -> None:
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    style_sheet(ws, table_name)


def update_kit_cells(ws, kit_map_id: dict[str, str], kit_map_numero: dict[str, str]) -> int:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Kit" not in headers:
        return 0
    id_col = headers.index("Proposta ID") + 1 if "Proposta ID" in headers else None
    numero_col = headers.index("Proposta Numero") + 1 if "Proposta Numero" in headers else None
    if not id_col and not numero_col:
        return 0
    kit_col = headers.index("Kit") + 1
    changed = 0
    for row in range(2, ws.max_row + 1):
        proposal_id = str(ws.cell(row, id_col).value or "") if id_col else ""
        proposal_numero = str(ws.cell(row, numero_col).value or "") if numero_col else ""
        new_kit = kit_map_id.get(proposal_id) or kit_map_numero.get(proposal_numero)
        if new_kit:
            if ws.cell(row, kit_col).value != new_kit:
                ws.cell(row, kit_col).value = new_kit
                changed += 1
    return changed


def make_ajustes(summary: dict[str, object]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ajustes"
    rows = [
        ["Ajuste", "Check", "Status", "Evidencia"],
        [
            "Corrigir propostas que apareciam com dois kits por causa de marcadores Expositor/Mostruario.",
            "ckeck",
            "Concluido",
            f"{summary['kits_corrigidos_total']} linhas tiveram o Kit recalculado.",
        ],
        [
            "Adicionar em Retiradas a proposta mantida, status, data de abertura e marcador em colunas separadas.",
            "ckeck",
            "Concluido",
            f"{summary['retiradas_enriquecidas']} retiradas vinculadas a proposta mantida.",
        ],
        [
            "Salvar nova versao mantendo o nome base e alterando somente a versao.",
            "ckeck",
            "Concluido",
            OUT_XLSX.name,
        ],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws, "TabelaAjustes7")
    wb.save(AJUSTES_XLSX)


def make_listagem(summary: dict[str, object]) -> None:
    wb = Workbook()
    sheets = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao gerada", OUT_XLSX.name],
            ["Quando", summary["gerado_em"]],
            ["Por que havia 2 kits", "A regra anterior tambem lia palavras como Expositor M/G e Mostruario P como kit."],
            ["Como ficou", "Kit agora prioriza marcadores explicitos que comecam com Kit; expositor/mostruario ficam como contexto."],
            ["Retiradas", "Foram adicionadas colunas separadas da proposta mantida: ID, numero, status, data e marcador."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Kit", "Usar marcador explicito Kit PP/P/M/G quando existir."],
            ["Excecao", "Expositor e Mostruario nao definem Kit quando existe marcador Kit explicito."],
            ["Sem marcador Kit", "Preserva o Kit anterior quando nao ha marcador explicito para evitar mudanca indevida."],
            ["Retiradas", "Quando _id_mantido existe, buscar os dados na base mantida e preencher colunas separadas."],
        ],
        "Linha do tempo": [
            ["Ordem", "Acao", "Resultado"],
            [1, "Leitura do lote AJUSTE_6_REGRAS e bases CSV de apoio", "Identificada origem dos kits duplos."],
            [2, "Recalculo do campo Kit", f"{summary['kits_corrigidos_total']} linhas ajustadas nos CSVs."],
            [3, "Enriquecimento das retiradas", f"{summary['retiradas_enriquecidas']} linhas com proposta mantida encontrada."],
            [4, "Geracao do Excel final", OUT_XLSX.name],
            [5, "Copia do ultimo arquivo para a raiz", ROOT_XLSX.name],
        ],
    }
    for idx, (name, rows) in enumerate(sheets.items()):
        ws = wb.active if idx == 0 else wb.create_sheet(name)
        ws.title = name
        for row in rows:
            ws.append(row)
        style_sheet(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def main() -> None:
    props = load_csv("base_oportunidades_propostas.csv")
    ret = load_csv("base_retiradas_propostas.csv")
    all_decisions = load_csv("base_todas_propostas_com_decisao.csv")

    props, props_multi_removed, props_changed = apply_kit_fix(props)
    ret, ret_multi_removed, ret_changed = apply_kit_fix(ret)
    all_decisions, all_multi_removed, all_changed = apply_kit_fix(all_decisions)
    ret = enrich_with_kept_proposal(ret, props, all_decisions)

    save_csv(props, "base_oportunidades_propostas.csv")
    save_csv(ret, "base_retiradas_propostas.csv")
    save_csv(all_decisions, "base_todas_propostas_com_decisao.csv")

    wb = load_workbook(SRC_XLSX)
    kit_map_id = dict(zip(all_decisions["Proposta ID"].astype(str), all_decisions["Kit"].astype(str), strict=False))
    kit_map_numero = dict(zip(all_decisions["Proposta Numero"].astype(str), all_decisions["Kit"].astype(str), strict=False))
    changed_cells = 0
    for sheet in ["Abertas mes vigente", "Todas propostas", "Auditoria ERP", "Auditoria outro segmento", "Revisao Cristina", "Todas com decisao"]:
        if sheet in wb.sheetnames:
            changed_cells += update_kit_cells(wb[sheet], kit_map_id, kit_map_numero)

    retirada_cols = [
        "Proposta Numero",
        "Proposta Data",
        "Cliente Nome",
        "CPF/CNPJ",
        "Status consolidado",
        "Valor total formatado",
        "Nome do Vendedor",
        "Setor",
        "Kit",
        "Tipo de agrupamento de produto",
        "Marcadores agrupados",
        "Cliente Fone",
        "Motivo da retirada",
        "Proposta mantida ID",
        "Proposta mantida numero",
        "Proposta mantida status",
        "Proposta mantida data de abertura",
        "Proposta mantida marcador",
    ]
    replace_sheet_from_df(wb, "Retiradas", ret[retirada_cols], "TabelaRetiradasAjuste7")
    wb.save(OUT_XLSX)
    shutil.copy2(OUT_XLSX, ROOT_XLSX)

    summary = {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "kits_corrigidos_total": props_changed + ret_changed + all_changed,
        "multi_kits_removidos_props": props_multi_removed,
        "multi_kits_removidos_ret": ret_multi_removed,
        "multi_kits_removidos_todas": all_multi_removed,
        "celulas_excel_corrigidas": changed_cells,
        "retiradas_enriquecidas": int((ret["Proposta mantida ID"].astype(str) != "").sum()),
        "arquivo_saida": str(OUT_XLSX),
        "arquivo_raiz": str(ROOT_XLSX),
    }
    make_ajustes(summary)
    make_listagem(summary)
    (ROOT / "05 - Apoio" / "resumo_ajuste_7.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
