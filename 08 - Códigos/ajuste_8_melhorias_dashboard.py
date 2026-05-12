from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
DASHBOARD = LOTE / "dashboard.html"
ROOT_DASHBOARD = ROOT / "dashboard.html"
VERSIONED = LOTE / "dashboard_AJUSTE_8_MELHORIAS.html"
ROOT_VERSIONED = ROOT / "dashboard_AJUSTE_8_MELHORIAS.html"
DATA_JSON = LOTE / "apoio" / "dashboard_data.json"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_8.json"


GREEN = "107C41"
WHITE = "FFFFFF"
PALE = "F3FAF2"
BORDER = "D9EAD3"


CSS = """
<style id="ajuste8-dashboard">
.auditIntro{display:grid;grid-template-columns:1.2fr .8fr;gap:14px;margin-bottom:16px}
.auditBox{border:1px solid var(--line);background:#fbfffc;border-radius:12px;padding:14px}
.auditBox h3{margin:0 0 8px;color:var(--green);font-size:17px}
.auditBox p{margin:0;color:var(--muted);line-height:1.45}
.auditKpis{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:14px}
.auditKpi{border:1px solid var(--line);background:#f8fffa;border-radius:12px;padding:12px}
.auditKpi span{display:block;font-size:11px;text-transform:uppercase;color:var(--muted);font-weight:900}
.auditKpi strong{display:block;font-size:22px;color:var(--green);margin-top:5px}
.decisionGrid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:16px}
.decisionCard{border:1px solid var(--line);background:#fff;border-radius:12px;padding:12px}
.decisionCard strong{display:block;color:var(--green);margin-bottom:6px}
.auditTools{display:flex;gap:10px;flex-wrap:wrap;margin:8px 0 14px}
.auditTools button{border:1px solid var(--line);background:#fff;color:var(--green);border-radius:999px;padding:8px 12px;font-weight:900}
.auditTools button.active{background:var(--leaf);border-color:var(--leaf);color:#fff}
.auditTable{overflow:auto;border:1px solid var(--line);border-radius:12px;max-height:420px}
.auditTable table{min-width:860px}
.auditTable th{background:#e2f0d9}
.statusDot{display:inline-flex;align-items:center;gap:6px}
.statusDot:before{content:"";width:8px;height:8px;border-radius:99px;background:var(--lime);display:inline-block}
@media(max-width:1050px){.auditIntro,.decisionGrid{grid-template-columns:1fr}.auditKpis{grid-template-columns:1fr 1fr}}
@media(max-width:680px){.auditKpis{grid-template-columns:1fr}}
</style>
"""


NEW_PANEL = """
<section class="expandPanel open" id="auditoriaPanel">
  <div class="closeLine">
    <h2>Auditoria ERP: decidir o que entra no painel</h2>
    <button data-close="auditoriaPanel">Fechar</button>
  </div>
  <div class="auditIntro">
    <div class="auditBox">
      <h3>O que e esta auditoria?</h3>
      <p>Propostas reais encontradas no ERP, mas separadas porque o vendedor nao estava na lista validada. Elas ficam fora dos indicadores principais ate alguem aprovar a entrada.</p>
    </div>
    <div class="auditBox">
      <h3>Como usar</h3>
      <p>Comece pelos vendedores com mais propostas. Se o vendedor pertence ao time, inclua na regra/base. Se for outro fluxo ou segmento, mantenha separado.</p>
    </div>
  </div>
  <div class="auditKpis">
    <div class="auditKpi"><span>ERP em auditoria</span><strong id="auditTotal">0</strong></div>
    <div class="auditKpi"><span>Vendedores</span><strong id="auditVendors">0</strong></div>
    <div class="auditKpi"><span>Outro segmento</span><strong id="auditOther">0</strong></div>
    <div class="auditKpi"><span>Revisao Cristina</span><strong id="auditCristina">0</strong></div>
    <div class="auditKpi"><span>Duplicidade tecnica</span><strong id="auditTech">0</strong></div>
  </div>
  <div class="decisionGrid">
    <div class="decisionCard"><strong>Prioridade 1</strong><span>Validar vendedores da Auditoria ERP com maior volume.</span></div>
    <div class="decisionCard"><strong>Prioridade 2</strong><span>Confirmar se outro segmento deve ficar fora do comercial.</span></div>
    <div class="decisionCard"><strong>Impacto</strong><span>Nada aqui foi apagado; apenas nao entra no painel principal sem aprovacao.</span></div>
  </div>
  <div class="auditTools">
    <button class="active" data-audit-view="erp">Vendedores ERP</button>
    <button data-audit-view="other">Outro segmento</button>
    <button data-audit-view="rules">Regra de decisao</button>
  </div>
  <div class="dupGrid" id="auditGrid"></div>
  <div class="auditTable" id="auditTable"></div>
</section>
"""


def update_dashboard_html(html: str) -> str:
    html = html.replace("relatorios_e_dashboard_base_AJUSTE_6_REGRAS.xlsx", "relatorios_e_dashboard_base_AJUSTE_7_REGRAS.xlsx")
    if 'id="ajuste8-dashboard"' not in html:
        html = html.replace("</head>", CSS + "</head>", 1)

    start = html.find('<section class="expandPanel open" id="auditoriaPanel">')
    if start == -1:
        raise RuntimeError("Painel auditoriaPanel nao encontrado.")
    end = html.find('<section class="tablePanel">', start)
    if end == -1:
        raise RuntimeError("Fim do painel auditoriaPanel nao encontrado.")
    html = html[:start] + NEW_PANEL + html[end:]

    new = """let auditView='erp';
function renderAudit(){const rows=[...(DATA.auditoria_vendedores||[])].sort((a,b)=>(b.Propostas||0)-(a.Propostas||0)); const vendors=uniq(rows.map(r=>r['Nome do Vendedor']||'(sem vendedor)')); if($('auditTotal'))$('auditTotal').textContent=fmt.format(DATA.auditoria_total_propostas||0); if($('auditVendors'))$('auditVendors').textContent=fmt.format(vendors.length); if($('auditOther'))$('auditOther').textContent=fmt.format(DATA.auditoria_outro_segmento_total||0); if($('auditCristina'))$('auditCristina').textContent=fmt.format(DATA.auditoria_cristina_revisao_total||0); if($('auditTech'))$('auditTech').textContent=fmt.format(DATA.auditoria_duplicidade_tecnica_total||0); const grid=$('auditGrid'); if(!grid)return; grid.innerHTML=''; const table=$('auditTable'); if(auditView==='rules'){grid.innerHTML='<div class="dupCard"><strong>1</strong><span>Vendedor pertence ao time?</span><p class="muted">Incluir na base somente apos validacao.</p></div><div class="dupCard"><strong>2</strong><span>Produto e do foco comercial?</span><p class="muted">Outro segmento continua separado.</p></div><div class="dupCard"><strong>3</strong><span>Impacto nos indicadores</span><p class="muted">Auditoria nao soma no painel principal.</p></div>'; if(table)table.innerHTML=''; return} if(auditView==='other'){grid.innerHTML=`<div class="dupCard"><strong>${fmt.format(DATA.auditoria_outro_segmento_total||0)}</strong><span>Outro segmento</span><p class="muted">Doces, festas, embalagens ou produtos fora do foco.</p></div><div class="dupCard"><strong>${fmt.format(DATA.auditoria_cristina_revisao_total||0)}</strong><span>Revisao Cristina</span><p class="muted">Caso isolado para decisao manual.</p></div><div class="dupCard"><strong>${fmt.format(DATA.auditoria_duplicidade_tecnica_total||0)}</strong><span>Duplicidade tecnica</span><p class="muted">Conferencia de arquivo/pasta, nao venda nova.</p></div>`; if(table)table.innerHTML=''; return} rows.slice(0,10).forEach(x=>{let d=document.createElement('div');d.className='dupCard';d.innerHTML=`<strong>${fmt.format(x.Propostas||0)}</strong><span>${x['Nome do Vendedor']||'(sem vendedor)'}</span><p class="muted">${brl.format(x.Valor||0)} em propostas</p>`;grid.appendChild(d)}); if(table){table.innerHTML=`<table><thead><tr><th>Vendedor</th><th>Propostas</th><th>Valor</th><th>Origem</th><th>Decisao sugerida</th></tr></thead><tbody>${rows.slice(0,80).map(x=>`<tr><td>${x['Nome do Vendedor']||'(sem vendedor)'}</td><td>${fmt.format(x.Propostas||0)}</td><td class="money">${brl.format(x.Valor||0)}</td><td>${x.Origem||''}</td><td><span class="statusDot">Validar se entra na base</span></td></tr>`).join('')}</tbody></table>`}}"""
    html, count = re.subn(r"function renderAudit\(\)\{.*?\nfunction render\(\)", new + "\nfunction render()", html, count=1, flags=re.S)
    if count != 1:
        raise RuntimeError("Funcao renderAudit antiga nao encontrada.")

    old_init = "document.querySelectorAll('button[data-close]').forEach(b=>b.onclick=()=>$(b.dataset.close).classList.remove('open'));"
    new_init = old_init + " document.querySelectorAll('button[data-audit-view]').forEach(b=>b.onclick=()=>{auditView=b.dataset.auditView;document.querySelectorAll('button[data-audit-view]').forEach(x=>x.classList.toggle('active',x===b));renderAudit()});"
    html = html.replace(old_init, new_init, 1)
    return html


def write_listagem(data: dict) -> None:
    wb = Workbook()
    rows_by_sheet = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao", "dashboard_AJUSTE_8_MELHORIAS.html"],
            ["Gerado em", data["gerado_em"]],
            ["Contexto", "Melhorias foram aplicadas no dashboard, nao na planilha de ajustes."],
            ["Auditoria ERP", f"{data['auditoria_erp']} propostas em painel explicativo para decisao."],
            ["Outro segmento", f"{data['outro_segmento']} propostas destacadas no dashboard."],
            ["Kits duplos", "Conferencia preservada: 0 nas abas principais do relatorio."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Auditoria ERP", "Mostrar no dashboard como fila de decisao, fora dos indicadores principais ate aprovacao."],
            ["Outro segmento", "Mostrar contagem separada para nao misturar com o foco comercial."],
            ["Retiradas", "Manter proposta mantida nas colunas do relatorio; dashboard exibe motivos de retirada."],
            ["Kit", "Marcador Kit segue como regra; expositor/mostruario nao vira kit."],
        ],
        "Linha do tempo": [
            ["Ordem", "Data", "Acao", "Resultado"],
            [1, data["gerado_em"], "Desfeito Ajustes 8 em xlsx", "Arquivos de melhoria da planilha foram removidos."],
            [2, data["gerado_em"], "Melhoria movida para dashboard", "Painel Auditoria ERP ficou explicativo e acionavel."],
            [3, data["gerado_em"], "Dashboard versionado", "Criado dashboard_AJUSTE_8_MELHORIAS.html e atualizado dashboard.html."],
        ],
    }
    first = True
    for name, rows in rows_by_sheet.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        for row in rows:
            ws.append(row)
        style_sheet(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def style_sheet(ws, table_name: str) -> None:
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PALE)
    for idx, cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value or "") for cell in cells[:100]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(map(len, sample)) + 2, 70)
    if ws.max_row > 1 and ws.max_column > 1:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws.add_table(table)


def main() -> None:
    data = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    source_dashboard = DASHBOARD if DASHBOARD.exists() else ROOT_DASHBOARD
    html = source_dashboard.read_text(encoding="utf-8")
    new_html = update_dashboard_html(html)
    DASHBOARD.parent.mkdir(parents=True, exist_ok=True)
    DASHBOARD.write_text(new_html, encoding="utf-8")
    VERSIONED.write_text(new_html, encoding="utf-8")
    shutil.copy2(DASHBOARD, ROOT_DASHBOARD)
    shutil.copy2(VERSIONED, ROOT_VERSIONED)

    summary = {
        "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "dashboard": str(DASHBOARD.resolve()),
        "dashboard_raiz": str(ROOT_DASHBOARD.resolve()),
        "dashboard_versionado": str(VERSIONED.resolve()),
        "dashboard_versionado_raiz": str(ROOT_VERSIONED.resolve()),
        "auditoria_erp": int(data.get("auditoria_total_propostas", 0) or 0),
        "vendedores_auditoria": len({x.get("Nome do Vendedor", "") for x in data.get("auditoria_vendedores", []) if x.get("Nome do Vendedor", "")}),
        "outro_segmento": int(data.get("auditoria_outro_segmento_total", 0) or 0),
        "cristina_revisao": int(data.get("auditoria_cristina_revisao_total", 0) or 0),
        "duplicidade_tecnica": int(data.get("auditoria_duplicidade_tecnica_total", 0) or 0),
    }
    write_listagem(summary)
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
