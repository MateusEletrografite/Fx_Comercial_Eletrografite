from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
LOTE = ROOT / "06 - Lote final organizado"
DASHBOARD = ROOT / "dashboard.html"
LOTE_DASHBOARD = LOTE / "dashboard.html"
VERSIONED = ROOT / "dashboard_AJUSTE_9_ORDENACAO_CABECALHO.html"
LOTE_VERSIONED = LOTE / "dashboard_AJUSTE_9_ORDENACAO_CABECALHO.html"
LISTAGEM = ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"
SUMMARY = ROOT / "05 - Apoio" / "resumo_dashboard_ajuste_9.json"


SORT_CSS = """
<style id="ajuste9-sort">
th.sortable{cursor:pointer;user-select:none;position:sticky}
th.sortable:after{content:" ⇅";font-size:11px;color:#607267}
th.sortable.activeSort{background:#c6e0b4;color:#0e5a2a}
th.sortable.activeSort.asc:after{content:" ▲";color:#0e5a2a}
th.sortable.activeSort.desc:after{content:" ▼";color:#0e5a2a}
</style>
"""


NEW_RENDER_TABLE = """function renderTable(rows){const head=$('theadRow'),tb=$('tbody'); head.innerHTML=''; tb.innerHTML=''; selectedColumns.forEach(c=>{let th=document.createElement('th');th.dataset.sort=c;th.textContent=c;th.title='Clique para ordenar por '+c;th.className='sortable'+((sortKey===c||(sortKey==='Data ISO'&&c==='Proposta Data')||(sortKey==='Valor total da proposta'&&c==='Valor total formatado'))?' activeSort '+sortDir:'');th.onclick=()=>{if(sortKey===c||(sortKey==='Data ISO'&&c==='Proposta Data')||(sortKey==='Valor total da proposta'&&c==='Valor total formatado')){sortDir=sortDir==='asc'?'desc':'asc'}else{sortKey=c;sortDir='asc'}render()};head.appendChild(th)}); rows.slice(0,1200).forEach(r=>{let tr=document.createElement('tr'); selectedColumns.forEach(c=>{let td=document.createElement('td'); const v=valueForCell(r,c); if(c==='Proposta Numero'||c==='Proposta ID'){const b=document.createElement('button');b.className='linkBtn prop';b.textContent=v||r['Proposta Numero']||r['Proposta ID'];b.onclick=()=>openProposal(r);td.appendChild(b)}else if(c==='Cliente Nome'){const b=document.createElement('button');b.className='linkBtn client';b.textContent=v;b.onclick=()=>openClient(r);td.appendChild(b)}else if(c==='Status painel'||c==='Status consolidado'){td.innerHTML=`<span class="pill">${v}</span>`}else if(c==='Valor total da proposta'||c==='Valor total formatado'){td.className='money';td.textContent=c==='Valor total formatado'?(v||brl.format(+r['Valor total da proposta']||0)):v}else if(c==='Marcadores agrupados'||c==='Produtos agrupados'||c==='Propostas relacionadas do cliente'){td.className='muted';td.textContent=v}else{td.textContent=v} tr.appendChild(td)}); tb.appendChild(tr)})}"""


def update_html(path: Path) -> str:
    html = path.read_text(encoding="utf-8")
    if 'id="ajuste9-sort"' not in html:
        html = html.replace("</head>", SORT_CSS + "</head>", 1)
    html, count = re.subn(r"function renderTable\(rows\)\{.*?\nfunction openSeller", NEW_RENDER_TABLE + "\nfunction openSeller", html, count=1, flags=re.S)
    if count != 1:
        raise RuntimeError(f"renderTable nao encontrado em {path}")
    path.write_text(html, encoding="utf-8")
    return html


def write_listagem(generated: str) -> None:
    wb = Workbook()
    sheets = {
        "Resumo": [
            ["Item", "Detalhe"],
            ["Versao", "dashboard_AJUSTE_9_ORDENACAO_CABECALHO.html"],
            ["Gerado em", generated],
            ["Correcao do entendimento", "O pedido era ordenar ao clicar no cabecalho, nao fixar a ordem das colunas."],
            ["Aplicacao", "Todos os cabecalhos da tabela ganharam indicacao visual de ordenacao."],
        ],
        "Regras": [
            ["Regra", "Aplicacao"],
            ["Ordenacao por cabecalho", "Clique no titulo da coluna para ordenar; novo clique inverte crescente/decrescente."],
            ["Data", "Proposta Data ordena usando Data ISO para nao ordenar como texto."],
            ["Valor", "Valor total formatado ordena usando Valor total da proposta numerico."],
            ["Texto/numero", "Demais colunas usam comparacao pt-BR com numeric=true."],
        ],
        "Linha do tempo": [
            ["Ordem", "Data", "Acao", "Resultado"],
            [1, generated, "Usuario corrigiu interpretacao", "Item passou de ordem fixa de colunas para ordenacao clicavel por cabecalho."],
            [2, generated, "Dashboard atualizado", "Cabecalhos com cursor, dica e seta de ordenacao."],
            [3, generated, "Versao criada", "dashboard_AJUSTE_9_ORDENACAO_CABECALHO.html"],
        ],
    }
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name
        for row in rows:
            ws.append(row)
        style(ws, f"Tabela{name.replace(' ', '')}")
    wb.save(LISTAGEM)


def style(ws, table_name: str) -> None:
    thin = Side(style="thin", color="D9EAD3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="107C41")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3FAF2")
    for idx, cells in enumerate(ws.columns, start=1):
        sample = [str(cell.value or "") for cell in cells[:100]]
        ws.column_dimensions[get_column_letter(idx)].width = min(max(map(len, sample)) + 2, 72)
    table = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)


def main() -> None:
    generated = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    update_html(DASHBOARD)
    shutil.copy2(DASHBOARD, LOTE_DASHBOARD)
    shutil.copy2(DASHBOARD, VERSIONED)
    shutil.copy2(DASHBOARD, LOTE_VERSIONED)
    write_listagem(generated)
    summary = {
        "gerado_em": generated,
        "ajuste": "Ordenacao por clique no cabecalho da tabela",
        "dashboard": str(DASHBOARD.resolve()),
        "dashboard_lote": str(LOTE_DASHBOARD.resolve()),
        "dashboard_versionado": str(VERSIONED.resolve()),
        "observacao": "Nao foi fixada ordem de colunas; foi melhorada a ordenacao clicavel por cabecalho.",
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
