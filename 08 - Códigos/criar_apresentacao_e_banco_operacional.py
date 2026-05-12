from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = ROOT.parent
ATEND = DOWNLOADS / "Atendimentos"
APRESENTACAO = ROOT / "Apresentação"
BANCO = DOWNLOADS / "Banco de Dados Operacional"
APOIO = ROOT / "05 - Apoio"
REGRAS = ROOT / "09 - Regras"
SUMMARY = APOIO / "resumo_apresentacao_banco_operacional.json"
LISTAGEM = REGRAS / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


PRESENTATION_FILES = [
    (ROOT / "dashboard.html", "01_DASHBOARD_ATUAL.html"),
    (ROOT / "dashboard_AJUSTE_16_CLIQUES_RX_CLIENTE.html", "02_DASHBOARD_AJUSTE_16_CLIQUES_RX_CLIENTE.html"),
    (ROOT / "04 - Fonte" / "MATRIZ_FONTES_ATUALIZACAO_DIARIA.xlsx", "03_MATRIZ_FONTES_ATUALIZACAO_DIARIA.xlsx"),
    (ROOT / "04 - Fonte" / "INVENTARIO_FONTES_ATUALIZACAO_DIARIA.xlsx", "04_INVENTARIO_FONTES_ATUALIZACAO_DIARIA.xlsx"),
    (ROOT / "relatorios_e_dashboard_base_AJUSTE_7_REGRAS_COM_ATENDIMENTOS_20260510-215511.xlsx", "05_BASE_PROPOSTAS_COM_ATENDIMENTOS_ATUAL.xlsx"),
    (ROOT / "relatorio_plano_ativacao_atendimentos_AJUSTE_13.xlsx", "06_PLANO_ATIVACAO_ATENDIMENTOS.xlsx"),
    (ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx", "07_REGRAS_LINHA_DO_TEMPO.xlsx"),
]


SOURCE_GROUPS = {
    "01 - ERP Olist/Propostas": [
        ROOT / "01 - Em aberto - 12-2024 a 05-2026",
        ROOT / "02 - Concluídas - 12-2024 a 05-2026",
        ROOT / "03 - Não aprovadas - 12-2024 a 05-2026",
        ROOT / "relatorios_e_dashboard_base_AJUSTE_7_REGRAS.xlsx",
        ROOT / "relatorios_e_dashboard_base_AJUSTE_7_REGRAS_COM_ATENDIMENTOS_20260510-215511.xlsx",
    ],
    "01 - ERP Olist/Marcadores": [
        ROOT / "Propostas por mes com marcadores",
    ],
    "01 - ERP Olist/Clientes e auxiliares": [
        ROOT / "Cnpjs Cenprot.xlsx",
        ROOT / "Compras clientes até 20-04.xlsx",
    ],
    "02 - Atendimentos/Consolidados": [
        ATEND / "relatorio_atendimentos_consolidado_classificado_empresa_fornecedor_20260511-085458.xlsx",
        ATEND / "relatorio_atendimentos_com_propostas_abertas_20260510-212628.xlsx",
        ATEND / "Analise de quantidade de atendimentos.xlsx",
        ATEND / "Atendimentos Sfx após 01-01-2026.xlsx",
        ATEND / "Classificação empresa - fornecedor.xlsx",
    ],
    "02 - Atendimentos/Compras": [
        ATEND / "Compras clientes até 20-04.xlsx",
    ],
    "03 - Regras internas/Propostas": [
        ROOT / "Regras.xlsx",
        ROOT / "Ajustes 7.xlsx",
        ROOT / "Layout e alterações.xlsx",
    ],
    "04 - Saídas consolidadas/Propostas do sistema": [
        ROOT / "dashboard.html",
        ROOT / "dashboard_AJUSTE_16_CLIQUES_RX_CLIENTE.html",
        ROOT / "relatorio_plano_ativacao_atendimentos_AJUSTE_13.xlsx",
        ROOT / "09 - Regras" / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx",
    ],
}


def copy_file(src: Path, dest: Path) -> dict[str, object]:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if not src.exists() or src.name.startswith("~$"):
        return {"origem": str(src), "destino": str(dest), "status": "não encontrado" if not src.exists() else "temporário ignorado", "tamanho": ""}
    if src.resolve() != dest.resolve():
        shutil.copy2(src, dest)
    return {"origem": str(src), "destino": str(dest), "status": "copiado", "tamanho": src.stat().st_size}


def copy_any(src: Path, dest_dir: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not src.exists():
        rows.append({"origem": str(src), "destino": str(dest_dir), "status": "não encontrado", "tamanho": ""})
        return rows
    if src.is_file():
        rows.append(copy_file(src, dest_dir / src.name))
        return rows
    for file in src.rglob("*"):
        if not file.is_file() or file.name.startswith("~$"):
            continue
        if file.suffix.lower() not in {".xlsx", ".xls", ".csv", ".txt", ".json", ".html"}:
            continue
        rows.append(copy_file(file, dest_dir / src.name / file.relative_to(src)))
    return rows


def style_ws(ws, table_name: str, header_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.value = "" if cell.value is None else str(cell.value)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, width in enumerate([32, 70, 70, 18, 16], start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    ws.freeze_panes = "A2"
    tab = Table(displayName=table_name, ref=f"A1:{ws.cell(ws.max_row, header_count).coordinate}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)


def write_manifest(path: Path, title: str, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    headers = ["Grupo", "Origem", "Destino", "Status", "Tamanho bytes"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get("grupo", ""), r["origem"], r["destino"], r["status"], r["tamanho"]])
    style_ws(ws, "Tabela" + "".join(ch for ch in title if ch.isalnum()), len(headers))
    wb.save(path)


def update_listagem(presentation_count: int, bank_count: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.append([now, "APRESENTACAO_E_BANCO_OPERACIONAL", "Apresentação", "Separar versões finais atuais em pasta Apresentação.", "Usuário pediu para visualizar sem se perder.", f"{presentation_count} arquivos copiados com nomes numerados."])
    ws.append([now, "APRESENTACAO_E_BANCO_OPERACIONAL", "Banco central", "Criar Banco de Dados Operacional para fontes originais de todos os workspaces.", "Usuário quer uma espécie de banco de dados central.", f"{bank_count} arquivos copiados/preservados; originais não foram removidos."])
    style_ws(ws, "TabelaLinhaTempoApresentacaoBanco", len(headers))
    wb.save(LISTAGEM)


def main() -> None:
    APRESENTACAO.mkdir(exist_ok=True)
    APOIO.mkdir(exist_ok=True)
    REGRAS.mkdir(exist_ok=True)
    BANCO.mkdir(exist_ok=True)

    presentation_rows = []
    for src, name in PRESENTATION_FILES:
        r = copy_file(src, APRESENTACAO / name)
        r["grupo"] = "Apresentação"
        presentation_rows.append(r)

    bank_rows = []
    for group, entries in SOURCE_GROUPS.items():
        for src in entries:
            for row in copy_any(src, BANCO / group):
                row["grupo"] = group
                bank_rows.append(row)

    write_manifest(APRESENTACAO / "00_MANIFESTO_APRESENTACAO.xlsx", "Apresentacao", presentation_rows)
    write_manifest(BANCO / "00_INVENTARIO_BANCO_OPERACIONAL.xlsx", "Banco", bank_rows)

    update_listagem(
        sum(1 for r in presentation_rows if r["status"] == "copiado"),
        sum(1 for r in bank_rows if r["status"] == "copiado"),
    )

    SUMMARY.write_text(json.dumps({
        "apresentacao": str(APRESENTACAO),
        "banco_operacional": str(BANCO),
        "arquivos_apresentacao": sum(1 for r in presentation_rows if r["status"] == "copiado"),
        "arquivos_banco": sum(1 for r in bank_rows if r["status"] == "copiado"),
        "manifesto_apresentacao": str(APRESENTACAO / "00_MANIFESTO_APRESENTACAO.xlsx"),
        "inventario_banco": str(BANCO / "00_INVENTARIO_BANCO_OPERACIONAL.xlsx"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({
        "ok": True,
        "apresentacao": str(APRESENTACAO),
        "banco_operacional": str(BANCO),
        "arquivos_apresentacao": sum(1 for r in presentation_rows if r["status"] == "copiado"),
        "arquivos_banco": sum(1 for r in bank_rows if r["status"] == "copiado"),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
