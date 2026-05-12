from __future__ import annotations

import json
import shutil
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = ROOT.parent
BANCO = DOWNLOADS / "Banco de Dados Operacional"
ENTRADA = BANCO / "00 - Entrada diaria"
APRESENTACAO = ROOT / "Apresentação"
FONTE = ROOT / "04 - Fonte"
REGRAS = ROOT / "09 - Regras"
APOIO = ROOT / "05 - Apoio"
CODIGOS = ROOT / "08 - Códigos"

OUT_XLSX = APRESENTACAO / "09_DOCUMENTO_UNICO_PROPOSTAS_COMPRAS_E_ATENDIMENTOS.xlsx"
OUT_PDF = APRESENTACAO / "09_DOCUMENTO_UNICO_PROPOSTAS_COMPRAS_E_ATENDIMENTOS.pdf"
SUMMARY = APOIO / "resumo_documento_unico_compras_atendimentos.json"
LISTAGEM = REGRAS / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


COMPRAS_DIR = ENTRADA / "01 - ERP Olist" / "Compras ERP"
PROPOSTAS_ABERTAS_DIR = ENTRADA / "01 - ERP Propostas" / "Em aberto"
PROPOSTAS_CONCLUIDAS_DIR = ENTRADA / "01 - ERP Propostas" / "Concluidas"
PROPOSTAS_NAO_APROVADAS_DIR = ENTRADA / "01 - ERP Propostas" / "Nao aprovadas"
PROPOSTAS_MARCADORES_DIR = ENTRADA / "01 - ERP Propostas" / "Marcadores"
SELLFLUX_DIR = ENTRADA / "02 - SellFlux Atendimentos"
LEGADO_COMPRAS_DIR = ENTRADA / "03 - Compras"
LEGADO_ATEND_DIR = ENTRADA / "02 - Atendimentos"


COMPRAS_COLS = [
    ("Data de inclusão", "Opcional", "Data em que o pedido entrou no ERP, se disponível.", "Data"),
    ("Pedido", "Obrigatória", "Número/ID do pedido no ERP.", "Texto ou número"),
    ("Duplicidade (Número de pedido)", "Opcional", "Campo auxiliar usado em auditoria de duplicidade.", "Texto"),
    ("CNPJ", "Obrigatória", "CPF/CNPJ do cliente conforme ERP.", "Texto"),
    ("Cnpj sem pontos para Cruzamento", "Obrigatória", "Mesmo CPF/CNPJ apenas com números, usado como chave.", "Texto"),
    ("Razão Social", "Obrigatória", "Nome/razão social do cliente.", "Texto"),
    ("Valor do pedido", "Obrigatória", "Valor do pedido/compra.", "Número"),
    ("Data da compra", "Obrigatória", "Data do pedido/compra.", "Data"),
    ("Vendedor da Compra", "Obrigatória", "Vendedor responsável no ERP.", "Texto"),
    ("Telefone Tiny", "Recomendada", "Telefone vindo do ERP/Tiny.", "Texto"),
    ("Telefone Biz", "Recomendada", "Telefone auxiliar, se existir.", "Texto"),
    ("Whatsapp Mande um Zap", "Recomendada", "Whatsapp auxiliar, se existir.", "Texto"),
    ("Whatsapp Biz", "Recomendada", "Whatsapp auxiliar, se existir.", "Texto"),
    ("Motivo da duplicidade", "Opcional", "Auditoria interna.", "Texto"),
    ("Indicadores da duplicidade", "Opcional", "Auditoria interna.", "Texto"),
    ("Tipo de duplicidade", "Opcional", "Classificação de duplicidade.", "Texto"),
    ("Ultimo Pedido", "Obrigatória", "Último pedido identificado para o cliente.", "Texto ou número"),
    ("Valor do último pedido", "Obrigatória", "Valor da última compra do cliente.", "Número"),
    ("Data última compra", "Obrigatória", "Última data de compra do cliente.", "Data"),
    ("Vendedor do ultimo pedido", "Obrigatória", "Vendedor do último pedido.", "Texto"),
]


PROPOSTAS_COLS = [
    ("ID", "Obrigatória", "ID interno da proposta no ERP.", "Texto ou número", "Chave principal da proposta."),
    ("Número da proposta", "Obrigatória", "Número comercial da proposta.", "Texto ou número", ""),
    ("Data", "Obrigatória", "Data de abertura/emissão da proposta.", "Data", ""),
    ("Data proximo contato", "Recomendada", "Data prevista de próximo contato.", "Data", ""),
    ("ID contato", "Obrigatória", "ID do contato/cliente no ERP.", "Texto ou número", "Ajuda a casar com clientes."),
    ("Nome do contato", "Obrigatória", "Nome/razão social do cliente.", "Texto", ""),
    ("Aos cuidados de", "Recomendada", "Pessoa de contato.", "Texto", ""),
    ("Lista de Preço", "Opcional", "Lista usada no ERP.", "Texto", ""),
    ("Tipo de Pessoa", "Recomendada", "F/J ou equivalente.", "Texto", ""),
    ("CPF/CNPJ", "Obrigatória", "Documento do cliente.", "Texto", "Chave principal de cliente."),
    ("RG/IE", "Opcional", "Inscrição estadual/RG.", "Texto", ""),
    ("CEP", "Recomendada", "CEP do cliente.", "Texto", ""),
    ("Município", "Recomendada", "Cidade do cliente.", "Texto", ""),
    ("UF", "Recomendada", "Estado do cliente.", "Texto", ""),
    ("Endereço", "Opcional", "Endereço.", "Texto", ""),
    ("Endereço Nro", "Opcional", "Número do endereço.", "Texto", ""),
    ("Complemento", "Opcional", "Complemento.", "Texto", ""),
    ("Bairro", "Opcional", "Bairro.", "Texto", ""),
    ("Fone", "Recomendada", "Telefone fixo/principal.", "Texto", "Chave alternativa com atendimentos."),
    ("Celular", "Recomendada", "Celular/WhatsApp.", "Texto", "Chave alternativa com atendimentos."),
    ("E-mail", "Recomendada", "Email do cliente.", "Texto", ""),
    ("Desconto", "Recomendada", "Desconto aplicado.", "Número/texto", ""),
    ("Frete", "Recomendada", "Frete aplicado.", "Número/texto", ""),
    ("Observações", "Opcional", "Observações comerciais.", "Texto", ""),
    ("Validade", "Opcional", "Validade da proposta.", "Texto ou número", ""),
    ("Prazo de Entrega", "Opcional", "Prazo informado.", "Texto", ""),
    ("Situação", "Obrigatória", "Situação/status vindo do ERP.", "Texto", "Base para Em aberto/Concluídas/Não aprovadas."),
    ("Introdução", "Opcional", "Texto de introdução.", "Texto", ""),
    ("ID produto", "Obrigatória", "ID do produto/item.", "Texto ou número", "Uma proposta pode ter várias linhas."),
    ("Descrição", "Obrigatória", "Descrição do produto.", "Texto", "Usado para categoria/kit quando necessário."),
    ("Quantidade", "Obrigatória", "Quantidade do item.", "Número", ""),
    ("Valor unitário", "Obrigatória", "Valor unitário do item.", "Número", ""),
    ("Descrição complementar", "Opcional", "Complemento do item.", "Texto", ""),
    ("Vendedor", "Obrigatória", "Vendedor responsável.", "Texto", "Base do RX da equipe."),
    ("Destinatário", "Opcional", "Destinatário de entrega.", "Texto", ""),
    ("CPF/CNPJ entrega", "Opcional", "Documento entrega.", "Texto", ""),
    ("CEP entrega", "Opcional", "CEP entrega.", "Texto", ""),
    ("Município entrega", "Opcional", "Cidade entrega.", "Texto", ""),
    ("UF entrega", "Opcional", "UF entrega.", "Texto", ""),
    ("Endereço entrega", "Opcional", "Endereço entrega.", "Texto", ""),
    ("Endereço Nro entrega", "Opcional", "Número entrega.", "Texto", ""),
    ("Complemento entrega", "Opcional", "Complemento entrega.", "Texto", ""),
    ("Bairro entrega", "Opcional", "Bairro entrega.", "Texto", ""),
    ("Fone entrega", "Opcional", "Telefone entrega.", "Texto", ""),
    ("Inscrição Estadual entrega", "Opcional", "IE entrega.", "Texto", ""),
]

PROPOSTAS_MARCADORES_COLS = [
    ("Marcadores", "Obrigatória no relatório com marcadores", "Marcadores/tags ligados à proposta/cliente.", "Texto", "Usado para setor, kit, carteira/tráfego e auditorias."),
    ("Razao social marcador", "Recomendada", "Razão social associada ao marcador, se houver.", "Texto", ""),
    ("CNPJ marcador", "Recomendada", "CNPJ associado ao marcador, se houver.", "Texto", ""),
    ("Arquivo marcador", "Recomendada", "Arquivo/origem do marcador.", "Texto", ""),
]


SELLFLUX_COLS = [
    ("Início do atendimento", "Obrigatória", "Data/hora em que o atendimento iniciou no SellFlux.", "Data/hora", "No arquivo atual apareceu com codificação quebrada como 'InÃ\xadcio do atendimento'."),
    ("Conclusão do atendimento", "Obrigatória", "Data/hora de conclusão do atendimento.", "Data/hora", "No arquivo atual apareceu como 'ConclusÃ£o do atendimento'."),
    ("Atendente que iniciou o atendimento", "Obrigatória", "Quem iniciou o atendimento.", "Texto", ""),
    ("Nome do cliente", "Obrigatória", "Nome do cliente no SellFlux.", "Texto", ""),
    ("Telefone", "Obrigatória", "Telefone do cliente, de preferência com DDI/DDDs ou só números.", "Texto", "É a principal chave de casamento com propostas/compras."),
    ("Email", "Recomendada", "Email do cliente.", "Texto", ""),
    ("Nota", "Opcional", "Nota/avaliação, se houver.", "Texto", ""),
    ("Comentário", "Opcional", "Comentário do atendimento.", "Texto", "No arquivo atual apareceu como 'ComentÃ¡rio'."),
    ("Assunto", "Recomendada", "Assunto/motivo do contato.", "Texto", ""),
    ("Resumo", "Recomendada", "Resumo do atendimento.", "Texto", ""),
    ("Atendente que finalizou o atendimento", "Recomendada", "Quem encerrou/finalizou.", "Texto", ""),
    ("Tempo total do atendimento", "Opcional", "Duração total.", "Duração/texto", ""),
    ("Tempo total do atendimento (após o atendente iniciar o atendimento)", "Opcional", "Duração após início pelo atendente.", "Duração/texto", "No arquivo atual apareceu com 'apÃ³s'."),
    ("Tempo total na fila (minutos)", "Opcional", "Tempo em fila.", "Duração/texto", ""),
    ("Tempo total do departamento (minutos)", "Opcional", "Tempo por departamento.", "Duração/texto", ""),
    ("Tempo médio entre as respostas", "Opcional", "Tempo médio entre respostas.", "Duração/texto", "No arquivo atual apareceu como 'Tempo mÃ©dio...'."),
    ("Dentro da jornada de trabalho?", "Opcional", "Sim/Não para jornada.", "Texto", "No arquivo atual apareceu como 'NÃ£o' nos valores."),
]


def ensure_dirs() -> None:
    for folder in [
        APRESENTACAO, FONTE, REGRAS, APOIO,
        PROPOSTAS_ABERTAS_DIR, PROPOSTAS_CONCLUIDAS_DIR, PROPOSTAS_NAO_APROVADAS_DIR, PROPOSTAS_MARCADORES_DIR,
        COMPRAS_DIR, SELLFLUX_DIR, LEGADO_COMPRAS_DIR, LEGADO_ATEND_DIR,
        BANCO / "01 - ERP Olist" / "Propostas", BANCO / "01 - ERP Olist" / "Compras", BANCO / "02 - SellFlux" / "Atendimentos",
    ]:
        folder.mkdir(parents=True, exist_ok=True)


def fix_handler() -> None:
    handler = CODIGOS / "atualizar_tudo_ao_salvar.py"
    if not handler.exists():
        return
    text = handler.read_text(encoding="utf-8")
    text = text.replace(
        'if "02 - Atendimentos" in parts:\n        return BANCO / "02 - Atendimentos" / "Consolidados" / path.name, FONTE / "02 - Atendimentos" / path.name\n    if "03 - Compras" in parts:\n        return BANCO / "02 - Atendimentos" / "Compras" / path.name, FONTE / "03 - Compras" / path.name',
        'if "02 - SellFlux Atendimentos" in parts or "02 - Atendimentos" in parts:\n        return BANCO / "02 - SellFlux" / "Atendimentos" / path.name, FONTE / "02 - Atendimentos" / path.name\n    if "Compras ERP" in parts or "03 - Compras" in parts:\n        return BANCO / "01 - ERP Olist" / "Compras" / path.name, FONTE / "03 - Compras" / path.name',
    )
    handler.write_text(text, encoding="utf-8")


def style_ws(ws, table_name: str, header_count: int, widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor="1F7A4D")
    thin = Side(style="thin", color="D5E7DC")
    for cell in ws[1]:
        cell.value = "" if cell.value is None else str(cell.value)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        tab = Table(displayName=table_name, ref=f"A1:{ws.cell(ws.max_row, header_count).coordinate}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws.add_table(tab)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]], table_name: str, widths: list[int]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_ws(ws, table_name, len(headers), widths)
    ws.sheet_view.showGridLines = False


def build_workbook() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    resumo_rows = [
        ["Propostas ERP", "Fonte: ERP Olist/Tiny. Usar os CSVs padrão de propostas.", f"Em aberto: {PROPOSTAS_ABERTAS_DIR}\nConcluídas: {PROPOSTAS_CONCLUIDAS_DIR}\nNão aprovadas: {PROPOSTAS_NAO_APROVADAS_DIR}\nMarcadores: {PROPOSTAS_MARCADORES_DIR}", "Atualizar diariamente por status e por marcadores.", "É a base central do dashboard comercial."],
        ["Compras ERP", "Fonte: ERP Olist/Tiny. Referência atual: Compras clientes até 20-04.xlsx.", str(COMPRAS_DIR), "Atualizar quando exportar novo relatório de compras/pedidos do ERP.", "Não é fonte de atendimentos."],
        ["Atendimentos SellFlux", "Fonte: relatório de atendimentos SellFlux.", str(SELLFLUX_DIR), "Atualizar com o relatório bruto do SellFlux.", "Não usar mais como regra principal o consolidado antigo de atendimentos, salvo como saída tratada."],
        ["Automação", "Salvar arquivo novo nas pastas acima.", str(ENTRADA), "O monitor copia para Banco/Fonte e atualiza inventários.", "Compras agora vai para Banco/01 - ERP Olist/Compras."],
    ]
    add_sheet(wb, "Resumo", ["Base", "Origem correta", "Onde salvar", "Quando atualizar", "Observação"], resumo_rows, "TabelaResumoComprasAtend", [24, 54, 72, 44, 48])

    add_sheet(
        wb,
        "Propostas ERP",
        ["Coluna", "Obrigatoriedade", "Descrição", "Formato", "Observação"],
        [[*row] for row in PROPOSTAS_COLS],
        "TabelaPropostasERP",
        [34, 22, 66, 22, 52],
    )

    add_sheet(
        wb,
        "Propostas marcadores",
        ["Coluna", "Obrigatoriedade", "Descrição", "Formato", "Observação"],
        [[*row] for row in PROPOSTAS_MARCADORES_COLS],
        "TabelaPropostasMarcadores",
        [34, 32, 66, 22, 52],
    )

    add_sheet(
        wb,
        "Compras ERP",
        ["Coluna", "Obrigatoriedade", "Descrição", "Formato"],
        [[*row] for row in COMPRAS_COLS],
        "TabelaComprasERP",
        [36, 18, 70, 22],
    )

    add_sheet(
        wb,
        "Atendimentos SellFlux",
        ["Coluna esperada", "Obrigatoriedade", "Descrição", "Formato", "Observação"],
        [[*row] for row in SELLFLUX_COLS],
        "TabelaAtendSellFlux",
        [48, 18, 66, 22, 56],
    )

    regras_rows = [
        [1, "Propostas ERP", "ID + Número da proposta", "Identifica a proposta e evita duplicidade.", "Não alterar cabeçalho dos CSVs padrão."],
        [2, "Propostas ERP", "Situação", "Define status painel: Em aberto, Concluídas, Não aprovadas.", "Separar os arquivos por pasta de status ajuda a auditar."],
        [3, "Propostas ERP", "Descrição + ID produto", "Base para produtos, categorias e kits.", "Não inferir kit por Expositor/Mostruário isolado."],
        [4, "Propostas marcadores", "Marcadores", "Base para setor, carteira/tráfego, kits e auditorias.", "Relatório de marcadores deve vir junto dos CSVs de proposta."],
        [5, "Compras ERP", "CNPJ/Cnpj sem pontos para Cruzamento", "Chave principal de casamento com propostas e clientes.", "Se CNPJ vier vazio, telefone ajuda, mas não substitui totalmente."],
        [6, "Compras ERP", "Data última compra", "Base para clientes sem compra há 180/365 dias.", "Deve estar em formato de data."],
        [7, "Compras ERP", "Valor do último pedido", "Base para potencial e priorização comercial.", "Número; evitar R$ como texto quando possível."],
        [8, "SellFlux", "Telefone", "Chave principal para cruzar atendimento com cliente/proposta/compra.", "Preferir só números; DDI 55 aceito."],
        [9, "SellFlux", "Início/Conclusão do atendimento", "Base para último contato e tempo desde contato.", "Se vier com acento quebrado, o script deve aceitar alias."],
        [10, "SellFlux", "Atendente", "Base para distribuição e auditoria de atendimento.", "Usar atendente que iniciou/finalizou conforme objetivo."],
    ]
    add_sheet(wb, "Regras de uso", ["Ordem", "Base", "Campo", "Uso no dashboard", "Observação"], regras_rows, "TabelaRegrasUso", [10, 22, 34, 60, 56])

    checklist_rows = [
        ["Propostas em aberto foram salvas?", str(PROPOSTAS_ABERTAS_DIR), ""],
        ["Propostas concluídas foram salvas?", str(PROPOSTAS_CONCLUIDAS_DIR), ""],
        ["Propostas não aprovadas foram salvas?", str(PROPOSTAS_NAO_APROVADAS_DIR), ""],
        ["Propostas com marcadores foram salvas?", str(PROPOSTAS_MARCADORES_DIR), ""],
        ["Compras ERP foi salvo na pasta correta?", str(COMPRAS_DIR), ""],
        ["Atendimentos SellFlux foi salvo na pasta correta?", str(SELLFLUX_DIR), ""],
        ["CNPJ ou telefone está preenchido?", "Compras: CNPJ. SellFlux: Telefone.", ""],
        ["Datas estão atualizadas?", "Data última compra e início/conclusão do atendimento.", ""],
        ["Arquivo não está aberto como temporário ~$?", "Fechar Excel antes de processar, se necessário.", ""],
    ]
    add_sheet(wb, "Checklist", ["Check", "Como conferir", "OK?"], checklist_rows, "TabelaChecklistComprasAtend", [42, 72, 12])

    wb.save(OUT_XLSX)


def export_pdf() -> tuple[bool, str]:
    ps = f"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$wb = $excel.Workbooks.Open('{str(OUT_XLSX).replace("'", "''")}')
foreach ($ws in $wb.Worksheets) {{
  $ws.PageSetup.Orientation = 2
  $ws.PageSetup.Zoom = $false
  $ws.PageSetup.FitToPagesWide = 1
  $ws.PageSetup.FitToPagesTall = $false
}}
$wb.ExportAsFixedFormat(0, '{str(OUT_PDF).replace("'", "''")}')
$wb.Close($false)
$excel.Quit()
"""
    try:
        proc = subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps], capture_output=True, text=True, timeout=120)
        return proc.returncode == 0 and OUT_PDF.exists(), (proc.stdout + proc.stderr)[-2000:]
    except Exception as exc:
        return False, str(exc)


def update_listagem(pdf_ok: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.append([now, "DOCUMENTO_UNICO_PROPOSTAS_COMPRAS_ATENDIMENTOS", "Propostas", "Incluir propostas no documento único de alimentação diária.", "Usuário pediu tudo junto.", f"Pastas de entrada: {PROPOSTAS_ABERTAS_DIR}; {PROPOSTAS_CONCLUIDAS_DIR}; {PROPOSTAS_NAO_APROVADAS_DIR}; {PROPOSTAS_MARCADORES_DIR}"])
    ws.append([now, "DOCUMENTO_UNICO_PROPOSTAS_COMPRAS_ATENDIMENTOS", "Compras", "Corrigir regra: compras são alimentadas pelo relatório ERP de compras/pedidos.", "Usuário confirmou que compras vem do ERP.", f"Pasta de entrada: {COMPRAS_DIR}"])
    ws.append([now, "DOCUMENTO_UNICO_PROPOSTAS_COMPRAS_ATENDIMENTOS", "Atendimentos", "Corrigir regra: atendimentos agora vêm do relatório SellFlux.", "Usuário confirmou que atendimentos vêm como relatórios de atendimentos SellFlux.", f"Pasta de entrada: {SELLFLUX_DIR}"])
    ws.append([now, "DOCUMENTO_UNICO_PROPOSTAS_COMPRAS_ATENDIMENTOS", "Entregável", "Criar documento único em XLSX e PDF com colunas e locais de salvamento.", "Pedido do usuário.", "PDF gerado com sucesso." if pdf_ok else "PDF não gerado automaticamente; XLSX criado."])
    style_ws(ws, "TabelaLinhaTempoComprasAtend", len(headers), [24, 32, 22, 54, 54, 54])
    wb.save(LISTAGEM)


def main() -> None:
    ensure_dirs()
    fix_handler()
    build_workbook()
    pdf_ok, pdf_msg = export_pdf()
    update_listagem(pdf_ok)
    SUMMARY.write_text(json.dumps({
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "xlsx": str(OUT_XLSX),
        "pdf": str(OUT_PDF) if OUT_PDF.exists() else "",
        "pdf_ok": pdf_ok,
        "pdf_msg": pdf_msg,
        "compras_salvar_em": str(COMPRAS_DIR),
        "sellflux_salvar_em": str(SELLFLUX_DIR),
        "handler_corrigido": str(CODIGOS / "atualizar_tudo_ao_salvar.py"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "xlsx": str(OUT_XLSX), "pdf": str(OUT_PDF) if OUT_PDF.exists() else "", "pdf_ok": pdf_ok, "compras": str(COMPRAS_DIR), "sellflux": str(SELLFLUX_DIR)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
