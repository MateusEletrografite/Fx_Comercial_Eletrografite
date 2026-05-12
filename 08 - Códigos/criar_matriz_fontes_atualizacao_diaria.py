from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
FONTE = ROOT / "04 - Fonte"
REGRAS = ROOT / "09 - Regras"
APOIO = ROOT / "05 - Apoio"
OUT = FONTE / "MATRIZ_FONTES_ATUALIZACAO_DIARIA.xlsx"
SUMMARY = APOIO / "resumo_matriz_fontes_atualizacao_diaria.json"
LISTAGEM = REGRAS / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


GREEN = "1F7A4D"
LIGHT = "E7F5EC"
YELLOW = "FFF2CC"
BLUE = "DDEBF7"
RED = "FCE4D6"


def style_ws(ws, table_name: str, header_count: int | None = None) -> None:
    header_count = header_count or ws.max_column
    for cell in ws[1]:
        cell.value = "" if cell.value is None else str(cell.value)
    fill = PatternFill("solid", fgColor=GREEN)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D5E7DC")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        1: 24,
        2: 26,
        3: 34,
        4: 24,
        5: 20,
        6: 26,
        7: 44,
        8: 38,
        9: 34,
        10: 34,
        11: 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and header_count > 1:
        tab = Table(displayName=table_name, ref=f"A1:{ws.cell(ws.max_row, header_count).coordinate}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws.add_table(tab)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]], table_name: str):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_ws(ws, table_name, len(headers))
    return ws


def update_listagem() -> None:
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    rows = [
        [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            "MATRIZ_FONTES_ATUALIZACAO_DIARIA",
            "Fonte diária",
            "Separar todas as bases necessárias para atualização diária do dashboard.",
            "Usuário informou que o ERP gera relatórios personalizados e também possui API.",
            "Recomendação: iniciar com relatório personalizado controlado e migrar Propostas/Clientes/Produtos/Vendedores para API V3 quando as credenciais OAuth2 estiverem disponíveis.",
        ],
        [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            "MATRIZ_FONTES_ATUALIZACAO_DIARIA",
            "API Olist/Tiny",
            "Registrar que API V3 usa aplicativo com Client ID, Client Secret, permissões por módulo e limites por minuto.",
            "Documentação Olist/Tiny consultada em 11/05/2026.",
            "Não armazenar segredo dentro de XLSX; usar .env local quando automatizar.",
        ],
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_ws(ws, "TabelaLinhaTempoFontesDiarias", len(headers))
    wb.save(LISTAGEM)


def main() -> None:
    FONTE.mkdir(exist_ok=True)
    REGRAS.mkdir(exist_ok=True)
    APOIO.mkdir(exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    fontes_headers = [
        "Base / fonte",
        "Origem atual",
        "Arquivo/pasta atual",
        "Método recomendado",
        "Frequência",
        "Responsável",
        "Campos mínimos",
        "Uso no dashboard",
        "Validação diária",
        "Observações",
        "Prioridade",
    ]
    fontes_rows = [
        [
            "Propostas comerciais",
            "ERP Olist/Tiny",
            "Relatórios personalizados de propostas por mês/status + relatorios_e_dashboard_base_AJUSTE_7_REGRAS.xlsx",
            "Curto prazo: exportação personalizada CSV/XLSX padronizada. Médio prazo: API V3 de Pedidos/Propostas, se endpoint de propostas estiver disponível na conta.",
            "Diária, antes do início comercial",
            "Comercial / automação",
            "Proposta ID, número, data, status, situação, cliente, CPF/CNPJ, vendedor, valor, desconto, frete, observações, itens/produtos, marcadores",
            "KPIs, RX da equipe, propostas comerciais, kits, status, valores e clientes",
            "Contar linhas por status, validar datas do dia/mês, validar IDs únicos e comparar total com relatório do ERP",
            "É a base central. Deve vir completa, não apenas em aberto, para permitir concluídas e histórico.",
            "Alta",
        ],
        [
            "Clientes / contatos",
            "ERP Olist/Tiny",
            "Campos dentro da base de propostas; ideal separar em fonte própria",
            "API V3 de Contatos e Vendedores. Se ainda não houver API, relatório personalizado de clientes com marcadores.",
            "Diária",
            "Comercial / cadastro",
            "Cliente ID, nome, CPF/CNPJ, telefone, celular, email, cidade, UF, endereço, marcadores, vendedor/carteira",
            "Chave de cliente, deduplicação, filtros, RX do cliente, casamento com atendimentos",
            "Validar CPF/CNPJ, telefones normalizados, clientes sem documento e duplicados por telefone",
            "A documentação da API V3 lista Contatos como recurso disponível, incluindo clientes, fornecedores e vendedores.",
            "Alta",
        ],
        [
            "Vendedores e setores",
            "ERP + regra interna",
            "Marcadores agrupados e regra no script; Elenice ajustada para Tráfego",
            "API V3 de Contatos/Vendedores + planilha auxiliar de regras de setor quando houver exceções.",
            "Diária ou quando houver mudança de equipe",
            "Gestor comercial",
            "Nome vendedor, ID vendedor, setor atual, status ativo/inativo, exceções temporárias",
            "RX da equipe, filtros por setor, distribuição de atendimento",
            "Validar vendedor sem setor, vendedor fora da base e exceções manuais",
            "Manter arquivo auxiliar para exceções como 'Elenice = Tráfego por enquanto'.",
            "Alta",
        ],
        [
            "Produtos / categorias / kits",
            "ERP Olist/Tiny + regra de inferência",
            "Produtos agrupados nas propostas; regras em 09 - Regras",
            "API V3 de Produtos para cadastro e relatório de itens/produtos por proposta para movimento comercial.",
            "Diária para itens vendidos/propostos; semanal para cadastro",
            "Comercial / produtos",
            "SKU, nome produto, categoria, família, kit real P/M/G/PP, tipo de agrupamento, marcadores",
            "Kits, categorias, agrupamento de produto, análise por mix",
            "Validar que Expositor/Mostruário não vire kit indevido; validar kits com separador ';'",
            "Regra já corrigida: Expositor e Mostruário não definem kit sozinhos.",
            "Alta",
        ],
        [
            "Atendimentos consolidados",
            "Workspace Atendimentos",
            r"C:\Users\Roberto Moura\Downloads\Atendimentos\relatorio_atendimentos_consolidado_classificado_empresa_fornecedor_20260511-085458.xlsx",
            "Manter pipeline do workspace Atendimentos e consumir o XLSX consolidado mais recente. Se o sistema de atendimento tiver API, integrar depois.",
            "Diária, após fechamento/extração dos atendimentos",
            "Atendimento / automação",
            "Cliente, CNPJ, telefone, quantidade, último contato, protocolo, atendente, classificação, fornecedor/empresa",
            "Fila de atendimento, último contato, atendidos uma vez, 15/30/60/+60 dias, plano de ação",
            "Validar data do último contato, protocolos vazios, telefone/CNPJ normalizados e total por setor",
            "Hoje é fonte externa ao ERP. Deve entrar antes de gerar dashboard final.",
            "Alta",
        ],
        [
            "Compras / histórico de compra",
            "Workspace Atendimentos / ERP",
            r"C:\Users\Roberto Moura\Downloads\Atendimentos\Compras clientes até 20-04.xlsx",
            "Preferir API/relatório ERP de pedidos faturados/compras por cliente. Enquanto isso, atualizar XLSX de compras.",
            "Diária ou semanal, conforme necessidade comercial",
            "Comercial / financeiro",
            "Cliente, CPF/CNPJ, data última compra, valor, produtos, quantidade, situação",
            "Tempo desde última compra, ativação de cliente, clientes sem compra 180+ dias",
            "Validar data máxima, clientes sem chave e total de compras novas",
            "Importante para plano de ativação, mas não deve bloquear atualização de propostas.",
            "Média",
        ],
        [
            "Marcadores de propostas/clientes",
            "ERP Olist/Tiny",
            "Propostas por mes com marcadores/*.csv",
            "Relatório personalizado ou API de Contatos/Propostas com marcadores.",
            "Diária",
            "Comercial",
            "Proposta ID, Cliente ID, marcadores, data, vendedor",
            "Setor, kit, carteira/tráfego, auditorias e filtros",
            "Validar propostas sem marcador e marcadores duplicados por ID",
            "Fonte sensível para separar Tráfego, Carteira e ESM.",
            "Alta",
        ],
        [
            "Retiradas / auditoria",
            "Pipeline interno",
            r"06 - Lote final organizado\apoio\base_retiradas_propostas.csv",
            "Gerado pelo script de consolidação após carregar fontes novas.",
            "Diária, após consolidar propostas",
            "Automação",
            "Proposta retirada, motivo, proposta mantida, status mantido, data, marcador",
            "Auditoria, duplicadas por motivo, explicação do que saiu do dashboard",
            "Validar retiradas com proposta mantida preenchida e contagens por motivo",
            "Não é fonte primária; é saída de auditoria.",
            "Média",
        ],
    ]
    add_sheet(wb, "Fontes diárias", fontes_headers, fontes_rows, "TabelaFontesDiarias")

    fluxo_headers = ["Ordem", "Etapa", "Entrada", "Saída", "Automatizar como", "Observação"]
    fluxo_rows = [
        [1, "Baixar/consultar ERP", "Propostas, clientes, vendedores, produtos, marcadores", "CSVs/XLSX brutos em 04 - Fonte", "API V3 ou exportação personalizada com nome padronizado", "Nunca editar arquivo bruto manualmente."],
        [2, "Atualizar Atendimentos", "Workspace Atendimentos", "XLSX consolidado de atendimentos mais recente", "Script do workspace Atendimentos", "Rodar antes do dashboard de propostas."],
        [3, "Normalizar chaves", "ERP + Atendimentos", "Base padronizada com CPF/CNPJ, telefone e nome normalizado", "Script Python", "Ordem de casamento recomendada: telefone, CPF/CNPJ, nome."],
        [4, "Aplicar regras comerciais", "Base normalizada", "Mantidas, retiradas, auditorias, kits e setores", "Script Python", "Manter exceções em 09 - Regras."],
        [5, "Gerar entregáveis", "Bases finais", "dashboard.html, XLSX final, auditorias", "Script Python único diário", "Atualizar raiz e 06 - Lote final organizado."],
        [6, "Validar", "Entregáveis", "Resumo de contagem e divergências", "Checklist automático", "Conferir totais por status, vendedor, setor e atendimentos."],
    ]
    add_sheet(wb, "Fluxo diário", fluxo_headers, fluxo_rows, "TabelaFluxoDiario")

    api_headers = ["Tema", "Documentação / regra", "Aplicação no projeto", "Status para automatizar"]
    api_rows = [
        ["Pré-requisito", "API V3 exige plano compatível e extensão Gestão de Aplicativos instalada.", "Antes de automatizar, confirmar plano e instalar/ativar Aplicativos.", "Pendente credenciais"],
        ["Credenciais", "Criar aplicativo no ERP e obter Client ID e Client Secret.", "Salvar em .env local, nunca dentro do XLSX.", "Pendente credenciais"],
        ["Permissões", "Permissões são definidas por módulo: leitura, incluir/editar e excluir.", "Para dashboard, começar com leitura em Contatos, Produtos, Pedidos/Propostas e Informações da Conta.", "Pendente configuração"],
        ["Autenticação", "API V3 usa OAuth2 com client_id e client_secret.", "Criar rotina de token/refresh token antes da coleta diária.", "Pendente desenvolvimento"],
        ["Limite", "Limites por minuto variam por plano e são compartilhados por conta.", "Implementar pausa/retry e registrar consumo.", "Pendente desenvolvimento"],
        ["Módulos úteis", "Contatos/Vendedores, Produtos, Pedidos, Estoque, Listas de Preço, Informações da Conta.", "Prioridade: Contatos, Vendedores, Pedidos/Propostas, Produtos e Marcadores.", "Mapeado"],
    ]
    add_sheet(wb, "API Olist Tiny", api_headers, api_rows, "TabelaApiOlistTiny")

    checklist_headers = ["Check diário", "Como validar", "OK?", "Observação"]
    checklist_rows = [
        ["Arquivo bruto de propostas do dia/mês está em 04 - Fonte", "Conferir data de modificação e quantidade de linhas", "", ""],
        ["Atendimentos consolidados atualizado", "Conferir arquivo mais recente do workspace Atendimentos", "", ""],
        ["Clientes com chave válida", "CPF/CNPJ ou telefone preenchido para casamento", "", ""],
        ["Vendedores sem setor", "Listar e decidir regra antes de publicar dashboard", "", ""],
        ["Kits sem duplicidade indevida", "Validar P/M/G/PP e 'Sem tag de kit'", "", ""],
        ["Totais por status conferem com ERP", "Em aberto, concluídas, não aprovadas", "", ""],
        ["Dashboard copiado para raiz", "dashboard.html atualizado", "", ""],
        ["Listagem de regras atualizada", "09 - Regras com linha do tempo", "", ""],
    ]
    add_sheet(wb, "Checklist diário", checklist_headers, checklist_rows, "TabelaChecklistDiario")

    wb.save(OUT)

    update_listagem()

    SUMMARY.write_text(
        json.dumps(
            {
                "arquivo": str(OUT),
                "abas": ["Fontes diárias", "Fluxo diário", "API Olist Tiny", "Checklist diário"],
                "gerado_em": datetime.now().isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"ok": True, "out": str(OUT), "listagem": str(LISTAGEM)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
