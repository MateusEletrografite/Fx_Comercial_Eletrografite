from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = ROOT.parent
ATEND = DOWNLOADS / "Atendimentos"
FONTE = ROOT / "04 - Fonte"
REGRAS = ROOT / "09 - Regras"
APOIO = ROOT / "05 - Apoio"
INVENTARIO = FONTE / "INVENTARIO_FONTES_ATUALIZACAO_DIARIA.xlsx"
SUMMARY = APOIO / "resumo_organizacao_fontes_atualizacao_diaria.json"
LISTAGEM = REGRAS / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


GROUPS = {
    "01 - ERP Propostas": [
        ROOT / "relatorios_e_dashboard_base_AJUSTE_7_REGRAS.xlsx",
        ROOT / "relatorios_e_dashboard_base_AJUSTE_7_REGRAS_COM_ATENDIMENTOS_20260510-215511.xlsx",
        ROOT / "Analises das propostas em aberto.xlsx",
        ROOT / "Analises propostas 2.xlsx",
        ROOT / "Analise 2 proposta.xlsx",
    ],
    "02 - Atendimentos": [
        ATEND / "relatorio_atendimentos_consolidado_classificado_empresa_fornecedor_20260511-085458.xlsx",
        ATEND / "relatorio_atendimentos_com_propostas_abertas_20260510-212628.xlsx",
        ATEND / "Analise de quantidade de atendimentos.xlsx",
        ATEND / "Atendimentos Sfx após 01-01-2026.xlsx",
        ATEND / "Classificação empresa - fornecedor.xlsx",
    ],
    "03 - Compras": [
        ROOT / "Compras clientes até 20-04.xlsx",
        ATEND / "Compras clientes até 20-04.xlsx",
    ],
    "04 - Marcadores e setores": [
        ROOT / "Regras.xlsx",
        ROOT / "Cnpjs Cenprot.xlsx",
        ROOT / "Ajustes 7.xlsx",
    ],
    "05 - Apoio layout e configuração": [
        ROOT / "Padrão Codex.xlsx",
        ROOT / "Layout e alterações.xlsx",
    ],
}

COPY_DIRS = {
    "01 - ERP Propostas/CSVs por status - Em aberto": ROOT / "01 - Em aberto - 12-2024 a 05-2026",
    "01 - ERP Propostas/CSVs por status - Concluidas": ROOT / "02 - Concluídas - 12-2024 a 05-2026",
    "01 - ERP Propostas/CSVs por status - Nao aprovadas": ROOT / "03 - Não aprovadas - 12-2024 a 05-2026",
    "04 - Marcadores e setores/Propostas por mes com marcadores": ROOT / "Propostas por mes com marcadores",
}


def safe_copy(src: Path, dest_dir: Path) -> dict[str, object]:
    dest_dir.mkdir(parents=True, exist_ok=True)
    if not src.exists() or src.name.startswith("~$"):
        return {
            "grupo": dest_dir.name,
            "arquivo": src.name,
            "origem": str(src),
            "destino": "",
            "status": "não encontrado" if not src.exists() else "temporário ignorado",
            "tamanho": "",
            "modificado": "",
        }
    dest = dest_dir / src.name
    if src.resolve() != dest.resolve():
        shutil.copy2(src, dest)
    return {
        "grupo": dest_dir.relative_to(FONTE).as_posix(),
        "arquivo": src.name,
        "origem": str(src),
        "destino": str(dest),
        "status": "copiado",
        "tamanho": src.stat().st_size,
        "modificado": datetime.fromtimestamp(src.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
    }


def copy_tree_files(src_dir: Path, dest_dir: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not src_dir.exists():
        rows.append({
            "grupo": dest_dir.relative_to(FONTE).as_posix(),
            "arquivo": src_dir.name,
            "origem": str(src_dir),
            "destino": "",
            "status": "pasta não encontrada",
            "tamanho": "",
            "modificado": "",
        })
        return rows
    for src in src_dir.rglob("*"):
        if not src.is_file() or src.name.startswith("~$"):
            continue
        if src.suffix.lower() not in {".xlsx", ".xls", ".csv", ".txt", ".json"}:
            continue
        rel = src.relative_to(src_dir)
        rows.append(safe_copy(src, dest_dir / rel.parent))
    return rows


def style_sheet(ws, table_name: str, header_count: int | None = None) -> None:
    header_count = header_count or ws.max_column
    for cell in ws[1]:
        cell.value = "" if cell.value is None else str(cell.value)
    green = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.value = "" if cell.value is None else str(cell.value)
    for cell in ws[1]:
        cell.fill = green
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [28, 42, 70, 70, 18, 16, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    ws.freeze_panes = "A2"
    tab = Table(displayName=table_name, ref=f"A1:{ws.cell(ws.max_row, header_count).coordinate}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)


def write_inventory(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário fontes"
    headers = ["Grupo", "Arquivo", "Origem", "Destino", "Status", "Tamanho bytes", "Modificado origem"]
    ws.append(headers)
    for item in rows:
        ws.append([item["grupo"], item["arquivo"], item["origem"], item["destino"], item["status"], item["tamanho"], item["modificado"]])
    style_sheet(ws, "TabelaInventarioFontes", len(headers))

    ws2 = wb.create_sheet("Orientação diária")
    ws2.append(["Ordem", "Ação", "Pasta", "Observação"])
    steps = [
        [1, "Substituir/baixar relatórios de propostas do ERP", "04 - Fonte/01 - ERP Propostas", "Manter nomes padronizados ou colocar novo arquivo bruto com data."],
        [2, "Atualizar atendimentos consolidados", "04 - Fonte/02 - Atendimentos", "Copiar o consolidado mais recente do workspace Atendimentos."],
        [3, "Atualizar compras", "04 - Fonte/03 - Compras", "Usar relatório ERP ou planilha de compras mais recente."],
        [4, "Atualizar marcadores/regras", "04 - Fonte/04 - Marcadores e setores", "Inclui marcadores, CNPJs auxiliares, regras de setor e exceções."],
        [5, "Rodar script diário de consolidação", "08 - Códigos", "Depois de alimentar as fontes, gerar dashboard e auditorias."],
    ]
    for step in steps:
        ws2.append(step)
    style_sheet(ws2, "TabelaOrientacaoFontes", 4)
    wb.save(INVENTARIO)


def update_listagem(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    ws.append([
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        "FONTES_ATUALIZACAO_DIARIA",
        "Organização de fontes",
        "Copiar para 04 - Fonte as planilhas e CSVs que precisam ser alimentados diariamente.",
        "Pedido do usuário para centralizar as bases de atualização diária.",
        f"{sum(1 for r in rows if r['status'] == 'copiado')} arquivos copiados; temporários ~$ ignorados.",
    ])
    green = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.fill = green
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in range(1, 7):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 30 if col != 4 else 48
    tab = Table(displayName="TabelaLinhaTempoFontesCopiadas", ref=f"A1:F{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)
    wb.save(LISTAGEM)


def main() -> None:
    FONTE.mkdir(exist_ok=True)
    APOIO.mkdir(exist_ok=True)
    REGRAS.mkdir(exist_ok=True)

    rows: list[dict[str, object]] = []
    for group, files in GROUPS.items():
        dest_dir = FONTE / group
        for src in files:
            rows.append(safe_copy(src, dest_dir))
    for rel_dest, src_dir in COPY_DIRS.items():
        rows.extend(copy_tree_files(src_dir, FONTE / rel_dest))

    write_inventory(rows)
    update_listagem(rows)

    SUMMARY.write_text(json.dumps({
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "pasta_fonte": str(FONTE),
        "inventario": str(INVENTARIO),
        "copiados": sum(1 for r in rows if r["status"] == "copiado"),
        "nao_encontrados": [r for r in rows if r["status"] != "copiado"],
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "fonte": str(FONTE), "inventario": str(INVENTARIO), "copiados": sum(1 for r in rows if r["status"] == "copiado")}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
