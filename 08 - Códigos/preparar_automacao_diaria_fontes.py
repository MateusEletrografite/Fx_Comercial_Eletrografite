from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = ROOT.parent
BANCO = DOWNLOADS / "Banco de Dados Operacional"
ENTRADA = BANCO / "00 - Entrada diaria"
APRESENTACAO = ROOT / "Apresentação"
FONTE = ROOT / "04 - Fonte"
APOIO = ROOT / "05 - Apoio"
REGRAS = ROOT / "09 - Regras"
CODIGOS = ROOT / "08 - Códigos"
GUIA = APRESENTACAO / "08_GUIA_RELATORIOS_E_AUTOMACAO_DIARIA.xlsx"
SUMMARY = APOIO / "resumo_automacao_diaria_fontes.json"
LISTAGEM = REGRAS / "LISTAGEM_REGRAS_EXCECOES_CONTEXTO_LINHA_DO_TEMPO.xlsx"


PROPOSTAS_COLS = [
    "ID", "Número da proposta", "Data", "Data proximo contato", "ID contato", "Nome do contato",
    "Aos cuidados de", "Lista de Preço", "Tipo de Pessoa", "CPF/CNPJ", "RG/IE", "CEP",
    "Município", "UF", "Endereço", "Endereço Nro", "Complemento", "Bairro", "Fone",
    "Celular", "E-mail", "Desconto", "Frete", "Observações", "Validade", "Prazo de Entrega",
    "Situação", "Introdução", "ID produto", "Descrição", "Quantidade", "Valor unitário",
    "Descrição complementar", "Vendedor", "Destinatário", "CPF/CNPJ entrega", "CEP entrega",
    "Município entrega", "UF entrega", "Endereço entrega", "Endereço Nro entrega",
    "Complemento entrega", "Bairro entrega", "Fone entrega", "Inscrição Estadual entrega",
]
MARCADOR_EXTRA_COLS = ["Marcadores", "Razao social marcador", "CNPJ marcador", "Arquivo marcador"]
ATENDIMENTOS_COLS = [
    "Classificação consolidação", "Quantidade de atendimentos", "Chave de consolidação",
    "Protocolo de atendimentos", "Plataforma de atendimento", "Abrangência plataforma",
    "Telefone completo em número", "Nome na plataforma de atendimento", "CNPJ na plataforma de atendimento",
    "Data do contato", "Horário do contato", "Número do último pedido", "Data do último pedido",
    "Horário do último pedido", "Valor do último pedido", "Vendedor do ultimo Pedido",
    "Data do último contato", "Horário do último contato", "Data do ultimo encerramento",
    "Horário do ultimo encerramento", "Classificação do Contato", "Vendedor classificado que irá atender",
    "Tag1 - Contato", "Tag2 - Vendedor", "Tag3 - Atividade na carteira",
]
COMPRAS_COLS = [
    "Data de inclusão", "Pedido", "Duplicidade (Número de pedido)", "CNPJ",
    "Cnpj sem pontos para Cruzamento", "Razão Social", "Valor do pedido", "Data da compra",
    "Vendedor da Compra", "Telefone Tiny", "Telefone Biz", "Whatsapp Mande um Zap", "Whatsapp Biz",
    "Motivo da duplicidade", "Indicadores da duplicidade", "Tipo de duplicidade", "Ultimo Pedido",
    "Valor do último pedido", "Data última compra", "Vendedor do ultimo pedido",
]
VENDEDORES_COLS = ["Nome do Vendedor", "ID vendedor", "Setor", "Ativo?", "Observação / exceção"]
PRODUTOS_COLS = ["ID produto", "SKU", "Descrição", "Categoria do produto", "Tipo de agrupamento de produto", "Kit", "Ativo?"]


def ensure_dirs() -> None:
    folders = [
        ENTRADA / "01 - ERP Propostas" / "Em aberto",
        ENTRADA / "01 - ERP Propostas" / "Concluidas",
        ENTRADA / "01 - ERP Propostas" / "Nao aprovadas",
        ENTRADA / "01 - ERP Propostas" / "Marcadores",
        ENTRADA / "01 - ERP Propostas" / "Propostas comerciais xlsx",
        ENTRADA / "02 - Atendimentos",
        ENTRADA / "03 - Compras",
        ENTRADA / "04 - Cadastros ERP" / "Clientes",
        ENTRADA / "04 - Cadastros ERP" / "Vendedores setores",
        ENTRADA / "04 - Cadastros ERP" / "Produtos kits",
        ENTRADA / "99 - Processados",
        ENTRADA / "99 - Logs",
        APRESENTACAO,
        FONTE / "01 - ERP Propostas",
        BANCO / "01 - ERP Olist" / "Propostas",
        APOIO,
        REGRAS,
    ]
    for folder in folders:
        folder.mkdir(parents=True, exist_ok=True)


def copy_latest_propostas_comerciais() -> list[dict[str, str]]:
    src = APRESENTACAO / "Propostas comerciais.xlsx"
    targets = [
        APRESENTACAO / "05_BASE_PROPOSTAS_COM_ATENDIMENTOS_ATUAL.xlsx",
        FONTE / "01 - ERP Propostas" / "Propostas comerciais.xlsx",
        BANCO / "01 - ERP Olist" / "Propostas" / "Propostas comerciais.xlsx",
        BANCO / "04 - Saídas consolidadas" / "Propostas do sistema" / "Propostas comerciais.xlsx",
    ]
    results = []
    for target in targets:
        target.parent.mkdir(parents=True, exist_ok=True)
        if src.exists():
            shutil.copy2(src, target)
            status = "substituído"
        else:
            status = "origem não encontrada"
        results.append({"origem": str(src), "destino": str(target), "status": status})
    return results


def style(ws, table_name: str, header_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F7A4D")
    for cell in ws[1]:
        cell.value = "" if cell.value is None else str(cell.value)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [28, 42, 24, 24, 60, 70]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width
    ws.freeze_panes = "A2"
    tab = Table(displayName=table_name, ref=f"A1:{ws.cell(ws.max_row, header_count).coordinate}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(tab)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]], table_name: str) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style(ws, table_name, len(headers))


def write_guide() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    reports = [
        ["Propostas em aberto", "CSV", "ERP Olist/Tiny - relatório personalizado de propostas", str(ENTRADA / "01 - ERP Propostas" / "Em aberto"), "Diário", "Use o CSV padrão do ERP. Não alterar cabeçalhos."],
        ["Propostas concluídas", "CSV", "ERP Olist/Tiny - relatório personalizado de propostas", str(ENTRADA / "01 - ERP Propostas" / "Concluidas"), "Diário", "Use o CSV padrão do ERP. Não alterar cabeçalhos."],
        ["Propostas não aprovadas", "CSV", "ERP Olist/Tiny - relatório personalizado de propostas", str(ENTRADA / "01 - ERP Propostas" / "Nao aprovadas"), "Diário", "Use o CSV padrão do ERP. Não alterar cabeçalhos."],
        ["Propostas com marcadores", "CSV", "ERP Olist/Tiny - propostas + marcadores", str(ENTRADA / "01 - ERP Propostas" / "Marcadores"), "Diário", "Mesmo CSV padrão de propostas, com 4 colunas extras de marcador."],
        ["Propostas comerciais consolidado", "XLSX", "Arquivo final mais recente exportado/gerado", str(ENTRADA / "01 - ERP Propostas" / "Propostas comerciais xlsx"), "Quando gerar novo", "Se você gerar esse XLSX, ele vira base prioritária."],
        ["Atendimentos consolidados", "XLSX ou CSV", "Workspace Atendimentos", str(ENTRADA / "02 - Atendimentos"), "Diário", "Usar o consolidado de atendimentos mais recente."],
        ["Compras / histórico de compra", "XLSX ou CSV", "ERP ou workspace Atendimentos", str(ENTRADA / "03 - Compras"), "Diário ou semanal", "Necessário para última compra e ativação."],
        ["Clientes", "CSV", "ERP Olist/Tiny - cadastro de contatos", str(ENTRADA / "04 - Cadastros ERP" / "Clientes"), "Diário", "Pode ser montado via API depois."],
        ["Vendedores e setores", "CSV ou XLSX", "ERP + regra interna", str(ENTRADA / "04 - Cadastros ERP" / "Vendedores setores"), "Quando mudar equipe", "Inclui exceções, exemplo: Elenice no Tráfego."],
        ["Produtos, categorias e kits", "CSV ou XLSX", "ERP Olist/Tiny - produtos", str(ENTRADA / "04 - Cadastros ERP" / "Produtos kits"), "Semanal ou quando mudar mix", "Usado para categorias, kits e agrupamentos."],
    ]
    add_sheet(wb, "Onde salvar", ["Relatório", "Formato", "Origem", "Salvar em", "Frequência", "Observação"], reports, "TabelaOndeSalvar")

    cols_rows = []
    for c in PROPOSTAS_COLS:
        cols_rows.append(["Propostas CSV padrão", c, "Obrigatória", "Já vem no CSV do ERP", ""])
    for c in MARCADOR_EXTRA_COLS:
        cols_rows.append(["Propostas com marcadores", c, "Obrigatória nesse relatório", "Já vem no CSV com marcadores", ""])
    for c in ATENDIMENTOS_COLS:
        cols_rows.append(["Atendimentos consolidados", c, "Obrigatória/recomendada", "Montar se exportar manualmente", ""])
    for c in COMPRAS_COLS:
        cols_rows.append(["Compras", c, "Obrigatória/recomendada", "Montar se exportar manualmente", ""])
    for c in VENDEDORES_COLS:
        cols_rows.append(["Vendedores e setores", c, "Obrigatória", "Montar manual/API", ""])
    for c in PRODUTOS_COLS:
        cols_rows.append(["Produtos categorias kits", c, "Obrigatória/recomendada", "Montar manual/API", ""])
    add_sheet(wb, "Colunas necessárias", ["Relatório", "Coluna", "Obrigatoriedade", "Como obter", "Observação"], cols_rows, "TabelaColunasNecessarias")

    auto_rows = [
        [1, "Salvar um arquivo novo em uma subpasta de 00 - Entrada diaria", "FileSystemWatcher detecta arquivo criado/alterado", "Aguardar arquivo estabilizar", "Evita processar arquivo ainda sendo gravado."],
        [2, "Classificar pela pasta onde foi salvo", "atualizar_tudo_ao_salvar.py", "Copiar para Banco/Fonte canônico", "Mantém histórico no banco e cópia usada pelo workspace."],
        [3, "Registrar log", "99 - Logs", "JSONL com data, arquivo e ação", "Ajuda auditoria."],
        [4, "Rodar atualização do projeto", "scripts do workspace", "Regerar inventário/apresentação/dashboards", "A etapa de dashboard pode ser expandida com a API quando as credenciais chegarem."],
    ]
    add_sheet(wb, "Automação ao salvar", ["Ordem", "Evento", "Script", "Ação", "Observação"], auto_rows, "TabelaAutomacaoSalvar")

    wb.save(GUIA)


WATCHER = r"""param(
  [string]$WatchPath = "C:\Users\Roberto Moura\Downloads\Banco de Dados Operacional\00 - Entrada diaria"
)

$ErrorActionPreference = "Stop"
$script = "C:\Users\Roberto Moura\Downloads\Propostas do sistema\08 - Códigos\atualizar_tudo_ao_salvar.py"

Write-Host "Monitorando: $WatchPath"
Write-Host "Quando salvar arquivo novo, a atualização será disparada."

$fsw = New-Object System.IO.FileSystemWatcher
$fsw.Path = $WatchPath
$fsw.IncludeSubdirectories = $true
$fsw.EnableRaisingEvents = $true
$fsw.Filter = "*.*"

$action = {
  $path = $Event.SourceEventArgs.FullPath
  $name = $Event.SourceEventArgs.Name
  if ($name -like "~$*") { return }
  if ($path -match "\\99 - Logs\\" -or $path -match "\\99 - Processados\\") { return }
  $ext = [System.IO.Path]::GetExtension($path).ToLowerInvariant()
  if ($ext -notin @(".csv",".xlsx",".xls")) { return }
  Write-Host "Detectado: $path"
  Start-Process -FilePath "python" -ArgumentList @($script, "--arquivo", $path) -NoNewWindow
}

Register-ObjectEvent $fsw Created -Action $action | Out-Null
Register-ObjectEvent $fsw Changed -Action $action | Out-Null

while ($true) { Start-Sleep -Seconds 2 }
"""


HANDLER = r'''from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import time
from datetime import datetime
from pathlib import Path


ROOT = Path(r"C:\Users\Roberto Moura\Downloads\Propostas do sistema")
BANCO = Path(r"C:\Users\Roberto Moura\Downloads\Banco de Dados Operacional")
ENTRADA = BANCO / "00 - Entrada diaria"
FONTE = ROOT / "04 - Fonte"
LOG = ENTRADA / "99 - Logs" / "atualizacoes_ao_salvar.jsonl"


def wait_stable(path: Path, attempts: int = 10, delay: float = 1.0) -> bool:
    last = -1
    for _ in range(attempts):
        if not path.exists():
            time.sleep(delay)
            continue
        size = path.stat().st_size
        if size == last and size > 0:
            return True
        last = size
        time.sleep(delay)
    return path.exists() and path.stat().st_size > 0


def classify(path: Path) -> tuple[Path, Path]:
    rel = path.relative_to(ENTRADA)
    parts = rel.parts
    if "Em aberto" in parts:
        return BANCO / "01 - ERP Olist" / "Propostas" / "Em aberto" / path.name, FONTE / "01 - ERP Propostas" / "CSVs por status - Em aberto" / path.name
    if "Concluidas" in parts:
        return BANCO / "01 - ERP Olist" / "Propostas" / "Concluidas" / path.name, FONTE / "01 - ERP Propostas" / "CSVs por status - Concluidas" / path.name
    if "Nao aprovadas" in parts:
        return BANCO / "01 - ERP Olist" / "Propostas" / "Nao aprovadas" / path.name, FONTE / "01 - ERP Propostas" / "CSVs por status - Nao aprovadas" / path.name
    if "Marcadores" in parts:
        return BANCO / "01 - ERP Olist" / "Marcadores" / path.name, FONTE / "04 - Marcadores e setores" / "Propostas por mes com marcadores" / path.name
    if "Propostas comerciais xlsx" in parts:
        return BANCO / "01 - ERP Olist" / "Propostas" / "Propostas comerciais.xlsx", FONTE / "01 - ERP Propostas" / "Propostas comerciais.xlsx"
    if "02 - Atendimentos" in parts:
        return BANCO / "02 - Atendimentos" / "Consolidados" / path.name, FONTE / "02 - Atendimentos" / path.name
    if "03 - Compras" in parts:
        return BANCO / "02 - Atendimentos" / "Compras" / path.name, FONTE / "03 - Compras" / path.name
    if "Vendedores setores" in parts:
        return BANCO / "03 - Regras internas" / "Vendedores setores" / path.name, FONTE / "04 - Marcadores e setores" / path.name
    if "Produtos kits" in parts:
        return BANCO / "01 - ERP Olist" / "Produtos" / path.name, FONTE / "04 - Marcadores e setores" / path.name
    if "Clientes" in parts:
        return BANCO / "01 - ERP Olist" / "Clientes e auxiliares" / path.name, FONTE / "01 - ERP Propostas" / path.name
    return BANCO / "99 - A classificar" / path.name, FONTE / "99 - A classificar" / path.name


def copy_to(src: Path, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)


def log_event(payload: dict) -> None:
    LOG.parent.mkdir(parents=True, exist_ok=True)
    with LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def run_project_scripts() -> list[dict]:
    scripts = [
        ROOT / "08 - Códigos" / "organizar_fontes_para_atualizacao_diaria.py",
        ROOT / "08 - Códigos" / "criar_apresentacao_e_banco_operacional.py",
    ]
    results = []
    for script in scripts:
        if not script.exists():
            results.append({"script": str(script), "status": "não encontrado"})
            continue
        proc = subprocess.run(["python", str(script)], cwd=str(ROOT), capture_output=True, text=True, timeout=300)
        results.append({"script": str(script), "returncode": proc.returncode, "stdout": proc.stdout[-1000:], "stderr": proc.stderr[-1000:]})
    return results


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--arquivo", required=True)
    args = parser.parse_args()
    src = Path(args.arquivo)
    if src.name.startswith("~$"):
        return
    ok = wait_stable(src)
    banco_dest, fonte_dest = classify(src)
    if ok:
        copy_to(src, banco_dest)
        copy_to(src, fonte_dest)
        scripts = run_project_scripts()
    else:
        scripts = []
    log_event({
        "quando": datetime.now().isoformat(timespec="seconds"),
        "arquivo": str(src),
        "estavel": ok,
        "banco_destino": str(banco_dest),
        "fonte_destino": str(fonte_dest),
        "scripts": scripts,
    })
    print(json.dumps({"ok": ok, "banco": str(banco_dest), "fonte": str(fonte_dest)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
'''


def write_scripts() -> None:
    (CODIGOS / "monitorar_entrada_diaria.ps1").write_text(WATCHER, encoding="utf-8")
    (CODIGOS / "atualizar_tudo_ao_salvar.py").write_text(HANDLER, encoding="utf-8")


def update_listagem(copy_results: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linha do tempo"
    headers = ["Data hora", "Versão", "Tipo", "Regra / ação", "Contexto", "Exceções / observações"]
    ws.append(headers)
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.append([now, "AUTOMACAO_DIARIA_FONTES", "Substituição de base", "Usar Propostas comerciais.xlsx como versão mais recente de propostas.", "Usuário informou que este arquivo é o mais atual.", f"{sum(1 for r in copy_results if r['status'] == 'substituído')} destinos substituídos."])
    ws.append([now, "AUTOMACAO_DIARIA_FONTES", "Automação", "Criar pastas de entrada diária, guia de relatórios e watcher para atualizar ao salvar arquivo novo.", "Usuário exporta CSVs por enquanto e quer programar atualização automática.", "Scripts: monitorar_entrada_diaria.ps1 e atualizar_tudo_ao_salvar.py."])
    style(ws, "TabelaLinhaTempoAutomacaoFontes", len(headers))
    wb.save(LISTAGEM)


def main() -> None:
    ensure_dirs()
    copy_results = copy_latest_propostas_comerciais()
    write_guide()
    write_scripts()
    update_listagem(copy_results)
    SUMMARY.write_text(json.dumps({
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "guia": str(GUIA),
        "entrada_diaria": str(ENTRADA),
        "watcher": str(CODIGOS / "monitorar_entrada_diaria.ps1"),
        "handler": str(CODIGOS / "atualizar_tudo_ao_salvar.py"),
        "substituicoes": copy_results,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "guia": str(GUIA), "entrada": str(ENTRADA), "watcher": str(CODIGOS / "monitorar_entrada_diaria.ps1"), "handler": str(CODIGOS / "atualizar_tudo_ao_salvar.py")}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
